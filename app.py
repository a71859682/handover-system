from __future__ import annotations

import os
import sqlite3
import uuid
from collections.abc import Mapping
from datetime import datetime, timedelta
from functools import wraps
from pathlib import Path
from urllib.parse import urlsplit

import psycopg
from flask import (
    Flask,
    flash,
    g,
    has_app_context,
    has_request_context,
    jsonify,
    redirect,
    render_template,
    request,
    session,
    url_for,
)
from openpyxl import load_workbook
from werkzeug.security import check_password_hash, generate_password_hash
from database import init_database
from services.ai_read_intent_registry import get_supported_ai_read_intent
from sqlite_db_path import get_sqlite_db_path


BASE_DIR = Path(__file__).resolve().parent
DB_PATH = get_sqlite_db_path()
SOURCE_XLSX = BASE_DIR / "source.xlsx"
MAX_WORK_COL = 60  # D:BH
DONE_VALUE = "O"
WORKING_VALUE = "X"
BUILTIN_EXTRA_FIELDS = {
    "initial_check": {"name": "初驗時間", "type": "date", "sort_order": 1},
    "recheck_1": {"name": "複驗1", "type": "date", "sort_order": 2},
    "recheck_2": {"name": "複驗2", "type": "date", "sort_order": 3},
    "handover": {"name": "已交屋", "type": "status", "sort_order": 4},
}
EXTRA_FIELDS = tuple(BUILTIN_EXTRA_FIELDS)
EXTRA_FIELD_TYPES = ("date", "status")

DEFAULT_SETTINGS = {
    "site_title": "大英營造-新埔段",
    "sheet_title": "內裝管制表",
    "tab_title": "內裝管制表(室內)",
    "instruction_text": "子項目輸入 O 代表完成；X 代表施作中或未開始。",
    "floor_header": "樓層",
    "count_header": "戶數",
    "unit_header": "棟別/戶別",
    "vendor_header": "廠商",
    "task_header": "工項",
}
DEFAULT_SITE_NAME = "大英營造-新埔段"
DEFAULT_SITE_CODE = ""
DUAL_WRITE_FLOORS_STRATEGY_LOG = (
    "DUAL_WRITE_FLOORS_SECONDARY table=floors strategy=reuse_primary_postgres_connection"
)
DUAL_WRITE_USERS_STRATEGY_LOG = (
    "DUAL_WRITE_USERS_SECONDARY table=users strategy=reuse_primary_postgres_connection"
)


app = Flask(__name__)
app.config["SECRET_KEY"] = os.environ.get("APP_SECRET_KEY", "dev-secret-change-me")
app.config.setdefault("SQLALCHEMY_DATABASE_URI", f"sqlite:///{DB_PATH.resolve().as_posix()}")
app.config.setdefault("SQLALCHEMY_TRACK_MODIFICATIONS", False)
init_database(app)
ASSET_VERSION = "20260627-010"
_USERS_READ_COMPARE_ORM_READY = True
CREW_BUSINESS_DAY_RESET_HOUR = 8
CREW_BUSINESS_DAY_RESET_MINUTE = 30


@app.context_processor
def inject_asset_version():
    return {"asset_version": ASSET_VERSION}


def db() -> sqlite3.Connection:
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def table_columns(conn: sqlite3.Connection, table: str) -> set[str]:
    return {row["name"] for row in conn.execute(f"PRAGMA table_info({table})")}


def env_flag(name: str) -> bool:
    return os.environ.get(name, "").strip().lower() in {"1", "true", "yes", "on"}


def dual_write_tables() -> set[str]:
    raw = os.environ.get("DUAL_WRITE_TABLES", "")
    return {item.strip() for item in raw.split(",") if item.strip()}


def controlled_dual_write_enabled(*, table: str, operation: str) -> bool:
    if not env_flag("DUAL_WRITE_ENABLED"):
        return False
    if env_flag("USE_SQLALCHEMY_WRITES"):
        return False
    if operation == "update":
        return table in dual_write_tables()
    if operation in {"create", "delete"} and table == "users":
        return table in dual_write_tables()
    return False


def dual_write_dry_run_enabled() -> bool:
    return env_flag("DUAL_WRITE_DRY_RUN")


def dual_write_strict_enabled() -> bool:
    return env_flag("DUAL_WRITE_STRICT")


def dual_write_log(message: str) -> None:
    print(message, flush=True)


def users_read_compare_enabled() -> bool:
    return env_flag("USERS_READ_COMPARE")


def _ensure_users_read_compare_orm_ready() -> None:
    if not _USERS_READ_COMPARE_ORM_READY:
        raise RuntimeError("users_read_compare_orm_not_initialized")


def _run_with_app_context(fn):
    if has_app_context():
        return fn()
    with app.app_context():
        return fn()


class UsersReadCompareError(RuntimeError):
    def __init__(self, *, stage: str, exc: Exception):
        super().__init__(f"users_read_compare_{stage}")
        self.stage = stage
        self.exc_class = type(exc).__name__


def _user_row_payload(
    row: sqlite3.Row | object | None,
    *,
    include_password_hash: bool,
) -> dict[str, object] | None:
    if row is None:
        return None

    is_mapping = isinstance(row, sqlite3.Row) or isinstance(row, dict)
    payload = {
        "id": row["id"] if is_mapping else row.id,
        "username": row["username"] if is_mapping else row.username,
        "display_name": row["display_name"] if is_mapping else row.display_name,
        "role": row["role"] if is_mapping else row.role,
        "created_at": row["created_at"] if is_mapping else row.created_at,
    }
    if include_password_hash:
        payload["password_hash"] = row["password_hash"] if is_mapping else row.password_hash
    return payload


def _list_user_row_payload(row: sqlite3.Row | object) -> dict[str, object]:
    is_mapping = isinstance(row, sqlite3.Row) or isinstance(row, dict)
    return {
        "id": row["id"] if is_mapping else row.id,
        "username": row["username"] if is_mapping else row.username,
        "display_name": row["display_name"] if is_mapping else row.display_name,
        "role": row["role"] if is_mapping else row.role,
        "created_at": row["created_at"] if is_mapping else row.created_at,
    }


def _sqlite_get_user_by_username(username: str) -> sqlite3.Row | None:
    with db() as conn:
        return conn.execute(
            """
            SELECT id, username, display_name, password_hash, role, created_at
            FROM users
            WHERE username = ?
            """,
            (username,),
        ).fetchone()


def _sqlite_get_user_by_id(user_id: int) -> sqlite3.Row | None:
    with db() as conn:
        return conn.execute(
            """
            SELECT id, username, display_name, password_hash, role, created_at
            FROM users
            WHERE id = ?
            """,
            (user_id,),
        ).fetchone()


def _sqlite_get_vendor_account_by_username(username: str) -> sqlite3.Row | None:
    with db() as conn:
        return conn.execute(
            """
            SELECT id, username, password_hash, vendor_name, is_active, created_at, updated_at
            FROM vendor_accounts
            WHERE username = ?
            """,
            (username,),
        ).fetchone()


def _sqlite_list_users() -> list[sqlite3.Row]:
    with db() as conn:
        return conn.execute(
            """
            SELECT id, username, display_name, role, created_at
            FROM users
            ORDER BY id
            """
        ).fetchall()


def _shadow_get_user_by_username(username: str):
    from services.users_orm_service import get_user_by_username_orm

    try:
        _ensure_users_read_compare_orm_ready()
    except Exception as exc:
        raise UsersReadCompareError(stage="orm_init", exc=exc) from exc
    try:
        return _run_with_app_context(lambda: get_user_by_username_orm(username))
    except Exception as exc:
        raise UsersReadCompareError(stage="orm_read", exc=exc) from exc


def _shadow_get_user_by_id(user_id: int):
    from services.users_orm_service import get_user_by_id_orm

    try:
        _ensure_users_read_compare_orm_ready()
    except Exception as exc:
        raise UsersReadCompareError(stage="orm_init", exc=exc) from exc
    try:
        return _run_with_app_context(lambda: get_user_by_id_orm(user_id))
    except Exception as exc:
        raise UsersReadCompareError(stage="orm_read", exc=exc) from exc


def _shadow_list_users():
    from services.users_orm_service import list_users_orm

    try:
        _ensure_users_read_compare_orm_ready()
    except Exception as exc:
        raise UsersReadCompareError(stage="orm_init", exc=exc) from exc
    try:
        return _run_with_app_context(list_users_orm)
    except Exception as exc:
        raise UsersReadCompareError(stage="orm_read", exc=exc) from exc


def _log_users_read_compare_match(*, helper: str, key: str | None = None) -> None:
    message = f"USERS_READ_COMPARE helper={helper}"
    if key:
        message += f" key={key}"
    message += " status=match"
    dual_write_log(message)


def _log_users_read_compare_mismatch(
    *,
    helper: str,
    fields: list[str],
    key: str | None = None,
    error_class: str | None = None,
    error_stage: str | None = None,
) -> None:
    message = f"USERS_READ_COMPARE helper={helper}"
    if key:
        message += f" key={key}"
    message += f" status=mismatch fields={','.join(fields)}"
    if error_class:
        message += f" compare_error_class={error_class}"
    if error_stage:
        message += f" compare_error_stage={error_stage}"
    dual_write_log(message)


def resolve_crew_business_date(now: datetime | None = None) -> str:
    current = now or datetime.now()
    reset_point = current.replace(
        hour=CREW_BUSINESS_DAY_RESET_HOUR,
        minute=CREW_BUSINESS_DAY_RESET_MINUTE,
        second=0,
        microsecond=0,
    )
    if current < reset_point:
        current = current - timedelta(days=1)
    return current.date().isoformat()


def _compare_user_lookup(
    *,
    helper: str,
    key: str,
    primary_row: sqlite3.Row | None,
    shadow_row,
    log_result: bool,
) -> dict[str, object]:
    primary_payload = _user_row_payload(primary_row, include_password_hash=True)
    shadow_payload = _user_row_payload(shadow_row, include_password_hash=True)
    mismatch_fields: list[str] = []

    if (primary_payload is None) != (shadow_payload is None):
        mismatch_fields.append("exists")
    elif primary_payload and shadow_payload:
        for field_name in ("id", "username", "display_name", "role", "created_at"):
            if primary_payload[field_name] != shadow_payload[field_name]:
                mismatch_fields.append(field_name)
        if primary_payload["password_hash"] != shadow_payload["password_hash"]:
            mismatch_fields.append("password_hash_match")

    result = {
        "helper": helper,
        "key": key,
        "status": "match" if not mismatch_fields else "mismatch",
        "fields": mismatch_fields,
        "exists_match": (primary_payload is None) == (shadow_payload is None),
        "password_hash_match": "password_hash_match" not in mismatch_fields,
    }
    if log_result:
        if mismatch_fields:
            _log_users_read_compare_mismatch(helper=helper, key=key, fields=mismatch_fields)
        else:
            _log_users_read_compare_match(helper=helper, key=key)
    return result


def _compare_list_users(*, primary_rows: list[sqlite3.Row], shadow_rows, log_result: bool) -> dict[str, object]:
    primary_payloads = [_list_user_row_payload(row) for row in primary_rows]
    shadow_payloads = [_list_user_row_payload(row) for row in shadow_rows]
    row_count_match = len(primary_payloads) == len(shadow_payloads)
    primary_ids = [row["id"] for row in primary_payloads]
    shadow_ids = [row["id"] for row in shadow_payloads]
    ordered_ids_match = primary_ids == shadow_ids
    detail_mismatches: list[dict[str, object]] = []

    shadow_by_id = {row["id"]: row for row in shadow_payloads}
    for primary_row in primary_payloads:
        row_id = primary_row["id"]
        shadow_row = shadow_by_id.get(row_id)
        mismatch_fields: list[str] = []
        if shadow_row is None:
            mismatch_fields.append("missing_in_shadow")
        else:
            for field_name in ("username", "display_name", "role", "created_at"):
                if primary_row[field_name] != shadow_row[field_name]:
                    mismatch_fields.append(field_name)
        if mismatch_fields:
            detail_mismatches.append({"id": row_id, "fields": mismatch_fields})

    primary_id_set = set(primary_ids)
    for shadow_row in shadow_payloads:
        if shadow_row["id"] not in primary_id_set:
            detail_mismatches.append({"id": shadow_row["id"], "fields": ["missing_in_primary"]})

    status = "match" if row_count_match and ordered_ids_match and not detail_mismatches else "mismatch"
    result = {
        "helper": "list_users",
        "status": status,
        "row_count_match": row_count_match,
        "ordered_ids_match": ordered_ids_match,
        "details": detail_mismatches,
    }

    if log_result:
        if status == "match":
            dual_write_log("USERS_READ_COMPARE helper=list_users status=match")
        else:
            dual_write_log(
                "USERS_READ_COMPARE helper=list_users "
                f"status=mismatch row_count_match={str(row_count_match).lower()} "
                f"ordered_ids_match={str(ordered_ids_match).lower()}"
            )
            for detail in detail_mismatches:
                dual_write_log(
                    "USERS_READ_COMPARE_DETAIL helper=list_users "
                    f"id={detail['id']} status=mismatch fields={','.join(detail['fields'])}"
                )
    return result


def _users_read_compare_error_result(
    *,
    helper: str,
    key: str | None = None,
    exc: Exception,
    list_mode: bool = False,
) -> dict[str, object]:
    error_class = type(exc).__name__
    error_stage = getattr(exc, "stage", "compare")
    if list_mode:
        return {
            "helper": helper,
            "status": "mismatch",
            "row_count_match": "unknown",
            "ordered_ids_match": "unknown",
            "details": [{"id": -1, "fields": ["compare_error"]}],
            "compare_error_class": error_class,
            "compare_error_stage": error_stage,
        }
    return {
        "helper": helper,
        "key": key,
        "status": "mismatch",
        "fields": ["compare_error"],
        "exists_match": False,
        "password_hash_match": False,
        "compare_error_class": error_class,
        "compare_error_stage": error_stage,
    }


def run_users_read_compare_by_username(username: str, *, log_result: bool = False) -> dict[str, object]:
    primary_row = _sqlite_get_user_by_username(username)
    try:
        shadow_row = _shadow_get_user_by_username(username)
        return _compare_user_lookup(
            helper="get_user_by_username",
            key=f"username:{username}",
            primary_row=primary_row,
            shadow_row=shadow_row,
            log_result=log_result,
        )
    except Exception as exc:
        result = _users_read_compare_error_result(
            helper="get_user_by_username",
            key=f"username:{username}",
            exc=exc,
        )
        if log_result:
            _log_users_read_compare_mismatch(
                helper="get_user_by_username",
                key=f"username:{username}",
                fields=["compare_error"],
                error_class=result["compare_error_class"],
                error_stage=result["compare_error_stage"],
            )
        return result


def run_users_read_compare_by_id(user_id: int, *, log_result: bool = False) -> dict[str, object]:
    primary_row = _sqlite_get_user_by_id(user_id)
    try:
        shadow_row = _shadow_get_user_by_id(user_id)
        return _compare_user_lookup(
            helper="get_user_by_id",
            key=f"id:{user_id}",
            primary_row=primary_row,
            shadow_row=shadow_row,
            log_result=log_result,
        )
    except Exception as exc:
        result = _users_read_compare_error_result(
            helper="get_user_by_id",
            key=f"id:{user_id}",
            exc=exc,
        )
        if log_result:
            _log_users_read_compare_mismatch(
                helper="get_user_by_id",
                key=f"id:{user_id}",
                fields=["compare_error"],
                error_class=result["compare_error_class"],
                error_stage=result["compare_error_stage"],
            )
        return result


def run_users_list_compare(*, log_result: bool = False) -> dict[str, object]:
    primary_rows = _sqlite_list_users()
    try:
        shadow_rows = _shadow_list_users()
        return _compare_list_users(primary_rows=primary_rows, shadow_rows=shadow_rows, log_result=log_result)
    except Exception as exc:
        result = _users_read_compare_error_result(helper="list_users", exc=exc, list_mode=True)
        if log_result:
            dual_write_log(
                "USERS_READ_COMPARE helper=list_users status=mismatch "
                f"row_count_match={result['row_count_match']} "
                f"ordered_ids_match={result['ordered_ids_match']} "
                f"compare_error_class={result['compare_error_class']} "
                f"compare_error_stage={result['compare_error_stage']}"
            )
            dual_write_log(
                "USERS_READ_COMPARE_DETAIL helper=list_users id=-1 status=mismatch fields=compare_error"
            )
        return result


def get_primary_postgres_connection() -> psycopg.Connection:
    database_url = os.environ.get("DATABASE_URL", "").strip()
    if not database_url:
        raise RuntimeError("DATABASE_URL is required for PostgreSQL secondary writes.")
    scheme = urlsplit(database_url).scheme.lower()
    if scheme not in {"postgresql", "postgres"}:
        raise RuntimeError(f"DATABASE_URL must point to PostgreSQL, got scheme '{scheme or 'missing'}'.")

    pg_conn = getattr(g, "_dual_write_pg_conn", None)
    if pg_conn is None or pg_conn.closed:
        pg_conn = psycopg.connect(database_url)
        g._dual_write_pg_conn = pg_conn
    return pg_conn


@app.teardown_appcontext
def close_primary_postgres_connection(exc: BaseException | None) -> None:
    pg_conn = g.pop("_dual_write_pg_conn", None)
    if pg_conn is None or pg_conn.closed:
        return
    try:
        if exc is None:
            pg_conn.commit()
        else:
            pg_conn.rollback()
    finally:
        pg_conn.close()


def update_floor_fields_sqlite(
    conn: sqlite3.Connection,
    floor_id: int,
    *,
    name: str,
    block_name: str,
) -> None:
    conn.execute(
        "UPDATE floors SET name = ?, block_name = ? WHERE id = ?",
        (name, block_name, floor_id),
    )


def update_user_display_name_sqlite(
    conn: sqlite3.Connection,
    user_id: int,
    *,
    display_name: str,
) -> None:
    conn.execute(
        "UPDATE users SET display_name = ? WHERE id = ?",
        (display_name, user_id),
    )


def update_user_role_sqlite(
    conn: sqlite3.Connection,
    user_id: int,
    *,
    role: str,
) -> None:
    conn.execute(
        "UPDATE users SET role = ? WHERE id = ?",
        (role, user_id),
    )


SITE_PERMISSION_ROLE_LABELS = {
    "supervisor": "工區主管",
    "member": "工地成員",
}


def get_site_permission_role_options() -> list[dict[str, str]]:
    return [{"value": value, "label": label} for value, label in SITE_PERMISSION_ROLE_LABELS.items()]


def normalize_site_permission_role(role: str) -> str:
    normalized = str(role or "").strip()
    if normalized not in SITE_PERMISSION_ROLE_LABELS:
        raise ValueError("site_role_invalid")
    return normalized


def delete_user_site_permissions_sqlite(conn: sqlite3.Connection, user_id: int) -> None:
    conn.execute("DELETE FROM user_site_permissions WHERE user_id = ?", (user_id,))


def create_user_site_permission_sqlite(
    conn: sqlite3.Connection,
    *,
    user_id: int,
    site_id: int,
    role: str,
) -> sqlite3.Row:
    normalized_role = normalize_site_permission_role(role)
    cur = conn.execute(
        """
        INSERT INTO user_site_permissions (user_id, site_id, role)
        VALUES (?, ?, ?)
        """,
        (user_id, site_id, normalized_role),
    )
    row = conn.execute(
        """
        SELECT usp.id, usp.user_id, usp.site_id, usp.role, s.site_name, s.site_code, s.is_active
        FROM user_site_permissions usp
        JOIN sites s ON s.id = usp.site_id
        WHERE usp.id = ?
        """,
        (cur.lastrowid,),
    ).fetchone()
    if row is None:
        raise RuntimeError(f"SQLite user_site_permissions create could not reload id={cur.lastrowid}.")
    return row


def update_user_site_permission_role_sqlite(
    conn: sqlite3.Connection,
    permission_id: int,
    *,
    role: str,
) -> sqlite3.Row:
    normalized_role = normalize_site_permission_role(role)
    cur = conn.execute(
        "UPDATE user_site_permissions SET role = ?, updated_at = CURRENT_TIMESTAMP WHERE id = ?",
        (normalized_role, permission_id),
    )
    if cur.rowcount != 1:
        raise LookupError(f"SQLite user_site_permissions update could not find id={permission_id}.")
    row = conn.execute(
        """
        SELECT usp.id, usp.user_id, usp.site_id, usp.role, s.site_name, s.site_code, s.is_active
        FROM user_site_permissions usp
        JOIN sites s ON s.id = usp.site_id
        WHERE usp.id = ?
        """,
        (permission_id,),
    ).fetchone()
    if row is None:
        raise RuntimeError(f"SQLite user_site_permissions update could not reload id={permission_id}.")
    return row


def delete_user_site_permission_sqlite(conn: sqlite3.Connection, permission_id: int) -> None:
    cur = conn.execute("DELETE FROM user_site_permissions WHERE id = ?", (permission_id,))
    if cur.rowcount != 1:
        raise LookupError(f"SQLite user_site_permissions delete could not find id={permission_id}.")


def create_user_sqlite(
    conn: sqlite3.Connection,
    *,
    username: str,
    display_name: str,
    password_hash: str,
    role: str,
) -> dict[str, object]:
    cur = conn.execute(
        "INSERT INTO users (username, display_name, password_hash, role) VALUES (?, ?, ?, ?)",
        (username, display_name, password_hash, role),
    )
    row = conn.execute(
        """
        SELECT id, username, display_name, password_hash, role, created_at
        FROM users
        WHERE id = ?
        """,
        (cur.lastrowid,),
    ).fetchone()
    if row is None:
        raise RuntimeError(f"SQLite users primary create could not reload id={cur.lastrowid}.")
    return {
        "id": row["id"],
        "username": row["username"],
        "display_name": row["display_name"],
        "password_hash": row["password_hash"],
        "role": row["role"],
        "created_at": row["created_at"],
    }


def is_protected_user_row(user_row: sqlite3.Row | dict[str, object]) -> bool:
    return (
        int(user_row["id"]) == 1
        or str(user_row["username"]) == "admin"
        or str(user_row["role"]) == "admin"
    )


def delete_user_sqlite(
    conn: sqlite3.Connection,
    user_id: int,
) -> dict[str, object]:
    row = conn.execute(
        """
        SELECT id, username, display_name, password_hash, role, created_at
        FROM users
        WHERE id = ?
        """,
        (user_id,),
    ).fetchone()
    if row is None:
        raise LookupError(f"SQLite users primary delete could not find id={user_id}.")
    if is_protected_user_row(row):
        raise ValueError(f"SQLite users primary delete refused protected id={user_id}.")
    delete_user_site_permissions_sqlite(conn, user_id)
    cur = conn.execute("DELETE FROM users WHERE id = ?", (user_id,))
    if cur.rowcount != 1:
        raise RuntimeError(f"SQLite users primary delete affected {cur.rowcount} rows for user_id={user_id}.")
    return {
        "id": row["id"],
        "username": row["username"],
        "display_name": row["display_name"],
        "role": row["role"],
        "created_at": row["created_at"],
    }


def update_floor_fields_postgres(
    pg_conn: psycopg.Connection,
    floor_id: int,
    *,
    name: str,
    block_name: str,
) -> None:
    savepoint_name = f"dual_write_floors_update_{floor_id}"
    dual_write_log(f"{DUAL_WRITE_FLOORS_STRATEGY_LOG} floor_id={floor_id}")
    with pg_conn.cursor() as cur:
        dual_write_log(f"DUAL_WRITE_FLOORS_SECONDARY table=floors event=SAVEPOINT_START floor_id={floor_id}")
        cur.execute(f"SAVEPOINT {savepoint_name}")
        dual_write_log(f"DUAL_WRITE_FLOORS_SECONDARY table=floors event=SAVEPOINT_OK floor_id={floor_id}")
        try:
            dual_write_log(f"DUAL_WRITE_FLOORS_SECONDARY table=floors event=EXECUTE_SQL_START floor_id={floor_id}")
            cur.execute(
                "UPDATE floors SET name = %s, block_name = %s WHERE id = %s",
                (name, block_name, floor_id),
            )
            if cur.rowcount != 1:
                raise RuntimeError(
                    f"PostgreSQL floors secondary update affected {cur.rowcount} rows for floor_id={floor_id}."
                )
            dual_write_log(f"DUAL_WRITE_FLOORS_SECONDARY table=floors event=EXECUTE_SQL_OK floor_id={floor_id}")
            cur.execute(f"RELEASE SAVEPOINT {savepoint_name}")
            dual_write_log(
                f"DUAL_WRITE_FLOORS_SECONDARY table=floors event=RELEASE_SAVEPOINT_OK floor_id={floor_id}"
            )
            pg_conn.commit()
        except Exception:
            try:
                cur.execute(f"ROLLBACK TO SAVEPOINT {savepoint_name}")
            finally:
                cur.execute(f"RELEASE SAVEPOINT {savepoint_name}")
            pg_conn.commit()
            raise


def update_user_display_name_postgres(
    pg_conn: psycopg.Connection,
    user_id: int,
    *,
    display_name: str,
) -> None:
    savepoint_name = f"dual_write_users_update_{user_id}"
    dual_write_log(f"{DUAL_WRITE_USERS_STRATEGY_LOG} user_id={user_id}")
    with pg_conn.cursor() as cur:
        dual_write_log(f"DUAL_WRITE_USERS_SECONDARY table=users event=SAVEPOINT_START user_id={user_id}")
        cur.execute(f"SAVEPOINT {savepoint_name}")
        dual_write_log(f"DUAL_WRITE_USERS_SECONDARY table=users event=SAVEPOINT_OK user_id={user_id}")
        try:
            dual_write_log(f"DUAL_WRITE_USERS_SECONDARY table=users event=EXECUTE_SQL_START user_id={user_id}")
            cur.execute(
                "UPDATE users SET display_name = %s WHERE id = %s",
                (display_name, user_id),
            )
            if cur.rowcount != 1:
                raise RuntimeError(
                    f"PostgreSQL users secondary update affected {cur.rowcount} rows for user_id={user_id}."
                )
            dual_write_log(f"DUAL_WRITE_USERS_SECONDARY table=users event=EXECUTE_SQL_OK user_id={user_id}")
            cur.execute(f"RELEASE SAVEPOINT {savepoint_name}")
            dual_write_log(
                f"DUAL_WRITE_USERS_SECONDARY table=users event=RELEASE_SAVEPOINT_OK user_id={user_id}"
            )
            pg_conn.commit()
        except Exception:
            try:
                cur.execute(f"ROLLBACK TO SAVEPOINT {savepoint_name}")
            finally:
                cur.execute(f"RELEASE SAVEPOINT {savepoint_name}")
            pg_conn.commit()
            raise


def update_user_role_postgres(
    pg_conn: psycopg.Connection,
    user_id: int,
    *,
    role: str,
) -> None:
    savepoint_name = f"dual_write_users_update_{user_id}"
    dual_write_log(f"{DUAL_WRITE_USERS_STRATEGY_LOG} user_id={user_id}")
    with pg_conn.cursor() as cur:
        dual_write_log(f"DUAL_WRITE_USERS_SECONDARY table=users event=SAVEPOINT_START user_id={user_id}")
        cur.execute(f"SAVEPOINT {savepoint_name}")
        dual_write_log(f"DUAL_WRITE_USERS_SECONDARY table=users event=SAVEPOINT_OK user_id={user_id}")
        try:
            dual_write_log(f"DUAL_WRITE_USERS_SECONDARY table=users event=EXECUTE_SQL_START user_id={user_id}")
            cur.execute(
                "UPDATE users SET role = %s WHERE id = %s",
                (role, user_id),
            )
            if cur.rowcount != 1:
                raise RuntimeError(
                    f"PostgreSQL users secondary update affected {cur.rowcount} rows for user_id={user_id}."
                )
            dual_write_log(f"DUAL_WRITE_USERS_SECONDARY table=users event=EXECUTE_SQL_OK user_id={user_id}")
            cur.execute(f"RELEASE SAVEPOINT {savepoint_name}")
            dual_write_log(
                f"DUAL_WRITE_USERS_SECONDARY table=users event=RELEASE_SAVEPOINT_OK user_id={user_id}"
            )
            pg_conn.commit()
        except Exception:
            try:
                cur.execute(f"ROLLBACK TO SAVEPOINT {savepoint_name}")
            finally:
                cur.execute(f"RELEASE SAVEPOINT {savepoint_name}")
            pg_conn.commit()
            raise


def create_user_postgres(
    pg_conn: psycopg.Connection,
    *,
    id: int,
    username: str,
    display_name: str,
    password_hash: str,
    role: str,
    created_at: str,
) -> None:
    savepoint_name = f"dual_write_users_create_{id}"
    dual_write_log(f"{DUAL_WRITE_USERS_STRATEGY_LOG} operation=create user_id={id}")
    with pg_conn.cursor() as cur:
        dual_write_log(f"DUAL_WRITE_USERS_SECONDARY table=users operation=create event=SAVEPOINT_START user_id={id}")
        cur.execute(f"SAVEPOINT {savepoint_name}")
        dual_write_log(f"DUAL_WRITE_USERS_SECONDARY table=users operation=create event=SAVEPOINT_OK user_id={id}")
        try:
            dual_write_log(
                f"DUAL_WRITE_USERS_SECONDARY table=users operation=create event=EXECUTE_SQL_START user_id={id}"
            )
            cur.execute(
                """
                INSERT INTO users (id, username, display_name, password_hash, role, created_at)
                VALUES (%s, %s, %s, %s, %s, %s)
                """,
                (id, username, display_name, password_hash, role, created_at),
            )
            if cur.rowcount != 1:
                raise RuntimeError(
                    f"PostgreSQL users secondary create affected {cur.rowcount} rows for user_id={id}."
                )
            dual_write_log(f"DUAL_WRITE_USERS_SECONDARY table=users operation=create event=EXECUTE_SQL_OK user_id={id}")
            cur.execute(f"RELEASE SAVEPOINT {savepoint_name}")
            dual_write_log(
                f"DUAL_WRITE_USERS_SECONDARY table=users operation=create event=RELEASE_SAVEPOINT_OK user_id={id}"
            )
            pg_conn.commit()
        except Exception:
            try:
                cur.execute(f"ROLLBACK TO SAVEPOINT {savepoint_name}")
            finally:
                cur.execute(f"RELEASE SAVEPOINT {savepoint_name}")
            pg_conn.commit()
            raise


def delete_user_postgres(
    pg_conn: psycopg.Connection,
    *,
    id: int,
    username: str,
) -> None:
    savepoint_name = f"dual_write_users_delete_{id}"
    dual_write_log(f"{DUAL_WRITE_USERS_STRATEGY_LOG} operation=delete user_id={id}")
    with pg_conn.cursor() as cur:
        dual_write_log(f"DUAL_WRITE_USERS_SECONDARY table=users operation=delete event=SAVEPOINT_START user_id={id}")
        cur.execute(f"SAVEPOINT {savepoint_name}")
        dual_write_log(f"DUAL_WRITE_USERS_SECONDARY table=users operation=delete event=SAVEPOINT_OK user_id={id}")
        try:
            dual_write_log(
                f"DUAL_WRITE_USERS_SECONDARY table=users operation=delete event=EXECUTE_SQL_START user_id={id}"
            )
            cur.execute(
                "DELETE FROM users WHERE id = %s AND username = %s",
                (id, username),
            )
            if cur.rowcount == 0:
                dual_write_log(
                    "DUAL_WRITE_USERS_SECONDARY table=users operation=delete "
                    f"event=SECONDARY_NOT_FOUND user_id={id} username={username!r}"
                )
            elif cur.rowcount != 1:
                raise RuntimeError(
                    f"PostgreSQL users secondary delete affected {cur.rowcount} rows for user_id={id}."
                )
            dual_write_log(f"DUAL_WRITE_USERS_SECONDARY table=users operation=delete event=EXECUTE_SQL_OK user_id={id}")
            cur.execute(f"RELEASE SAVEPOINT {savepoint_name}")
            dual_write_log(
                f"DUAL_WRITE_USERS_SECONDARY table=users operation=delete event=RELEASE_SAVEPOINT_OK user_id={id}"
            )
            pg_conn.commit()
        except Exception:
            try:
                cur.execute(f"ROLLBACK TO SAVEPOINT {savepoint_name}")
            finally:
                cur.execute(f"RELEASE SAVEPOINT {savepoint_name}")
            pg_conn.commit()
            raise


def maybe_dual_write_floor_update(
    floor_id: int,
    *,
    name: str,
    block_name: str,
) -> None:
    if not controlled_dual_write_enabled(table="floors", operation="update"):
        return

    dry_run = dual_write_dry_run_enabled()
    if dry_run:
        dual_write_log(f"DUAL_WRITE_DRY_RUN operation=update table=floors floor_id={floor_id}")
        dual_write_log(f"{DUAL_WRITE_FLOORS_STRATEGY_LOG} floor_id={floor_id}")
        dual_write_log(f"DUAL_WRITE_FLOORS_SECONDARY table=floors event=SAVEPOINT_START floor_id={floor_id}")
        dual_write_log(f"DUAL_WRITE_FLOORS_SECONDARY table=floors event=SAVEPOINT_OK floor_id={floor_id}")
        dual_write_log(f"DUAL_WRITE_FLOORS_SECONDARY table=floors event=EXECUTE_SQL_START floor_id={floor_id}")
        dual_write_log(f"DUAL_WRITE_FLOORS_SECONDARY table=floors event=EXECUTE_SQL_OK floor_id={floor_id}")
        dual_write_log(
            f"DUAL_WRITE_FLOORS_SECONDARY table=floors event=RELEASE_SAVEPOINT_OK floor_id={floor_id}"
        )
        dual_write_log(
            "DUAL_WRITE operation=update table=floors "
            f"floor_id={floor_id} dry_run=true postgres_result=success"
        )
        return

    try:
        update_floor_fields_postgres(
            get_primary_postgres_connection(),
            floor_id,
            name=name,
            block_name=block_name,
        )
        dual_write_log(
            "DUAL_WRITE operation=update table=floors "
            f"floor_id={floor_id} dry_run=false postgres_result=success"
        )
    except Exception as exc:
        dual_write_log(
            "DUAL_WRITE operation=update table=floors "
            f"floor_id={floor_id} dry_run=false postgres_result=failed error={exc!r}"
        )
        if dual_write_strict_enabled():
            raise


def maybe_dual_write_user_display_name_update(
    user_id: int,
    *,
    display_name: str,
) -> None:
    if not controlled_dual_write_enabled(table="users", operation="update"):
        return

    dry_run = dual_write_dry_run_enabled()
    if dry_run:
        dual_write_log(f"DUAL_WRITE_DRY_RUN operation=update table=users user_id={user_id}")
        dual_write_log(f"{DUAL_WRITE_USERS_STRATEGY_LOG} user_id={user_id}")
        dual_write_log(f"DUAL_WRITE_USERS_SECONDARY table=users event=SAVEPOINT_START user_id={user_id}")
        dual_write_log(f"DUAL_WRITE_USERS_SECONDARY table=users event=SAVEPOINT_OK user_id={user_id}")
        dual_write_log(f"DUAL_WRITE_USERS_SECONDARY table=users event=EXECUTE_SQL_START user_id={user_id}")
        dual_write_log(f"DUAL_WRITE_USERS_SECONDARY table=users event=EXECUTE_SQL_OK user_id={user_id}")
        dual_write_log(
            f"DUAL_WRITE_USERS_SECONDARY table=users event=RELEASE_SAVEPOINT_OK user_id={user_id}"
        )
        dual_write_log(
            "DUAL_WRITE operation=update table=users "
            f"user_id={user_id} dry_run=true postgres_result=success"
        )
        return

    try:
        update_user_display_name_postgres(
            get_primary_postgres_connection(),
            user_id,
            display_name=display_name,
        )
        dual_write_log(
            "DUAL_WRITE operation=update table=users "
            f"user_id={user_id} dry_run=false postgres_result=success"
        )
    except Exception as exc:
        dual_write_log(
            "DUAL_WRITE operation=update table=users "
            f"user_id={user_id} dry_run=false postgres_result=failed error={exc!r}"
        )
        if dual_write_strict_enabled():
            raise


def maybe_dual_write_user_role_update(
    user_id: int,
    *,
    role: str,
) -> None:
    if not controlled_dual_write_enabled(table="users", operation="update"):
        return

    dry_run = dual_write_dry_run_enabled()
    if dry_run:
        dual_write_log(f"DUAL_WRITE_DRY_RUN operation=update table=users user_id={user_id}")
        dual_write_log(f"{DUAL_WRITE_USERS_STRATEGY_LOG} user_id={user_id}")
        dual_write_log(f"DUAL_WRITE_USERS_SECONDARY table=users event=SAVEPOINT_START user_id={user_id}")
        dual_write_log(f"DUAL_WRITE_USERS_SECONDARY table=users event=SAVEPOINT_OK user_id={user_id}")
        dual_write_log(f"DUAL_WRITE_USERS_SECONDARY table=users event=EXECUTE_SQL_START user_id={user_id}")
        dual_write_log(f"DUAL_WRITE_USERS_SECONDARY table=users event=EXECUTE_SQL_OK user_id={user_id}")
        dual_write_log(
            f"DUAL_WRITE_USERS_SECONDARY table=users event=RELEASE_SAVEPOINT_OK user_id={user_id}"
        )
        dual_write_log(
            "DUAL_WRITE operation=update table=users "
            f"user_id={user_id} dry_run=true postgres_result=success"
        )
        return

    try:
        update_user_role_postgres(
            get_primary_postgres_connection(),
            user_id,
            role=role,
        )
        dual_write_log(
            "DUAL_WRITE operation=update table=users "
            f"user_id={user_id} dry_run=false postgres_result=success"
        )
    except Exception as exc:
        dual_write_log(
            "DUAL_WRITE operation=update table=users "
            f"user_id={user_id} dry_run=false postgres_result=failed error={exc!r}"
        )
        if dual_write_strict_enabled():
            raise


def maybe_dual_write_user_create(user_row: dict[str, object]) -> None:
    if not controlled_dual_write_enabled(table="users", operation="create"):
        return

    user_id = int(user_row["id"])
    username = str(user_row["username"])
    dry_run = dual_write_dry_run_enabled()
    if dry_run:
        dual_write_log(
            f"DUAL_WRITE_DRY_RUN operation=create table=users user_id={user_id} username={username!r}"
        )
        dual_write_log(f"{DUAL_WRITE_USERS_STRATEGY_LOG} operation=create user_id={user_id}")
        dual_write_log(
            f"DUAL_WRITE_USERS_SECONDARY table=users operation=create event=SAVEPOINT_START user_id={user_id}"
        )
        dual_write_log(
            f"DUAL_WRITE_USERS_SECONDARY table=users operation=create event=SAVEPOINT_OK user_id={user_id}"
        )
        dual_write_log(
            f"DUAL_WRITE_USERS_SECONDARY table=users operation=create event=EXECUTE_SQL_START user_id={user_id}"
        )
        dual_write_log(
            f"DUAL_WRITE_USERS_SECONDARY table=users operation=create event=EXECUTE_SQL_OK user_id={user_id}"
        )
        dual_write_log(
            f"DUAL_WRITE_USERS_SECONDARY table=users operation=create event=RELEASE_SAVEPOINT_OK user_id={user_id}"
        )
        dual_write_log(
            "DUAL_WRITE operation=create table=users "
            f"user_id={user_id} username={username!r} dry_run=true postgres_result=success"
        )
        return

    try:
        create_user_postgres(
            get_primary_postgres_connection(),
            id=user_id,
            username=username,
            display_name=str(user_row["display_name"]),
            password_hash=str(user_row["password_hash"]),
            role=str(user_row["role"]),
            created_at=str(user_row["created_at"]),
        )
        dual_write_log(
            "DUAL_WRITE operation=create table=users "
            f"user_id={user_id} username={username!r} dry_run=false postgres_result=success"
        )
    except Exception as exc:
        dual_write_log(
            "DUAL_WRITE operation=create table=users "
            f"user_id={user_id} username={username!r} dry_run=false postgres_result=failed error={exc!r}"
        )
        if dual_write_strict_enabled():
            raise


def maybe_dual_write_user_delete(user_row: dict[str, object]) -> None:
    if not controlled_dual_write_enabled(table="users", operation="delete"):
        return

    user_id = int(user_row["id"])
    username = str(user_row["username"])
    dry_run = dual_write_dry_run_enabled()
    if dry_run:
        dual_write_log(
            f"DUAL_WRITE_DRY_RUN operation=delete table=users user_id={user_id} username={username!r}"
        )
        dual_write_log(f"{DUAL_WRITE_USERS_STRATEGY_LOG} operation=delete user_id={user_id}")
        dual_write_log(
            f"DUAL_WRITE_USERS_SECONDARY table=users operation=delete event=SAVEPOINT_START user_id={user_id}"
        )
        dual_write_log(
            f"DUAL_WRITE_USERS_SECONDARY table=users operation=delete event=SAVEPOINT_OK user_id={user_id}"
        )
        dual_write_log(
            f"DUAL_WRITE_USERS_SECONDARY table=users operation=delete event=EXECUTE_SQL_START user_id={user_id}"
        )
        dual_write_log(
            f"DUAL_WRITE_USERS_SECONDARY table=users operation=delete event=EXECUTE_SQL_OK user_id={user_id}"
        )
        dual_write_log(
            f"DUAL_WRITE_USERS_SECONDARY table=users operation=delete event=RELEASE_SAVEPOINT_OK user_id={user_id}"
        )
        dual_write_log(
            "DUAL_WRITE operation=delete table=users "
            f"user_id={user_id} username={username!r} dry_run=true postgres_result=success"
        )
        return

    try:
        delete_user_postgres(
            get_primary_postgres_connection(),
            id=user_id,
            username=username,
        )
        dual_write_log(
            "DUAL_WRITE operation=delete table=users "
            f"user_id={user_id} username={username!r} dry_run=false postgres_result=success"
        )
    except Exception as exc:
        dual_write_log(
            "DUAL_WRITE operation=delete table=users "
            f"user_id={user_id} username={username!r} dry_run=false postgres_result=failed error={exc!r}"
        )
        if dual_write_strict_enabled():
            raise


def query_one(sql: str, params: tuple = ()) -> sqlite3.Row | None:
    with db() as conn:
        return conn.execute(sql, params).fetchone()


def get_user_by_username(username: str) -> sqlite3.Row | None:
    row = _sqlite_get_user_by_username(username)
    if users_read_compare_enabled():
        run_users_read_compare_by_username(username, log_result=True)
    return row


def get_user_by_id(user_id: int) -> sqlite3.Row | None:
    row = _sqlite_get_user_by_id(user_id)
    if users_read_compare_enabled():
        run_users_read_compare_by_id(user_id, log_result=True)
    return row


def get_vendor_account_by_username(username: str) -> sqlite3.Row | None:
    return _sqlite_get_vendor_account_by_username(username)


def list_users() -> list[sqlite3.Row]:
    rows = _sqlite_list_users()
    if users_read_compare_enabled():
        run_users_list_compare(log_result=True)
    return rows


def crew_api_error(code: str, message: str, *, status: int = 400):
    return jsonify({"ok": False, "error": {"code": code, "message": message}}), status


def parse_crew_business_date(value: str) -> str:
    raw = value.strip()
    if not raw:
        raise ValueError("business_date is required.")
    try:
        return datetime.strptime(raw, "%Y-%m-%d").date().isoformat()
    except ValueError as exc:
        raise ValueError("business_date must use YYYY-MM-DD.") from exc


def parse_crew_planned_at(value: str) -> str:
    raw = value.strip()
    if not raw:
        return ""
    for fmt in ("%Y-%m-%d %H:%M", "%Y-%m-%dT%H:%M", "%Y-%m-%d %H:%M:%S", "%Y-%m-%dT%H:%M:%S"):
        try:
            parsed = datetime.strptime(raw, fmt)
            return parsed.strftime("%Y-%m-%d %H:%M")
        except ValueError:
            continue
    raise ValueError("planned_at must use YYYY-MM-DD HH:MM.")


def parse_non_negative_int(value, *, field_name: str) -> int:
    try:
        parsed = int(value)
    except (TypeError, ValueError) as exc:
        raise ValueError(f"{field_name} must be a non-negative integer.") from exc
    if parsed < 0:
        raise ValueError(f"{field_name} must be a non-negative integer.")
    return parsed


def parse_optional_positive_int(value, *, field_name: str) -> int | None:
    if value in (None, ""):
        return None
    try:
        parsed = int(value)
    except (TypeError, ValueError) as exc:
        raise ValueError(f"{field_name} must be a positive integer.") from exc
    if parsed <= 0:
        raise ValueError(f"{field_name} must be a positive integer.")
    return parsed


def parse_optional_non_negative_int(value, *, field_name: str) -> int | None:
    if value in (None, ""):
        return None
    return parse_non_negative_int(value, field_name=field_name)


def parse_contact_primary_flag(value) -> int:
    if isinstance(value, bool):
        return 1 if value else 0
    if value in (0, 1):
        return int(value)
    if isinstance(value, str):
        raw = value.strip().lower()
        if raw in ("0", "1"):
            return int(raw)
        if raw in ("true", "false"):
            return 1 if raw == "true" else 0
    raise ValueError("is_primary must be 0, 1, true, or false.")


def require_sheet_exists(conn: sqlite3.Connection, sheet_id: int) -> None:
    row = conn.execute("SELECT id FROM sheets WHERE id = ?", (sheet_id,)).fetchone()
    if row is None:
        raise LookupError(f"sheet_id={sheet_id} was not found.")


def _get_sheet_row(conn: sqlite3.Connection, sheet_id: int) -> sqlite3.Row | None:
    return conn.execute("SELECT * FROM sheets WHERE id = ?", (sheet_id,)).fetchone()


def _resolve_non_admin_read_site_id(conn: sqlite3.Connection, user) -> int:
    current_site_id = get_current_site_id()
    if current_site_id is None:
        raise LookupError("site_context_invalid")
    site_row = _fetch_site_row_by_id(conn, int(current_site_id))
    if site_row is None or int(site_row["is_active"]) != 1:
        raise LookupError("site_context_invalid")
    permission_row = conn.execute(
        """
        SELECT 1
        FROM user_site_permissions usp
        JOIN sites s ON s.id = usp.site_id
        WHERE usp.user_id = ? AND usp.site_id = ? AND s.is_active = 1
        """,
        (int(user["id"]), int(current_site_id)),
    ).fetchone()
    if permission_row is None:
        raise LookupError("site_permission_missing")
    return int(current_site_id)


def _resolve_read_scope(conn: sqlite3.Connection) -> tuple[object | None, int | None, bool]:
    user = _current_internal_user() if has_request_context() else None
    if user is None:
        return None, None, True
    if is_global_admin(user):
        return user, None, True
    current_site_id = _resolve_non_admin_read_site_id(conn, user)
    return user, current_site_id, False


def site_read_api_error(code: str, message: str, *, status: int) -> tuple[Response, int]:
    return jsonify({"ok": False, "error": {"code": code, "message": message}}), status


def _handle_sheet_read_lookup_error(exc: LookupError):
    code = str(exc)
    if code == "site_context_invalid":
        flash("\u8acb\u5148\u9078\u64c7\u76ee\u524d\u5de5\u5730\u3002", "error")
        return redirect(url_for("site_selector"))
    if code == "sheet_not_in_current_site":
        flash("\u8a72\u8868\u55ae\u4e0d\u5c6c\u65bc\u76ee\u524d\u5de5\u5730\u3002", "error")
        return redirect(url_for("sheet"))
    if code == "site_permission_missing":
        flash("\u4f60\u76ee\u524d\u6c92\u6709\u6b64\u5de5\u5730\u7684\u8b80\u53d6\u6b0a\u9650\u3002", "error")
        return redirect(url_for("site_selector"))
    if code == "no_sheets_in_current_site":
        flash("\u76ee\u524d\u5de5\u5730\u5c1a\u7121\u53ef\u8b80\u53d6\u7684\u8868\u55ae\u3002", "error")
        return redirect(url_for("site_selector"))
    if code == "sheet_not_found":
        flash("\u627e\u4e0d\u5230\u6307\u5b9a\u8868\u55ae\u3002", "error")
        return redirect(url_for("sheet"))
    raise exc


def _handle_grid_read_lookup_error(exc: LookupError):
    code = str(exc)
    if code == "sheet_not_found":
        return site_read_api_error(code, "sheet_id was not found.", status=404)
    if code == "sheet_not_in_current_site":
        return site_read_api_error(code, "sheet_id does not belong to the current site.", status=403)
    if code == "site_permission_missing":
        return site_read_api_error(code, "current user no longer has permission for the current site.", status=403)
    if code == "site_context_invalid":
        return site_read_api_error(code, "current_site_id is missing or invalid.", status=403)
    if code == "no_sheets_in_current_site":
        return site_read_api_error(code, "no sheets are available in the current site.", status=403)
    raise exc


def _handle_dashboard_lookup_error(exc: LookupError):
    code = str(exc)
    if code == "vendor_auth_forbidden":
        return crew_api_error("vendor_auth_forbidden", "vendor authentication cannot access dashboard.", status=403)
    if code == "auth_required":
        return crew_api_error("auth_required", "authentication is required.", status=403)
    return _handle_grid_read_lookup_error(exc)


def _handle_scheduling_lookup_error(exc: LookupError):
    code = str(exc)
    if code == "vendor_auth_forbidden":
        return crew_api_error("vendor_auth_forbidden", "vendor authentication cannot access scheduling.", status=403)
    if code == "auth_required":
        return crew_api_error("auth_required", "authentication is required.", status=403)
    return _handle_grid_read_lookup_error(exc)


def _handle_work_hub_runtime_lookup_error(exc: LookupError):
    code = str(exc)
    if code == "vendor_auth_forbidden":
        return crew_api_error("vendor_auth_forbidden", "vendor authentication cannot access work hub runtime.", status=403)
    if code == "auth_required":
        return crew_api_error("auth_required", "authentication is required.", status=403)
    return _handle_grid_read_lookup_error(exc)


def _handle_management_read_model_lookup_error(exc: LookupError):
    code = str(exc)
    if code == "vendor_auth_forbidden":
        return crew_api_error("vendor_auth_forbidden", "vendor authentication cannot access management read model.", status=403)
    if code == "auth_required":
        return crew_api_error("auth_required", "authentication is required.", status=403)
    return _handle_grid_read_lookup_error(exc)


def _handle_unit_handover_report_lookup_error(exc: LookupError):
    code = str(exc)
    if code == "unit_not_found":
        return "找不到指定的單戶報表。", 404
    if code == "sheet_not_in_current_site":
        return "無權存取指定的單戶報表。", 403
    if code in {"site_context_invalid", "site_permission_missing"}:
        return redirect(url_for("site_selector"))
    if code == "vendor_auth_forbidden":
        return "此身分無法存取單戶報表。", 403
    if code == "auth_required":
        return redirect(url_for("login"))
    raise exc


def _handle_floor_handover_report_lookup_error(exc: LookupError):
    code = str(exc)
    if code == "floor_not_found":
        return "找不到指定的樓層報表。", 404
    if code == "sheet_not_in_current_site":
        return "無權存取指定的樓層報表。", 403
    if code in {"site_context_invalid", "site_permission_missing"}:
        return redirect(url_for("site_selector"))
    if code == "vendor_auth_forbidden":
        return "此身分無法存取樓層報表。", 403
    if code == "auth_required":
        return redirect(url_for("login"))
    raise exc


def resolve_admin_current_site_id(conn: sqlite3.Connection) -> int:
    user = _current_internal_user()
    if user is None:
        raise LookupError("auth_required")
    current_site_id = get_current_site_id()
    if current_site_id is None:
        raise LookupError("site_context_invalid")
    site_row = _fetch_site_row_by_id(conn, int(current_site_id))
    if site_row is None or int(site_row["is_active"]) != 1:
        raise LookupError("site_context_invalid")
    return int(site_row["id"])


def resolve_sheet_site_for_admin_write(conn: sqlite3.Connection, *, sheet_id: int) -> dict[str, int]:
    row = conn.execute("SELECT id, site_id FROM sheets WHERE id = ?", (sheet_id,)).fetchone()
    if row is None:
        raise LookupError("sheet_not_found")
    site_id = row["site_id"]
    if site_id in (None, ""):
        raise LookupError("sheet_site_missing")
    return {
        "sheet_id": int(row["id"]),
        "site_id": int(site_id),
    }


def authorize_admin_site_scoped_write(conn: sqlite3.Connection, *, sheet_id: int) -> dict[str, int]:
    context = resolve_sheet_site_for_admin_write(conn, sheet_id=sheet_id)
    current_site_id = resolve_admin_current_site_id(conn)
    if int(context["site_id"]) != int(current_site_id):
        raise LookupError("write_target_not_in_current_site")
    context["current_site_id"] = int(current_site_id)
    return context


def authorize_admin_create_sheet_site(conn: sqlite3.Connection) -> dict[str, int]:
    current_site_id = resolve_admin_current_site_id(conn)
    return {"site_id": int(current_site_id)}


def resolve_task_sheet_for_admin_write(conn: sqlite3.Connection, *, task_id: int) -> dict[str, int]:
    row = conn.execute("SELECT id, sheet_id FROM tasks WHERE id = ?", (task_id,)).fetchone()
    if row is None:
        raise LookupError("task_not_found")
    sheet_id = row["sheet_id"]
    if sheet_id in (None, ""):
        raise LookupError("task_sheet_missing")
    return {
        "task_id": int(row["id"]),
        "sheet_id": int(sheet_id),
    }


def resolve_floor_sheet_for_admin_write(conn: sqlite3.Connection, *, floor_id: int) -> dict[str, int]:
    row = conn.execute("SELECT id, sheet_id FROM floors WHERE id = ?", (floor_id,)).fetchone()
    if row is None:
        raise LookupError("floor_not_found")
    sheet_id = row["sheet_id"]
    if sheet_id in (None, ""):
        raise LookupError("floor_sheet_missing")
    return {
        "floor_id": int(row["id"]),
        "sheet_id": int(sheet_id),
    }


def resolve_unit_sheet_for_admin_write(conn: sqlite3.Connection, *, unit_id: int) -> dict[str, int]:
    row = conn.execute(
        """
        SELECT u.id AS unit_id, u.floor_id AS floor_id, f.sheet_id AS sheet_id
        FROM units u
        LEFT JOIN floors f ON f.id = u.floor_id
        WHERE u.id = ?
        """,
        (unit_id,),
    ).fetchone()
    if row is None:
        raise LookupError("unit_not_found")
    floor_id = row["floor_id"]
    if floor_id in (None, ""):
        raise LookupError("unit_floor_missing")
    sheet_id = row["sheet_id"]
    if sheet_id in (None, ""):
        raise LookupError("unit_sheet_missing")
    return {
        "unit_id": int(row["unit_id"]),
        "floor_id": int(floor_id),
        "sheet_id": int(sheet_id),
    }


def resolve_extra_field_sheet_for_admin_write(conn: sqlite3.Connection, *, field_id: int) -> dict[str, int]:
    row = conn.execute(
        "SELECT id, sheet_id, active FROM extra_fields WHERE id = ?",
        (field_id,),
    ).fetchone()
    if row is None:
        raise LookupError("extra_field_not_found")
    sheet_id = row["sheet_id"]
    if sheet_id in (None, ""):
        raise LookupError("extra_field_sheet_missing")
    return {
        "field_id": int(row["id"]),
        "sheet_id": int(sheet_id),
        "active": int(row["active"] or 0),
    }


def _handle_admin_site_write_lookup_error(exc: LookupError, *, sheet_id: int | None = None):
    code = str(exc)
    if code == "site_context_invalid":
        flash("\u8acb\u5148\u9078\u64c7\u76ee\u524d\u5de5\u5730\u3002", "error")
        return redirect(url_for("site_selector"))
    if code == "write_target_not_in_current_site":
        flash("\u76ee\u524d\u5de5\u5730\u4e0b\u4e0d\u53ef\u4fee\u6539\u5176\u4ed6\u5de5\u5730\u8cc7\u6599\u3002", "error")
        if sheet_id is not None:
            return redirect(url_for("table_admin", sheet_id=sheet_id))
        return redirect(url_for("table_admin"))
    if code in {"sheet_not_found", "sheet_site_missing"}:
        flash("\u627e\u4e0d\u5230\u76ee\u6a19\u8868\u55ae\u3002", "error")
        if sheet_id is not None:
            return redirect(url_for("table_admin", sheet_id=sheet_id))
        return redirect(url_for("table_admin"))
    if code in {"task_not_found", "task_sheet_missing", "task_sheet_mismatch"}:
        flash("\u627e\u4e0d\u5230\u76ee\u6a19\u8cc7\u6599\u3002", "error")
        if sheet_id is not None:
            return redirect(url_for("table_admin", sheet_id=sheet_id))
        return redirect(url_for("table_admin"))
    if code in {"floor_not_found", "floor_sheet_missing", "floor_sheet_mismatch"}:
        flash("\u627e\u4e0d\u5230\u76ee\u6a19\u8cc7\u6599\u3002", "error")
        if sheet_id is not None:
            return redirect(url_for("table_admin", sheet_id=sheet_id))
        return redirect(url_for("table_admin"))
    if code in {"unit_not_found", "unit_floor_missing", "unit_sheet_missing", "unit_sheet_mismatch"}:
        flash("\u627e\u4e0d\u5230\u76ee\u6a19\u8cc7\u6599\u3002", "error")
        if sheet_id is not None:
            return redirect(url_for("table_admin", sheet_id=sheet_id))
        return redirect(url_for("table_admin"))
    if code in {"extra_field_not_found", "extra_field_sheet_missing", "extra_field_sheet_mismatch"}:
        flash("\u627e\u4e0d\u5230\u76ee\u6a19\u8cc7\u6599\u3002", "error")
        if sheet_id is not None:
            return redirect(url_for("table_admin", sheet_id=sheet_id))
        return redirect(url_for("table_admin"))
    if code == "auth_required":
        flash("\u8acb\u5148\u767b\u5165\u3002", "error")
        return redirect(url_for("login"))
    raise exc


def progress_api_error(message: str, *, status: int = 400):
    return jsonify({"ok": False, "message": message}), status


def _handle_admin_reset_sheet_lookup_error(exc: LookupError):
    code = str(exc)
    if code in {"sheet_not_found", "sheet_site_missing"}:
        return progress_api_error("找不到目標表單。", status=404)
    if code == "site_context_invalid":
        return progress_api_error("請先選擇目前工地。", status=403)
    if code == "write_target_not_in_current_site":
        return progress_api_error("目前工地下不可重設其他工地資料。", status=403)
    if code == "invalid_request":
        return progress_api_error("sheet_id 無效。", status=400)
    raise exc


def build_admin_site_write_render_state(
    conn,
    *,
    sheet_id,
    missing_current_site_message,
    sheet_not_in_current_site_message,
):
    state = {
        "enabled": True,
        "block_reason": "",
        "block_message": "",
    }
    try:
        authorize_admin_site_scoped_write(conn, sheet_id=sheet_id)
    except LookupError as exc:
        code = exc.args[0] if exc.args else ""
        if code == "site_context_invalid":
            state["enabled"] = False
            state["block_reason"] = "missing_current_site"
            state["block_message"] = missing_current_site_message
        elif code == "write_target_not_in_current_site":
            state["enabled"] = False
            state["block_reason"] = "sheet_not_in_current_site"
            state["block_message"] = sheet_not_in_current_site_message
        else:
            raise
    return state


def authorize_sheet_read(conn: sqlite3.Connection, sheet_id: int) -> None:
    sheet_row = _get_sheet_row(conn, int(sheet_id))
    if sheet_row is None:
        raise LookupError("sheet_not_found")
    _user, current_site_id, is_admin = _resolve_read_scope(conn)
    if is_admin:
        return
    if int(sheet_row["site_id"] or 0) != int(current_site_id):
        raise LookupError("sheet_not_in_current_site")


def authorize_dashboard_read(conn: sqlite3.Connection, *, sheet_id: int) -> dict[str, int]:
    sheet_row = conn.execute("SELECT id, site_id FROM sheets WHERE id = ?", (sheet_id,)).fetchone()
    if sheet_row is None:
        raise LookupError("sheet_not_found")

    user = _current_internal_user()
    if user is None:
        if current_vendor_account() is not None:
            raise LookupError("vendor_auth_forbidden")
        raise LookupError("auth_required")

    if is_global_admin(user):
        current_site_id = resolve_admin_current_site_id(conn)
    else:
        current_site_id = _resolve_non_admin_read_site_id(conn, user)

    if int(sheet_row["site_id"] or 0) != int(current_site_id):
        raise LookupError("sheet_not_in_current_site")

    return {
        "sheet_id": int(sheet_row["id"]),
        "site_id": int(sheet_row["site_id"]),
        "current_site_id": int(current_site_id),
    }


def resolve_progress_write_context(conn: sqlite3.Connection, *, unit_id: int, task_id: int) -> dict[str, int]:
    unit_row = conn.execute(
        """
        SELECT u.id AS unit_id, f.sheet_id AS sheet_id, s.site_id AS site_id
        FROM units u
        JOIN floors f ON f.id = u.floor_id
        JOIN sheets s ON s.id = f.sheet_id
        WHERE u.id = ?
        """,
        (unit_id,),
    ).fetchone()
    if unit_row is None:
        raise LookupError("unit_not_found")

    task_row = conn.execute(
        "SELECT id, sheet_id FROM tasks WHERE id = ?",
        (task_id,),
    ).fetchone()
    if task_row is None:
        raise LookupError("task_not_found")

    if int(task_row["sheet_id"] or 0) != int(unit_row["sheet_id"] or 0):
        raise LookupError("unit_task_sheet_mismatch")

    return {
        "unit_id": int(unit_row["unit_id"]),
        "task_id": int(task_row["id"]),
        "sheet_id": int(unit_row["sheet_id"]),
        "site_id": int(unit_row["site_id"]),
    }


def authorize_progress_write(conn: sqlite3.Connection, *, unit_id: int, task_id: int) -> dict[str, int]:
    context = resolve_progress_write_context(conn, unit_id=unit_id, task_id=task_id)
    user = _current_internal_user()
    if user is None:
        raise LookupError("auth_required")
    if is_global_admin(user):
        return context

    current_site_id = _resolve_non_admin_read_site_id(conn, user)
    if int(context["site_id"]) != int(current_site_id):
        raise LookupError("write_target_not_in_current_site")
    return context


def _handle_progress_write_lookup_error(exc: LookupError):
    code = str(exc)
    if code == "unit_not_found":
        return progress_api_error("unit_id was not found.", status=404)
    if code == "task_not_found":
        return progress_api_error("task_id was not found.", status=404)
    if code == "unit_task_sheet_mismatch":
        return progress_api_error("unit_id and task_id do not belong to the same sheet.", status=409)
    if code == "site_context_invalid":
        return progress_api_error("current_site_id is missing or invalid.", status=403)
    if code == "site_permission_missing":
        return progress_api_error("current user no longer has permission for the current site.", status=403)
    if code == "write_target_not_in_current_site":
        return progress_api_error("write target does not belong to the current site.", status=403)
    if code == "auth_required":
        return progress_api_error("authentication is required.", status=403)
    raise exc


def unit_extra_api_error(message: str, *, status: int = 400):
    return jsonify({"ok": False, "message": message}), status


def resolve_unit_extra_write_context(conn: sqlite3.Connection, *, unit_id: int, field_key: str) -> dict[str, object]:
    unit_row = conn.execute(
        """
        SELECT u.id AS unit_id, f.id AS floor_id, f.sheet_id AS sheet_id, s.site_id AS site_id
        FROM units u
        JOIN floors f ON f.id = u.floor_id
        JOIN sheets s ON s.id = f.sheet_id
        WHERE u.id = ?
        """,
        (unit_id,),
    ).fetchone()
    if unit_row is None:
        raise LookupError("unit_not_found")

    field_row = conn.execute(
        """
        SELECT id, sheet_id, field_key, field_type
        FROM extra_fields
        WHERE sheet_id = ? AND field_key = ? AND active = 1
        """,
        (int(unit_row["sheet_id"]), field_key),
    ).fetchone()
    if field_row is None:
        any_field_row = conn.execute(
            """
            SELECT 1
            FROM extra_fields
            WHERE field_key = ? AND active = 1
            """,
            (field_key,),
        ).fetchone()
        if any_field_row is None:
            raise LookupError("field_not_found")
        raise LookupError("unit_field_sheet_mismatch")

    return {
        "unit_id": int(unit_row["unit_id"]),
        "floor_id": int(unit_row["floor_id"]),
        "sheet_id": int(unit_row["sheet_id"]),
        "site_id": int(unit_row["site_id"]),
        "field_id": int(field_row["id"]),
        "field_key": str(field_row["field_key"]),
        "field_type": str(field_row["field_type"]),
    }


def authorize_unit_extra_write(conn: sqlite3.Connection, *, unit_id: int, field_key: str) -> dict[str, object]:
    context = resolve_unit_extra_write_context(conn, unit_id=unit_id, field_key=field_key)
    user = _current_internal_user()
    if user is None:
        raise LookupError("auth_required")
    if is_global_admin(user):
        return context

    current_site_id = _resolve_non_admin_read_site_id(conn, user)
    if int(context["site_id"]) != int(current_site_id):
        raise LookupError("write_target_not_in_current_site")
    return context


def _handle_unit_extra_write_lookup_error(exc: LookupError):
    code = str(exc)
    if code == "unit_not_found":
        return unit_extra_api_error("unit_id was not found.", status=404)
    if code == "field_not_found":
        return unit_extra_api_error("field was not found.", status=404)
    if code == "unit_field_sheet_mismatch":
        return unit_extra_api_error("unit_id and field do not belong to the same sheet.", status=409)
    if code == "site_context_invalid":
        return unit_extra_api_error("current_site_id is missing or invalid.", status=403)
    if code == "site_permission_missing":
        return unit_extra_api_error("current user no longer has permission for the current site.", status=403)
    if code == "write_target_not_in_current_site":
        return unit_extra_api_error("write target does not belong to the current site.", status=403)
    if code == "auth_required":
        return unit_extra_api_error("authentication is required.", status=403)
    raise exc


def normalize_vendor_name(value: str) -> str:
    vendor_name = value.strip()
    if not vendor_name:
        raise ValueError("vendor_name is required.")
    if len(vendor_name) > 100:
        raise ValueError("vendor_name must be 100 characters or fewer.")
    return vendor_name


def validate_vendor_belongs_to_sheet(
    conn: sqlite3.Connection,
    *,
    sheet_id: int,
    vendor_name: str,
) -> None:
    row = conn.execute(
        """
        SELECT 1
        FROM tasks
        WHERE sheet_id = ? AND TRIM(COALESCE(vendor, '')) = ?
        LIMIT 1
        """,
        (sheet_id, vendor_name),
    ).fetchone()
    if row is None:
        raise LookupError("vendor_not_in_sheet")


def resolve_vendor_contact_write_context(
    conn: sqlite3.Connection,
    *,
    sheet_id: int,
    vendor_name: str,
    contact_id: int | None = None,
) -> dict[str, object]:
    sheet_row = conn.execute("SELECT id, site_id FROM sheets WHERE id = ?", (sheet_id,)).fetchone()
    if sheet_row is None:
        raise LookupError("sheet_not_found")

    validate_vendor_belongs_to_sheet(conn, sheet_id=sheet_id, vendor_name=vendor_name)

    existing_contact_order: int | None = None
    if contact_id is not None:
        existing_contact = conn.execute(
            """
            SELECT id, sheet_id, vendor_name, contact_order
            FROM vendor_contacts
            WHERE id = ?
            """,
            (contact_id,),
        ).fetchone()
        if existing_contact is None:
            raise LookupError("contact_not_found")
        if int(existing_contact["sheet_id"] or 0) != int(sheet_id):
            raise LookupError("cross_sheet_update_not_allowed")
        existing_contact_order = int(existing_contact["contact_order"] or 0)

    return {
        "sheet_id": int(sheet_row["id"]),
        "site_id": int(sheet_row["site_id"]),
        "vendor_name": vendor_name,
        "contact_id": contact_id,
        "existing_contact_order": existing_contact_order,
    }


def authorize_vendor_contact_write(
    conn: sqlite3.Connection,
    *,
    sheet_id: int,
    vendor_name: str,
    contact_id: int | None = None,
) -> dict[str, object]:
    context = resolve_vendor_contact_write_context(
        conn,
        sheet_id=sheet_id,
        vendor_name=vendor_name,
        contact_id=contact_id,
    )
    user = _current_internal_user()
    if user is None:
        raise LookupError("auth_required")
    if is_global_admin(user):
        return context

    current_site_id = _resolve_non_admin_read_site_id(conn, user)
    if int(context["site_id"]) != int(current_site_id):
        raise LookupError("write_target_not_in_current_site")
    return context


def _handle_vendor_contact_lookup_error(exc: LookupError):
    code = str(exc)
    if code == "sheet_not_found":
        return crew_api_error("sheet_not_found", "sheet_id was not found.", status=404)
    if code == "vendor_not_in_sheet":
        return crew_api_error("vendor_not_in_sheet", "vendor_name does not belong to the requested sheet.", status=404)
    if code == "contact_not_found":
        return crew_api_error("contact_not_found", "contact id was not found.", status=404)
    if code == "cross_sheet_update_not_allowed":
        return crew_api_error("cross_sheet_update_not_allowed", "contact does not belong to the requested sheet.", status=400)
    if code == "site_context_invalid":
        return crew_api_error("site_context_invalid", "current_site_id is missing or invalid.", status=403)
    if code == "site_permission_missing":
        return crew_api_error("site_permission_missing", "current user no longer has permission for the current site.", status=403)
    if code == "write_target_not_in_current_site":
        return crew_api_error("write_target_not_in_current_site", "write target does not belong to the current site.", status=403)
    if code == "auth_required":
        return crew_api_error("auth_required", "authentication is required.", status=403)
    raise exc


def resolve_vendor_work_entry_write_context(
    conn: sqlite3.Connection,
    *,
    sheet_id: int,
    vendor_name: str,
    entry_id: int | None = None,
) -> dict[str, object]:
    sheet_row = conn.execute("SELECT id, site_id FROM sheets WHERE id = ?", (sheet_id,)).fetchone()
    if sheet_row is None:
        raise LookupError("sheet_not_found")

    validate_vendor_belongs_to_sheet(conn, sheet_id=sheet_id, vendor_name=vendor_name)

    if entry_id is not None:
        existing_entry = conn.execute(
            """
            SELECT id, sheet_id
            FROM vendor_work_entries
            WHERE id = ?
            """,
            (entry_id,),
        ).fetchone()
        if existing_entry is None:
            raise LookupError("entry_not_found")
        if int(existing_entry["sheet_id"] or 0) != int(sheet_id):
            raise LookupError("sheet_mismatch")

    return {
        "sheet_id": int(sheet_row["id"]),
        "site_id": int(sheet_row["site_id"]),
        "vendor_name": vendor_name,
        "entry_id": entry_id,
    }


def authorize_vendor_work_entry_write(
    conn: sqlite3.Connection,
    *,
    sheet_id: int,
    vendor_name: str,
    entry_id: int | None = None,
) -> dict[str, object]:
    context = resolve_vendor_work_entry_write_context(
        conn,
        sheet_id=sheet_id,
        vendor_name=vendor_name,
        entry_id=entry_id,
    )
    user = _current_internal_user()
    if user is None:
        raise LookupError("auth_required")
    if is_global_admin(user):
        return context

    current_site_id = _resolve_non_admin_read_site_id(conn, user)
    if int(context["site_id"]) != int(current_site_id):
        raise LookupError("write_target_not_in_current_site")
    return context


def _handle_vendor_work_entry_lookup_error(exc: LookupError):
    code = str(exc)
    if code == "sheet_not_found":
        return crew_api_error("sheet_not_found", "sheet_id was not found.", status=404)
    if code == "vendor_not_in_sheet":
        return crew_api_error("vendor_not_in_sheet", "vendor_name does not belong to the requested sheet.", status=404)
    if code == "entry_not_found":
        return crew_api_error("entry_not_found", "vendor work entry id was not found.", status=404)
    if code == "sheet_mismatch":
        return crew_api_error("sheet_mismatch", "vendor work entry belongs to a different sheet_id.", status=409)
    if code == "site_context_invalid":
        return crew_api_error("site_context_invalid", "current_site_id is missing or invalid.", status=403)
    if code == "site_permission_missing":
        return crew_api_error("site_permission_missing", "current user no longer has permission for the current site.", status=403)
    if code == "write_target_not_in_current_site":
        return crew_api_error("write_target_not_in_current_site", "write target does not belong to the current site.", status=403)
    if code == "auth_required":
        return crew_api_error("auth_required", "authentication is required.", status=403)
    raise exc


def resolve_vendor_work_entry_requirement_confirmation_context(
    conn: sqlite3.Connection,
    *,
    sheet_id: int,
    entry_id: int,
) -> dict[str, object]:
    sheet_row = conn.execute("SELECT id, site_id FROM sheets WHERE id = ?", (sheet_id,)).fetchone()
    if sheet_row is None:
        raise LookupError("sheet_not_found")

    existing_entry = conn.execute(
        """
        SELECT id, sheet_id, vendor_name, pre_entry_requirement, requirement_status,
               requirement_confirmed_by, requirement_confirmed_at
        FROM vendor_work_entries
        WHERE id = ?
        """,
        (entry_id,),
    ).fetchone()
    if existing_entry is None:
        raise LookupError("entry_not_found")
    if int(existing_entry["sheet_id"] or 0) != int(sheet_id):
        raise LookupError("sheet_mismatch")

    return {
        "sheet_id": int(sheet_row["id"]),
        "site_id": int(sheet_row["site_id"]),
        "entry_id": int(existing_entry["id"]),
        "vendor_name": str(existing_entry["vendor_name"] or ""),
        "pre_entry_requirement": str(existing_entry["pre_entry_requirement"] or ""),
        "requirement_status": str(existing_entry["requirement_status"] or "pending"),
        "requirement_confirmed_by": str(existing_entry["requirement_confirmed_by"] or ""),
        "requirement_confirmed_at": str(existing_entry["requirement_confirmed_at"] or ""),
    }


def authorize_vendor_work_entry_requirement_confirmation(
    conn: sqlite3.Connection,
    *,
    sheet_id: int,
    entry_id: int,
) -> dict[str, object]:
    context = resolve_vendor_work_entry_requirement_confirmation_context(
        conn,
        sheet_id=sheet_id,
        entry_id=entry_id,
    )
    user = _current_internal_user()
    if user is None:
        if current_vendor_account() is not None:
            raise LookupError("vendor_auth_forbidden")
        raise LookupError("auth_required")
    if is_global_admin(user):
        current_site_context = authorize_admin_site_scoped_write(conn, sheet_id=sheet_id)
        context["current_site_id"] = int(current_site_context["current_site_id"])
        return context

    current_site_id = _resolve_non_admin_read_site_id(conn, user)
    if int(context["site_id"]) != int(current_site_id):
        raise LookupError("write_target_not_in_current_site")
    context["current_site_id"] = int(current_site_id)
    return context


def _handle_vendor_work_entry_requirement_confirmation_lookup_error(exc: LookupError):
    code = str(exc)
    if code == "sheet_not_found":
        return crew_api_error("sheet_not_found", "sheet_id was not found.", status=404)
    if code == "entry_not_found":
        return crew_api_error("entry_not_found", "vendor work entry id was not found.", status=404)
    if code == "sheet_mismatch":
        return crew_api_error("sheet_mismatch", "vendor work entry belongs to a different sheet_id.", status=409)
    if code == "site_context_invalid":
        return crew_api_error("site_context_invalid", "current_site_id is missing or invalid.", status=403)
    if code == "site_permission_missing":
        return crew_api_error("site_permission_missing", "current user no longer has permission for the current site.", status=403)
    if code == "write_target_not_in_current_site":
        return crew_api_error("write_target_not_in_current_site", "write target does not belong to the current site.", status=403)
    if code == "vendor_auth_forbidden":
        return crew_api_error("vendor_auth_forbidden", "vendor authentication cannot confirm requirements.", status=403)
    if code == "auth_required":
        return crew_api_error("auth_required", "authentication is required.", status=403)
    raise exc


def _handle_vendor_write_preflight_error(exc: LookupError):
    code = str(exc)
    if code == "vendor_auth_required":
        return crew_api_error("vendor_auth_required", "vendor authentication is required.", status=403)
    if code == "sheet_not_found":
        return crew_api_error("sheet_not_found", "sheet_id was not found.", status=404)
    if code == "vendor_not_in_sheet":
        return crew_api_error("vendor_not_in_sheet", "trusted vendor_name does not belong to the requested sheet.", status=404)
    if code == "entry_not_found":
        return crew_api_error("entry_not_found", "vendor work entry id was not found.", status=404)
    if code == "sheet_mismatch":
        return crew_api_error("sheet_mismatch", "vendor work entry belongs to a different sheet_id.", status=409)
    if code == "vendor_name_mismatch":
        return crew_api_error("vendor_name_mismatch", "payload vendor_name does not match authenticated vendor identity.", status=403)
    if code == "vendor_cross_vendor_write_forbidden":
        return crew_api_error("vendor_cross_vendor_write_forbidden", "authenticated vendor cannot write another vendor's entry.", status=403)
    if code == "vendor_business_date_mismatch":
        return crew_api_error("vendor_business_date_mismatch", "payload business_date must match the existing vendor work entry business_date.", status=409)
    raise exc


def apply_vendor_work_entry_gate_state(entry: dict[str, object]) -> dict[str, object]:
    requirement_text = str(entry.get("pre_entry_requirement") or "").strip()
    requirement_status = str(entry.get("requirement_status") or "").strip()
    if not requirement_text:
        entry["readiness_state"] = "ready"
        entry["readiness_reason"] = "no_requirement"
    elif requirement_status == "confirmed":
        entry["readiness_state"] = "ready"
        entry["readiness_reason"] = "requirement_confirmed"
    else:
        entry["readiness_state"] = "not_ready"
        entry["readiness_reason"] = "requirement_pending"
    if entry["readiness_state"] == "ready":
        entry["scheduling_gate_state"] = "allowed"
    else:
        entry["scheduling_gate_state"] = "warning"
    entry["scheduling_gate_reason"] = entry["readiness_reason"]
    return entry


def apply_vendor_work_entry_formal_approval_state(entry: dict[str, object]) -> dict[str, object]:
    approval_status = str(entry.get("formal_approval_status") or "").strip()
    if approval_status == "approved":
        entry["formal_approval_state"] = "approved"
        entry["formal_approval_status"] = "approved"
        entry["formal_approved_by"] = str(entry.get("formal_approved_by") or "")
        entry["formal_approved_at"] = str(entry.get("formal_approved_at") or "")
        return entry

    entry["formal_approval_state"] = "pending"
    entry["formal_approval_status"] = "pending"
    entry["formal_approved_by"] = ""
    entry["formal_approved_at"] = ""
    return entry


def resolve_schedule_entry_context(
    conn: sqlite3.Connection,
    *,
    sheet_id: int,
    entry_id: int,
) -> dict[str, object]:
    sheet_row = conn.execute("SELECT id, site_id FROM sheets WHERE id = ?", (sheet_id,)).fetchone()
    if sheet_row is None:
        raise LookupError("sheet_not_found")

    existing_entry = conn.execute(
        """
        SELECT vwe.id,
               vwe.sheet_id,
               vwe.vendor_name,
               vwe.pre_entry_requirement,
               vwe.requirement_status,
               fa.approval_status AS formal_approval_status,
               fa.approved_by AS formal_approved_by,
               fa.approved_at AS formal_approved_at
        FROM vendor_work_entries vwe
        LEFT JOIN formal_approvals fa
               ON fa.entry_id = vwe.id
              AND fa.action = 'crew_formal_approve_entry'
        WHERE vwe.id = ?
        """,
        (entry_id,),
    ).fetchone()
    if existing_entry is None:
        raise LookupError("entry_not_found")
    if int(existing_entry["sheet_id"] or 0) != int(sheet_id):
        raise LookupError("sheet_mismatch")

    context = apply_vendor_work_entry_gate_state(
        {
            "sheet_id": int(sheet_row["id"]),
            "site_id": int(sheet_row["site_id"]),
            "entry_id": int(existing_entry["id"]),
            "vendor_name": str(existing_entry["vendor_name"] or ""),
            "pre_entry_requirement": str(existing_entry["pre_entry_requirement"] or ""),
            "requirement_status": str(existing_entry["requirement_status"] or "pending"),
            "formal_approval_status": str(existing_entry["formal_approval_status"] or ""),
            "formal_approved_by": str(existing_entry["formal_approved_by"] or ""),
            "formal_approved_at": str(existing_entry["formal_approved_at"] or ""),
        }
    )
    return apply_vendor_work_entry_formal_approval_state(context)


def resolve_vendor_work_entry_formal_approve_context(
    conn: sqlite3.Connection,
    *,
    sheet_id: int,
    entry_id: int,
) -> dict[str, object]:
    sheet_row = conn.execute("SELECT id, site_id FROM sheets WHERE id = ?", (sheet_id,)).fetchone()
    if sheet_row is None:
        raise LookupError("sheet_not_found")

    existing_entry = conn.execute(
        """
        SELECT id, sheet_id, vendor_name, pre_entry_requirement, requirement_status
        FROM vendor_work_entries
        WHERE id = ?
        """,
        (entry_id,),
    ).fetchone()
    if existing_entry is None:
        raise LookupError("entry_not_found")
    if int(existing_entry["sheet_id"] or 0) != int(sheet_id):
        raise LookupError("sheet_mismatch")

    context = apply_vendor_work_entry_gate_state(
        {
            "sheet_id": int(sheet_row["id"]),
            "site_id": int(sheet_row["site_id"]),
            "entry_id": int(existing_entry["id"]),
            "vendor_name": str(existing_entry["vendor_name"] or ""),
            "pre_entry_requirement": str(existing_entry["pre_entry_requirement"] or ""),
            "requirement_status": str(existing_entry["requirement_status"] or "pending"),
        }
    )
    return context


def authorize_vendor_work_entry_formal_approve(
    conn: sqlite3.Connection,
    *,
    sheet_id: int,
    entry_id: int,
) -> dict[str, object]:
    context = resolve_vendor_work_entry_formal_approve_context(
        conn,
        sheet_id=sheet_id,
        entry_id=entry_id,
    )
    user = _current_internal_user()
    if user is None:
        if current_vendor_account() is not None:
            raise LookupError("vendor_auth_forbidden")
        raise LookupError("auth_required")
    if is_global_admin(user):
        current_site_context = authorize_admin_site_scoped_write(conn, sheet_id=sheet_id)
        context["current_site_id"] = int(current_site_context["current_site_id"])
        return context

    current_site_id = _resolve_non_admin_read_site_id(conn, user)
    if int(context["site_id"]) != int(current_site_id):
        raise LookupError("write_target_not_in_current_site")
    context["current_site_id"] = int(current_site_id)
    return context


def authorize_schedule_entry_write(
    conn: sqlite3.Connection,
    *,
    sheet_id: int,
    entry_id: int,
) -> dict[str, object]:
    context = resolve_schedule_entry_context(
        conn,
        sheet_id=sheet_id,
        entry_id=entry_id,
    )
    user = _current_internal_user()
    if user is None:
        if current_vendor_account() is not None:
            raise LookupError("vendor_auth_forbidden")
        raise LookupError("auth_required")
    if is_global_admin(user):
        current_site_context = authorize_admin_site_scoped_write(conn, sheet_id=sheet_id)
        context["current_site_id"] = int(current_site_context["current_site_id"])
        return context

    current_site_id = _resolve_non_admin_read_site_id(conn, user)
    if int(context["site_id"]) != int(current_site_id):
        raise LookupError("write_target_not_in_current_site")
    context["current_site_id"] = int(current_site_id)
    return context


def _handle_vendor_work_entry_formal_approve_lookup_error(exc: LookupError):
    code = str(exc)
    if code == "sheet_not_found":
        return crew_api_error("sheet_not_found", "sheet_id was not found.", status=404)
    if code == "entry_not_found":
        return crew_api_error("entry_not_found", "vendor work entry id was not found.", status=404)
    if code == "sheet_mismatch":
        return crew_api_error("sheet_mismatch", "vendor work entry belongs to a different sheet_id.", status=409)
    if code == "site_context_invalid":
        return crew_api_error("site_context_invalid", "current_site_id is missing or invalid.", status=403)
    if code == "site_permission_missing":
        return crew_api_error("site_permission_missing", "current user no longer has permission for the current site.", status=403)
    if code == "write_target_not_in_current_site":
        return crew_api_error("write_target_not_in_current_site", "write target does not belong to the current site.", status=403)
    if code == "vendor_auth_forbidden":
        return crew_api_error("vendor_auth_forbidden", "vendor authentication cannot approve entries.", status=403)
    if code == "auth_required":
        return crew_api_error("auth_required", "authentication is required.", status=403)
    raise exc


def _handle_schedule_entry_lookup_error(exc: LookupError):
    code = str(exc)
    if code == "sheet_not_found":
        return crew_api_error("sheet_not_found", "sheet_id was not found.", status=404)
    if code == "entry_not_found":
        return crew_api_error("entry_not_found", "vendor work entry id was not found.", status=404)
    if code == "sheet_mismatch":
        return crew_api_error("sheet_mismatch", "vendor work entry belongs to a different sheet_id.", status=409)
    if code == "site_context_invalid":
        return crew_api_error("site_context_invalid", "current_site_id is missing or invalid.", status=403)
    if code == "site_permission_missing":
        return crew_api_error("site_permission_missing", "current user no longer has permission for the current site.", status=403)
    if code == "write_target_not_in_current_site":
        return crew_api_error("write_target_not_in_current_site", "write target does not belong to the current site.", status=403)
    if code == "vendor_auth_forbidden":
        return crew_api_error("vendor_auth_forbidden", "vendor authentication cannot create schedules.", status=403)
    if code == "auth_required":
        return crew_api_error("auth_required", "authentication is required.", status=403)
    raise exc


def get_active_crew_vendors(sheet_id: int) -> list[str]:
    with db() as conn:
        tasks = conn.execute(
            """
            SELECT id, vendor
            FROM tasks
            WHERE sheet_id = ?
            ORDER BY col_index, id
            """,
            (sheet_id,),
        ).fetchall()
        active_task_ids = {
            row["task_id"]
            for row in conn.execute(
                """
                SELECT DISTINCT p.task_id
                FROM progress p
                JOIN tasks t ON t.id = p.task_id
                WHERE t.sheet_id = ? AND p.value = ?
                """,
                (sheet_id, WORKING_VALUE),
            ).fetchall()
        }

    seen: set[str] = set()
    active_vendors: list[str] = []
    for task in tasks:
        vendor_name = (task["vendor"] or "").strip()
        if not vendor_name or task["id"] not in active_task_ids or vendor_name in seen:
            continue
        seen.add(vendor_name)
        active_vendors.append(vendor_name)
    return active_vendors


def get_pending_items_by_vendor(sheet_id: int) -> dict[str, list[str]]:
    with db() as conn:
        tasks = conn.execute(
            """
            SELECT id, vendor, name
            FROM tasks
            WHERE sheet_id = ?
            ORDER BY col_index, id
            """,
            (sheet_id,),
        ).fetchall()
        active_task_ids = {
            row["task_id"]
            for row in conn.execute(
                """
                SELECT DISTINCT p.task_id
                FROM progress p
                JOIN tasks t ON t.id = p.task_id
                WHERE t.sheet_id = ? AND p.value = ?
                """,
                (sheet_id, WORKING_VALUE),
            ).fetchall()
        }

    pending_items: dict[str, list[str]] = {}
    for task in tasks:
        vendor_name = (task["vendor"] or "").strip()
        task_name = str(task["name"] or "").strip()
        if not vendor_name or not task_name or task["id"] not in active_task_ids:
            continue
        pending_items.setdefault(vendor_name, []).append(task_name)
    return pending_items


def build_contact_display_name(contact_or_fields) -> str:
    if isinstance(contact_or_fields, sqlite3.Row):
        contact_title = str(contact_or_fields["contact_title"] or "").strip()
        contact_name = str(contact_or_fields["contact_name"] or "").strip()
    else:
        contact_title = str((contact_or_fields or {}).get("contact_title", "") or "").strip()
        contact_name = str((contact_or_fields or {}).get("contact_name", "") or "").strip()
    if contact_title and contact_name:
        return f"{contact_title} {contact_name}"
    return contact_title or contact_name


def serialize_vendor_contact(row: sqlite3.Row) -> dict[str, object]:
    return {
        "id": row["id"],
        "sheet_id": row["sheet_id"],
        "vendor_name": row["vendor_name"],
        "contact_name": row["contact_name"],
        "contact_title": row["contact_title"],
        "contact_phone": row["contact_phone"],
        "display_name": build_contact_display_name(row),
        "is_primary": row["is_primary"],
        "contact_order": row["contact_order"],
        "created_at": row["created_at"],
        "updated_at": row["updated_at"],
    }


def empty_vendor_contact(sheet_id: int, vendor_name: str) -> dict[str, object]:
    return {
        "id": None,
        "sheet_id": sheet_id,
        "vendor_name": vendor_name,
        "contact_name": "",
        "contact_title": "",
        "contact_phone": "",
        "display_name": "",
        "is_primary": 0,
        "contact_order": 0,
        "created_at": "",
        "updated_at": "",
    }


def vendor_contact_order_clause() -> str:
    return "is_primary DESC, contact_order ASC, id ASC"


def fetch_vendor_contacts_grouped(
    conn: sqlite3.Connection, sheet_id: int
) -> dict[str, list[dict[str, object]]]:
    rows = conn.execute(
        f"""
        SELECT id, sheet_id, vendor_name, contact_name, contact_title, contact_phone,
               is_primary, contact_order, created_at, updated_at
        FROM vendor_contacts
        WHERE sheet_id = ?
        ORDER BY vendor_name, {vendor_contact_order_clause()}
        """,
        (sheet_id,),
    ).fetchall()
    contacts_by_vendor: dict[str, list[dict[str, object]]] = {}
    for row in rows:
        contacts_by_vendor.setdefault(row["vendor_name"], []).append(serialize_vendor_contact(row))
    return contacts_by_vendor


def first_or_empty_vendor_contact(
    contacts: list[dict[str, object]] | None,
    *,
    sheet_id: int,
    vendor_name: str,
) -> dict[str, object]:
    if contacts:
        return dict(contacts[0])
    return empty_vendor_contact(sheet_id, vendor_name)


def fetch_vendor_contact_map(conn: sqlite3.Connection, sheet_id: int) -> dict[str, dict[str, object]]:
    contact_map: dict[str, dict[str, object]] = {}
    for vendor_name, contacts in fetch_vendor_contacts_grouped(conn, sheet_id).items():
        contact_map[vendor_name] = first_or_empty_vendor_contact(contacts, sheet_id=sheet_id, vendor_name=vendor_name)
    return contact_map


def fetch_vendor_contacts_by_vendor(
    conn: sqlite3.Connection,
    *,
    sheet_id: int,
    vendor_name: str,
) -> list[dict[str, object]]:
    rows = conn.execute(
        f"""
        SELECT id, sheet_id, vendor_name, contact_name, contact_title, contact_phone,
               is_primary, contact_order, created_at, updated_at
        FROM vendor_contacts
        WHERE sheet_id = ? AND vendor_name = ?
        ORDER BY {vendor_contact_order_clause()}
        """,
        (sheet_id, vendor_name),
    ).fetchall()
    return [serialize_vendor_contact(row) for row in rows]


def next_contact_order(conn: sqlite3.Connection, *, sheet_id: int, vendor_name: str) -> int:
    row = conn.execute(
        """
        SELECT COALESCE(MAX(contact_order), -1) AS max_contact_order
        FROM vendor_contacts
        WHERE sheet_id = ? AND vendor_name = ?
        """,
        (sheet_id, vendor_name),
    ).fetchone()
    return int(row["max_contact_order"]) + 1


def set_primary_contact(
    conn: sqlite3.Connection,
    *,
    sheet_id: int,
    vendor_name: str,
    contact_id: int,
) -> None:
    conn.execute(
        """
        UPDATE vendor_contacts
        SET is_primary = 0, updated_at = CURRENT_TIMESTAMP
        WHERE sheet_id = ? AND vendor_name = ? AND id != ?
        """,
        (sheet_id, vendor_name, contact_id),
    )
    conn.execute(
        """
        UPDATE vendor_contacts
        SET is_primary = 1, updated_at = CURRENT_TIMESTAMP
        WHERE id = ? AND sheet_id = ? AND vendor_name = ?
        """,
        (contact_id, sheet_id, vendor_name),
    )


def fetch_vendor_work_entries(
    conn: sqlite3.Connection,
    *,
    sheet_id: int,
    business_date: str,
) -> dict[str, list[dict[str, object]]]:
    rows = conn.execute(
        """
        SELECT vwe.id, vwe.sheet_id, vwe.vendor_name, vwe.business_date, vwe.planned_at, vwe.planned_headcount,
               vwe.actual_headcount, vwe.work_content, vwe.pre_entry_requirement, vwe.requirement_status,
               vwe.requirement_confirmed_by, vwe.requirement_confirmed_at, vwe.work_headcount,
               vwe.entry_order, vwe.created_at, vwe.updated_at,
               fa.approval_status AS formal_approval_status,
               fa.approved_by AS formal_approved_by,
               fa.approved_at AS formal_approved_at
        FROM vendor_work_entries vwe
        LEFT JOIN formal_approvals fa
               ON fa.entry_id = vwe.id
              AND fa.action = 'crew_formal_approve_entry'
        WHERE vwe.sheet_id = ? AND vwe.business_date = ?
        ORDER BY vwe.vendor_name, vwe.entry_order, vwe.id
        """,
        (sheet_id, business_date),
    ).fetchall()
    entries_by_vendor: dict[str, list[dict[str, object]]] = {}
    for row in rows:
        entry = apply_vendor_work_entry_gate_state(dict(row))
        entry = apply_vendor_work_entry_formal_approval_state(entry)
        entries_by_vendor.setdefault(row["vendor_name"], []).append(entry)
    return entries_by_vendor


def fetch_dashboard_scheduled_entries(
    conn: sqlite3.Connection,
    *,
    sheet_id: int,
) -> list[dict[str, object]]:
    rows = conn.execute(
        """
        SELECT se.id AS schedule_id,
               se.entry_id,
               se.sheet_id,
               se.action AS schedule_action,
               se.schedule_status,
               se.scheduled_date,
               se.scheduled_time,
               se.scheduled_by,
               se.scheduled_at,
               vwe.vendor_name,
               vwe.business_date,
               vwe.planned_at,
               vwe.planned_headcount,
               vwe.actual_headcount,
               vwe.work_content,
               vwe.pre_entry_requirement,
               vwe.requirement_status,
               vwe.requirement_confirmed_by,
               vwe.requirement_confirmed_at,
               vwe.work_headcount,
               vwe.entry_order,
               vwe.created_at,
               vwe.updated_at,
               fa.approval_status AS formal_approval_status,
               fa.approved_by AS formal_approved_by,
               fa.approved_at AS formal_approved_at
        FROM scheduling_entries se
        JOIN vendor_work_entries vwe
          ON vwe.id = se.entry_id
        LEFT JOIN formal_approvals fa
          ON fa.entry_id = vwe.id
         AND fa.action = 'crew_formal_approve_entry'
        WHERE se.sheet_id = ?
        ORDER BY se.scheduled_date, se.scheduled_time, se.id
        """,
        (sheet_id,),
    ).fetchall()
    scheduled_entries: list[dict[str, object]] = []
    for row in rows:
        entry = {
            "id": int(row["entry_id"]),
            "sheet_id": int(row["sheet_id"]),
            "vendor_name": str(row["vendor_name"] or ""),
            "business_date": str(row["business_date"] or ""),
            "planned_at": str(row["planned_at"] or ""),
            "planned_headcount": int(row["planned_headcount"] or 0),
            "actual_headcount": int(row["actual_headcount"] or 0),
            "work_content": str(row["work_content"] or ""),
            "pre_entry_requirement": str(row["pre_entry_requirement"] or ""),
            "requirement_status": str(row["requirement_status"] or ""),
            "requirement_confirmed_by": str(row["requirement_confirmed_by"] or ""),
            "requirement_confirmed_at": str(row["requirement_confirmed_at"] or ""),
            "work_headcount": int(row["work_headcount"] or 0),
            "entry_order": int(row["entry_order"] or 0),
            "created_at": str(row["created_at"] or ""),
            "updated_at": str(row["updated_at"] or ""),
            "formal_approval_status": str(row["formal_approval_status"] or ""),
            "formal_approved_by": str(row["formal_approved_by"] or ""),
            "formal_approved_at": str(row["formal_approved_at"] or ""),
            "schedule_id": int(row["schedule_id"]),
            "schedule_action": str(row["schedule_action"] or ""),
            "schedule_status": str(row["schedule_status"] or ""),
            "scheduled_date": str(row["scheduled_date"] or ""),
            "scheduled_time": str(row["scheduled_time"] or ""),
            "scheduled_by": str(row["scheduled_by"] or ""),
            "scheduled_at": str(row["scheduled_at"] or ""),
        }
        entry = apply_vendor_work_entry_gate_state(entry)
        entry = apply_vendor_work_entry_formal_approval_state(entry)
        scheduled_entries.append(entry)
    return scheduled_entries


def build_dashboard_payload(
    conn: sqlite3.Connection,
    *,
    sheet_id: int,
    business_date: str,
) -> dict[str, object]:
    entries_by_vendor = fetch_vendor_work_entries(conn, sheet_id=sheet_id, business_date=business_date)
    scheduled_entries = fetch_dashboard_scheduled_entries(conn, sheet_id=sheet_id)
    today_schedule = [
        entry.copy()
        for entry in scheduled_entries
        if str(entry.get("scheduled_date")) == business_date
    ]
    today_entries = [
        entry.copy()
        for vendor_entries in entries_by_vendor.values()
        for entry in vendor_entries
    ]
    blocked_items = [entry.copy() for entry in today_entries if str(entry.get("scheduling_gate_state")) == "warning"]
    pending_approvals = [
        entry.copy()
        for entry in today_entries
        if str(entry.get("scheduling_gate_state")) == "allowed"
        and str(entry.get("formal_approval_state")) == "pending"
    ]
    pending_requirements = [
        entry.copy()
        for entry in today_entries
        if str(entry.get("readiness_reason")) == "requirement_pending"
    ]
    approved_today_count = sum(1 for entry in today_entries if str(entry.get("formal_approval_state")) == "approved")

    return {
        "summary": {
            "blocked_count": len(blocked_items),
            "pending_approval_count": len(pending_approvals),
            "pending_requirement_count": len(pending_requirements),
            "today_entry_count": len(today_entries),
            "approved_today_count": approved_today_count,
            "scheduled_count": len(scheduled_entries),
            "today_schedule_count": len(today_schedule),
        },
        "blocked_items": blocked_items,
        "pending_approvals": pending_approvals,
        "pending_requirements": pending_requirements,
        "today_entries": today_entries,
        "scheduled_entries": [entry.copy() for entry in scheduled_entries],
        "today_schedule": today_schedule,
        "quick_actions": [
            {
                "action": "review_blocked_items",
                "label": "Blocked Items",
                "href": f"/sheet?sheet_id={int(sheet_id)}",
                "count": len(blocked_items),
            },
            {
                "action": "review_pending_approvals",
                "label": "Pending Formal Approval",
                "href": f"/sheet?sheet_id={int(sheet_id)}",
                "count": len(pending_approvals),
            },
            {
                "action": "review_pending_requirements",
                "label": "Pending Requirement Confirmation",
                "href": f"/sheet?sheet_id={int(sheet_id)}",
                "count": len(pending_requirements),
            },
        ],
    }


def build_scheduling_payload(
    conn: sqlite3.Connection,
    *,
    sheet_id: int,
    business_date: str,
) -> dict[str, object]:
    entries_by_vendor = fetch_vendor_work_entries(conn, sheet_id=sheet_id, business_date=business_date)
    today_entries = [
        entry.copy()
        for vendor_entries in entries_by_vendor.values()
        for entry in vendor_entries
    ]
    schedulable_entries = [
        entry.copy()
        for entry in today_entries
        if str(entry.get("formal_approval_state")) == "approved"
        and str(entry.get("scheduling_gate_state")) == "allowed"
    ]
    blocked_entries = [
        entry.copy()
        for entry in today_entries
        if str(entry.get("formal_approval_state")) != "approved"
        or str(entry.get("scheduling_gate_state")) == "warning"
    ]
    scheduled_entries: list[dict[str, object]] = []
    unscheduled_entries = [entry.copy() for entry in today_entries]

    return {
        "summary": {
            "schedulable_count": len(schedulable_entries),
            "blocked_count": len(blocked_entries),
            "unscheduled_count": len(unscheduled_entries),
        },
        "schedulable_entries": schedulable_entries,
        "blocked_entries": blocked_entries,
        "scheduled_entries": scheduled_entries,
        "unscheduled_entries": unscheduled_entries,
    }


def build_work_hub_runtime_payload(
    conn: sqlite3.Connection,
    *,
    sheet_id: int,
    business_date: str,
) -> dict[str, object]:
    dashboard_payload = build_dashboard_payload(
        conn,
        sheet_id=sheet_id,
        business_date=business_date,
    )
    scheduling_payload = build_scheduling_payload(
        conn,
        sheet_id=sheet_id,
        business_date=business_date,
    )
    dashboard_summary = dict(dashboard_payload.get("summary") or {})
    scheduling_summary = dict(scheduling_payload.get("summary") or {})

    return {
        "sheet_id": int(sheet_id),
        "business_date": business_date,
        "dashboard": dashboard_payload,
        "scheduling": scheduling_payload,
        "work_hub": {
            "summary": {
                "blocked_count": int(scheduling_summary.get("blocked_count") or 0),
                "schedulable_count": int(scheduling_summary.get("schedulable_count") or 0),
                "pending_approval_count": int(dashboard_summary.get("pending_approval_count") or 0),
                "pending_requirement_count": int(dashboard_summary.get("pending_requirement_count") or 0),
                "today_entry_count": int(dashboard_summary.get("today_entry_count") or 0),
                "scheduled_count": int(dashboard_summary.get("scheduled_count") or 0),
                "today_schedule_count": int(dashboard_summary.get("today_schedule_count") or 0),
            },
            "blocked_entries": [entry.copy() for entry in scheduling_payload.get("blocked_entries", [])],
            "schedulable_entries": [entry.copy() for entry in scheduling_payload.get("schedulable_entries", [])],
            "today_entries": [entry.copy() for entry in dashboard_payload.get("today_entries", [])],
            "scheduled_entries": [entry.copy() for entry in dashboard_payload.get("scheduled_entries", [])],
            "today_schedule": [entry.copy() for entry in dashboard_payload.get("today_schedule", [])],
        },
    }


def _management_read_model_count(summary: dict[str, object], key: str) -> int:
    return int(summary.get(key) or 0)


def _management_read_model_entries(payload: dict[str, object], key: str) -> list[dict[str, object]]:
    return [entry.copy() for entry in payload.get(key, [])]


def _management_read_model_entry_ids(entries: list[dict[str, object]]) -> list[int]:
    return [int(entry["id"]) for entry in entries if entry.get("id") is not None]


def _management_read_model_drilldown_ref(target: str, count: int) -> dict[str, object]:
    return {
        "target": target,
        "count": int(count),
    }


def build_management_read_model_payload(
    conn: sqlite3.Connection,
    *,
    sheet_id: int,
    business_date: str,
) -> dict[str, object]:
    dashboard_payload = build_dashboard_payload(
        conn,
        sheet_id=sheet_id,
        business_date=business_date,
    )
    scheduling_payload = build_scheduling_payload(
        conn,
        sheet_id=sheet_id,
        business_date=business_date,
    )
    dashboard_summary = dict(dashboard_payload.get("summary") or {})
    scheduling_summary = dict(scheduling_payload.get("summary") or {})

    blocked_entry_refs = _management_read_model_entries(scheduling_payload, "blocked_entries")
    schedulable_entry_refs = _management_read_model_entries(scheduling_payload, "schedulable_entries")
    scheduled_entry_refs = _management_read_model_entries(dashboard_payload, "scheduled_entries")
    today_schedule_entry_refs = _management_read_model_entries(dashboard_payload, "today_schedule")
    pending_approval_refs = _management_read_model_entries(dashboard_payload, "pending_approvals")
    pending_requirement_refs = _management_read_model_entries(dashboard_payload, "pending_requirements")

    return {
        "management_summary": {
            "today_entry_count": _management_read_model_count(dashboard_summary, "today_entry_count"),
            "scheduled_count": _management_read_model_count(dashboard_summary, "scheduled_count"),
            "today_schedule_count": _management_read_model_count(dashboard_summary, "today_schedule_count"),
            "schedulable_count": _management_read_model_count(scheduling_summary, "schedulable_count"),
            "blocked_count": _management_read_model_count(scheduling_summary, "blocked_count"),
            "pending_approval_count": _management_read_model_count(dashboard_summary, "pending_approval_count"),
            "pending_requirement_count": _management_read_model_count(dashboard_summary, "pending_requirement_count"),
        },
        "scheduling_overview": {
            "scheduled_count": _management_read_model_count(dashboard_summary, "scheduled_count"),
            "today_schedule_count": _management_read_model_count(dashboard_summary, "today_schedule_count"),
            "schedulable_count": _management_read_model_count(scheduling_summary, "schedulable_count"),
            "blocked_count": _management_read_model_count(scheduling_summary, "blocked_count"),
            "scheduled_entry_ids": _management_read_model_entry_ids(scheduled_entry_refs),
            "today_schedule_entry_ids": _management_read_model_entry_ids(today_schedule_entry_refs),
            "schedulable_entry_ids": _management_read_model_entry_ids(schedulable_entry_refs),
            "blocked_entry_ids": _management_read_model_entry_ids(blocked_entry_refs),
        },
        "approval_overview": {
            "pending_approval_count": _management_read_model_count(dashboard_summary, "pending_approval_count"),
            "approved_today_count": _management_read_model_count(dashboard_summary, "approved_today_count"),
            "pending_approval_entry_ids": _management_read_model_entry_ids(pending_approval_refs),
        },
        "requirement_overview": {
            "pending_requirement_count": _management_read_model_count(dashboard_summary, "pending_requirement_count"),
            "pending_requirement_entry_ids": _management_read_model_entry_ids(pending_requirement_refs),
        },
        "operational_risk_overview": {
            "blocked_count": _management_read_model_count(scheduling_summary, "blocked_count"),
            "pending_approval_count": _management_read_model_count(dashboard_summary, "pending_approval_count"),
            "pending_requirement_count": _management_read_model_count(dashboard_summary, "pending_requirement_count"),
            "blocked_entry_ids": _management_read_model_entry_ids(blocked_entry_refs),
            "pending_approval_entry_ids": _management_read_model_entry_ids(pending_approval_refs),
            "pending_requirement_entry_ids": _management_read_model_entry_ids(pending_requirement_refs),
        },
        "drilldown_refs": {
            "blocked": _management_read_model_drilldown_ref(
                "blocked",
                _management_read_model_count(scheduling_summary, "blocked_count"),
            ),
            "schedulable": _management_read_model_drilldown_ref(
                "schedulable",
                _management_read_model_count(scheduling_summary, "schedulable_count"),
            ),
            "scheduled": _management_read_model_drilldown_ref(
                "scheduled",
                _management_read_model_count(dashboard_summary, "scheduled_count"),
            ),
            "pending_approval": _management_read_model_drilldown_ref(
                "pending-approval",
                _management_read_model_count(dashboard_summary, "pending_approval_count"),
            ),
            "pending_requirement": _management_read_model_drilldown_ref(
                "pending-requirement",
                _management_read_model_count(dashboard_summary, "pending_requirement_count"),
            ),
            "today_entries": _management_read_model_drilldown_ref(
                "today-entries",
                _management_read_model_count(dashboard_summary, "today_entry_count"),
            ),
            "today_schedule": _management_read_model_drilldown_ref(
                "today-schedule",
                _management_read_model_count(dashboard_summary, "today_schedule_count"),
            ),
        },
    }


class AIReadOrchestrationError(RuntimeError):
    def __init__(self, code: str):
        self.code = code
        super().__init__(code)


def build_role_scoped_ai_read_result(
    conn: sqlite3.Connection,
    *,
    intent_key: str,
    sheet_id: int,
    business_date: str | None = None,
    as_of: datetime | None = None,
) -> dict[str, object]:
    intent = get_supported_ai_read_intent(intent_key)
    if intent is None:
        raise AIReadOrchestrationError("unsupported_ai_read_intent")
    scope = authorize_dashboard_read(conn, sheet_id=sheet_id)
    if intent_key != "today_formally_scheduled_count":
        raise AIReadOrchestrationError("ai_read_intent_not_implemented")

    trusted_as_of = as_of if as_of is not None else datetime.now()
    try:
        resolved_business_date = (
            resolve_crew_business_date(now=trusted_as_of)
            if business_date is None
            else parse_crew_business_date(business_date)
        )
    except ValueError:
        raise AIReadOrchestrationError("invalid_business_date") from None

    actor = _current_internal_user()
    actor_role = str(actor["role"])
    actor_scope = "current_site_constrained_admin" if is_global_admin(actor) else "site_member"
    if actor_scope not in intent["actor_scope"]:
        raise AIReadOrchestrationError("ai_read_intent_not_implemented")

    payload = build_dashboard_payload(
        conn,
        sheet_id=int(scope["sheet_id"]),
        business_date=resolved_business_date,
    )
    if not isinstance(payload, Mapping):
        raise AIReadOrchestrationError("ai_read_source_shape_mismatch")
    summary = payload.get("summary")
    if not isinstance(summary, Mapping):
        raise AIReadOrchestrationError("ai_read_source_shape_mismatch")
    count = summary.get("today_schedule_count")
    if isinstance(count, bool) or not isinstance(count, int) or count < 0:
        raise AIReadOrchestrationError("ai_read_source_shape_mismatch")

    evidence = {
        "actor_identity_type": "internal",
        "actor_role": actor_role,
        "actor_scope": actor_scope,
        "current_site_id": int(scope["current_site_id"]),
        "sheet_id": int(scope["sheet_id"]),
        "intent_key": intent_key,
        "source_family": intent["source_family"],
        "business_date": resolved_business_date,
        "authorization_policy": intent["authorization_policy"],
        "as_of": trusted_as_of.isoformat(),
        "item_count": count,
        "empty_state_reason": "no_formal_schedule_for_business_date" if count == 0 else None,
    }
    required_evidence_fields = intent.get("evidence_fields")
    if not isinstance(required_evidence_fields, (list, tuple)) or not set(required_evidence_fields).issubset(
        evidence
    ):
        raise AIReadOrchestrationError("ai_read_source_shape_mismatch")

    return {
        "mode": "read_only",
        "intent_key": intent_key,
        "scope": {
            "current_site_id": int(scope["current_site_id"]),
            "sheet_id": int(scope["sheet_id"]),
            "business_date": resolved_business_date,
        },
        "evidence": evidence,
        "total_count": count,
        "returned_count": count,
        "truncated": False,
        "count": count,
    }


def parse_planned_at_datetime(value: str) -> datetime | None:
    raw = value.strip()
    if not raw:
        return None
    for fmt in ("%Y-%m-%d %H:%M", "%Y-%m-%dT%H:%M", "%Y-%m-%d %H:%M:%S", "%Y-%m-%dT%H:%M:%S"):
        try:
            return datetime.strptime(raw, fmt)
        except ValueError:
            continue
    return None


def parse_schedule_date(value: object) -> str:
    scheduled_date = str(value or "").strip()
    if not scheduled_date:
        raise ValueError("scheduled_date is required.")
    try:
        datetime.strptime(scheduled_date, "%Y-%m-%d")
    except ValueError as exc:
        raise ValueError("scheduled_date must be in YYYY-MM-DD format.") from exc
    return scheduled_date


def parse_schedule_time(value: object) -> str:
    scheduled_time = str(value or "").strip()
    if not scheduled_time:
        raise ValueError("scheduled_time is required.")
    try:
        datetime.strptime(scheduled_time, "%H:%M")
    except ValueError as exc:
        raise ValueError("scheduled_time must be in HH:MM format.") from exc
    return scheduled_time


def login_required(fn):
    @wraps(fn)
    def wrapper(*args, **kwargs):
        if not session.get("user_id"):
            return redirect(url_for("login"))
        return fn(*args, **kwargs)

    return wrapper


def admin_required(fn):
    @wraps(fn)
    def wrapper(*args, **kwargs):
        if not session.get("user_id"):
            return redirect(url_for("login"))
        if session.get("role") != "admin":
            flash("需要管理員權限。", "error")
            return redirect(url_for("sheet"))
        return fn(*args, **kwargs)

    return wrapper


def clear_vendor_session() -> None:
    session.pop("vendor_account_id", None)
    session.pop("vendor_username", None)
    session.pop("vendor_name", None)
    if session.get("identity_type") == "vendor":
        session.pop("identity_type", None)


def is_vendor_session() -> bool:
    return (
        session.get("identity_type") == "vendor"
        and session.get("vendor_account_id") is not None
        and session.get("vendor_username") is not None
        and session.get("vendor_name") is not None
    )


def current_vendor_account() -> dict[str, object] | None:
    if not is_vendor_session():
        return None
    return {
        "id": int(session["vendor_account_id"]),
        "username": str(session["vendor_username"]),
        "vendor_name": str(session["vendor_name"]),
    }


def current_vendor_scope() -> dict[str, object] | None:
    vendor_account = current_vendor_account()
    if vendor_account is None:
        return None
    return {
        "identity_type": "vendor",
        "vendor_account_id": int(vendor_account["id"]),
        "vendor_username": str(vendor_account["username"]),
        "vendor_name": str(vendor_account["vendor_name"]),
        "scope_type": "vendor_identity_only",
        "scope_version": 1,
    }


def set_vendor_session(vendor_account) -> None:
    session.clear()
    session["identity_type"] = "vendor"
    session["vendor_account_id"] = int(vendor_account["id"])
    session["vendor_username"] = str(vendor_account["username"])
    session["vendor_name"] = str(vendor_account["vendor_name"])


def verify_vendor_account(username: str, password: str) -> sqlite3.Row | None:
    vendor_account = get_vendor_account_by_username(username.strip())
    if vendor_account is None:
        return None
    if int(vendor_account["is_active"] or 0) != 1:
        return None
    if not check_password_hash(vendor_account["password_hash"], password):
        return None
    return vendor_account


def vendor_login_required(fn):
    @wraps(fn)
    def wrapper(*args, **kwargs):
        if current_vendor_account() is None:
            clear_vendor_session()
            return redirect(url_for("vendor_login"))
        return fn(*args, **kwargs)

    return wrapper


def require_current_vendor_account() -> dict[str, object]:
    vendor_account = current_vendor_account()
    if vendor_account is None:
        raise LookupError("vendor_auth_required")
    return vendor_account


def require_current_vendor_business_identity() -> dict[str, object]:
    vendor_account = require_current_vendor_account()
    return {
        "vendor_account_id": int(vendor_account["id"]),
        "vendor_username": str(vendor_account["username"]),
        "vendor_name": str(vendor_account["vendor_name"]),
    }


def vendor_write_preflight(
    conn: sqlite3.Connection,
    *,
    sheet_id: int,
    business_date: str,
    entry_id: int | None = None,
    payload_vendor_name: str | None = None,
    payload_business_date_provided: bool = False,
) -> dict[str, object]:
    business_identity = require_current_vendor_business_identity()
    trusted_vendor_name = str(business_identity["vendor_name"])
    if payload_vendor_name is not None and str(payload_vendor_name) != trusted_vendor_name:
        raise LookupError("vendor_name_mismatch")

    context = resolve_vendor_work_entry_write_context(
        conn,
        sheet_id=sheet_id,
        vendor_name=trusted_vendor_name,
        entry_id=entry_id,
    )
    if entry_id is not None:
        existing_entry = conn.execute(
            """
            SELECT vendor_name, business_date
            FROM vendor_work_entries
            WHERE id = ?
            """,
            (entry_id,),
        ).fetchone()
        if existing_entry is None:
            raise LookupError("entry_not_found")
        if str(existing_entry["vendor_name"]) != trusted_vendor_name:
            raise LookupError("vendor_cross_vendor_write_forbidden")
        if payload_business_date_provided and str(existing_entry["business_date"]) != str(business_date):
            raise LookupError("vendor_business_date_mismatch")

    return {
        "vendor_account_id": int(business_identity["vendor_account_id"]),
        "vendor_username": str(business_identity["vendor_username"]),
        "vendor_name": trusted_vendor_name,
        "sheet_id": int(context["sheet_id"]),
        "business_date": str(business_date),
        "entry_id": entry_id,
        "write_mode": "update" if entry_id is not None else "create",
    }


def build_vendor_work_entry_preflight_payload(
    conn: sqlite3.Connection,
    *,
    sheet_id: int,
    business_date: str,
    entry_id: int | None = None,
    payload_vendor_name: str | None = None,
    payload_business_date_provided: bool = False,
) -> dict[str, object]:
    return {
        "ok": True,
        "preflight": vendor_write_preflight(
            conn,
            sheet_id=sheet_id,
            business_date=business_date,
            entry_id=entry_id,
            payload_vendor_name=payload_vendor_name,
            payload_business_date_provided=payload_business_date_provided,
        ),
    }


def find_vendor_work_entry_preflight_sheet_id(
    conn: sqlite3.Connection,
    *,
    vendor_name: str,
) -> int | None:
    row = conn.execute(
        """
        SELECT DISTINCT sheet_id
        FROM tasks
        WHERE TRIM(COALESCE(vendor, '')) = ?
        ORDER BY sheet_id
        LIMIT 1
        """,
        (vendor_name,),
    ).fetchone()
    if row is None:
        return None
    return int(row["sheet_id"])


def authorize_vendor_business_read() -> dict[str, object]:
    return require_current_vendor_business_identity()


def fetch_vendor_business_read_preview(
    conn: sqlite3.Connection,
    *,
    vendor_name: str,
) -> list[sqlite3.Row]:
    return conn.execute(
        """
        SELECT id, vendor_name, business_date, planned_at, planned_headcount,
               actual_headcount, work_content, pre_entry_requirement, work_headcount, entry_order
        FROM vendor_work_entries
        WHERE vendor_name = ?
        ORDER BY business_date DESC, entry_order ASC, rowid ASC
        """,
        (vendor_name,),
    ).fetchall()


def serialize_vendor_business_read_preview(
    *,
    business_identity: dict[str, object],
    rows: list[sqlite3.Row],
) -> dict[str, object]:
    vendor_name = str(business_identity["vendor_name"])
    entries = [serialize_vendor_business_read_entry(row) for row in rows]
    business_dates = sorted({entry["business_date"] for entry in entries}, reverse=True)
    return {
        "ok": True,
        "vendor_account_id": int(business_identity["vendor_account_id"]),
        "vendor_username": str(business_identity["vendor_username"]),
        "vendor_name": vendor_name,
        "entry_count": len(entries),
        "business_dates": business_dates,
        "entries": entries,
    }


def serialize_vendor_business_read_entry(row: sqlite3.Row) -> dict[str, object]:
    return {
        "entry_id": int(row["id"]),
        "vendor_name": str(row["vendor_name"]),
        "business_date": str(row["business_date"]),
        "planned_at": str(row["planned_at"] or ""),
        "planned_headcount": int(row["planned_headcount"] or 0),
        "actual_headcount": int(row["actual_headcount"] or 0),
        "work_content": str(row["work_content"] or ""),
        "pre_entry_requirement": str(row["pre_entry_requirement"] or ""),
        "work_headcount": int(row["work_headcount"] or 0),
        "entry_order": int(row["entry_order"] or 0),
    }


def resolve_vendor_today_entry_selection(
    preview_entries: list[dict[str, object]],
    business_date: str,
    today_entry_index_raw: object,
    selected_entry_id_raw: object = None,
) -> dict[str, object]:
    today_entries = [
        entry
        for entry in preview_entries
        if str(entry.get("business_date", "")) == str(business_date)
    ]
    selected_entry_id = None
    try:
        if str(selected_entry_id_raw or "").strip():
            selected_entry_id = int(selected_entry_id_raw)
    except (TypeError, ValueError):
        selected_entry_id = None
    if selected_entry_id is not None:
        for index, entry in enumerate(today_entries):
            if int(entry.get("entry_id") or 0) == selected_entry_id:
                active_today_entry = entry
                return {
                    "today_entries": today_entries,
                    "active_index": index,
                    "active_today_entry": active_today_entry,
                    "active_entry_id": selected_entry_id,
                }
    try:
        active_index = int(today_entry_index_raw)
    except (TypeError, ValueError):
        active_index = 0
    if active_index < 0 or active_index >= len(today_entries):
        active_index = 0
    active_today_entry = today_entries[active_index] if today_entries else None
    active_entry_id = (
        int(active_today_entry["entry_id"])
        if active_today_entry is not None and active_today_entry.get("entry_id") is not None
        else None
    )
    return {
        "today_entries": today_entries,
        "active_index": active_index,
        "active_today_entry": active_today_entry,
        "active_entry_id": active_entry_id,
    }


def resolve_vendor_work_entry_preflight_target(
    *,
    preflight_sheet_id: int | None,
    active_entry_id: int | None,
    is_new_entry_mode: bool,
) -> dict[str, object]:
    entry_id = None if is_new_entry_mode else active_entry_id
    return {
        "sheet_id": preflight_sheet_id,
        "entry_id": entry_id,
        "preflight_available": preflight_sheet_id is not None,
    }


def build_vendor_work_entry_draft_entry_defaults(
    today_entries: list[dict[str, object]],
) -> dict[str, object]:
    return {
        "planned_at": "",
        "planned_headcount": 0,
        "actual_headcount": 0,
        "work_content": "",
        "pre_entry_requirement": "",
        "work_headcount": 0,
        "entry_order": len(today_entries),
    }


def resolve_vendor_work_entry_draft_entry_source(
    *,
    is_new_entry_mode: bool,
    active_today_entry: dict[str, object] | None,
    today_entries: list[dict[str, object]],
) -> dict[str, object]:
    if is_new_entry_mode:
        return build_vendor_work_entry_draft_entry_defaults(today_entries)
    return active_today_entry or {}


def build_vendor_work_entry_draft_submit_preparation(
    *,
    preflight_context: dict[str, object],
    business_date: str,
    vendor_name: str,
    draft_entry_source: dict[str, object],
) -> dict[str, object]:
    return {
        "sheet_id": preflight_context.get("sheet_id"),
        "business_date": business_date,
        "vendor_name": vendor_name,
        "entry_id": preflight_context.get("entry_id"),
        "write_mode": preflight_context.get("write_mode"),
        "planned_at": str(draft_entry_source.get("planned_at", "")),
        "planned_headcount": int(draft_entry_source.get("planned_headcount", 0) or 0),
        "actual_headcount": int(draft_entry_source.get("actual_headcount", 0) or 0),
        "work_content": str(draft_entry_source.get("work_content", "")),
        "pre_entry_requirement": str(draft_entry_source.get("pre_entry_requirement", "")),
        "work_headcount": int(draft_entry_source.get("work_headcount", 0) or 0),
        "entry_order": int(draft_entry_source.get("entry_order", 0) or 0),
    }


def normalize_vendor_work_entry_submit_payload(data) -> dict[str, object]:
    entry_id = parse_optional_positive_int(data.get("id"), field_name="id")
    sheet_id = parse_non_negative_int(data.get("sheet_id"), field_name="sheet_id")
    if sheet_id <= 0:
        raise ValueError("sheet_id must be a positive integer.")
    vendor_name = normalize_vendor_name(str(data.get("vendor_name", "")))
    business_date = (
        parse_crew_business_date(str(data.get("business_date", "")))
        if str(data.get("business_date", "")).strip()
        else resolve_crew_business_date()
    )
    planned_at = parse_crew_planned_at(str(data.get("planned_at", "")))
    planned_headcount = parse_non_negative_int(data.get("planned_headcount", 0), field_name="planned_headcount")
    actual_headcount = parse_non_negative_int(data.get("actual_headcount", 0), field_name="actual_headcount")
    work_content = str(data.get("work_content", "")).strip()
    pre_entry_requirement = str(data.get("pre_entry_requirement", "")).strip()
    work_headcount = parse_non_negative_int(data.get("work_headcount", 0), field_name="work_headcount")
    entry_order = parse_non_negative_int(data.get("entry_order", 0), field_name="entry_order")
    return {
        "entry_id": entry_id,
        "sheet_id": sheet_id,
        "vendor_name": vendor_name,
        "business_date": business_date,
        "planned_at": planned_at,
        "planned_headcount": planned_headcount,
        "actual_headcount": actual_headcount,
        "work_content": work_content,
        "pre_entry_requirement": pre_entry_requirement,
        "work_headcount": work_headcount,
        "entry_order": entry_order,
    }


def build_vendor_work_entry_page_context(
    conn: sqlite3.Connection,
    vendor_account,
    business_identity: dict[str, object],
    scope: dict[str, object],
    business_date: str,
    today_entry_index_raw: object,
    selected_entry_id_raw: object,
    new_entry_raw: object,
    submit_status_raw: object,
    submit_mode_raw: object,
) -> dict[str, object]:
    profile_payload = {
        "ok": True,
        "vendor_account_id": int(vendor_account["id"]),
        "vendor_username": str(vendor_account["username"]),
        "vendor_name": str(vendor_account["vendor_name"]),
    }
    scope_payload = {"ok": True, "scope": scope}

    rows = fetch_vendor_business_read_preview(
        conn,
        vendor_name=str(business_identity["vendor_name"]),
    )
    preview_payload = serialize_vendor_business_read_preview(
        business_identity=business_identity,
        rows=rows,
    )
    preflight_sheet_id = find_vendor_work_entry_preflight_sheet_id(
        conn,
        vendor_name=str(business_identity["vendor_name"]),
    )
    preview_entries = list(preview_payload.get("entries", []))
    selection = resolve_vendor_today_entry_selection(
        preview_entries,
        business_date,
        today_entry_index_raw,
        selected_entry_id_raw,
    )
    today_entries = list(selection["today_entries"])
    active_today_entry = selection["active_today_entry"]
    is_new_entry_mode = str(new_entry_raw).strip() == "1"
    preflight_target = resolve_vendor_work_entry_preflight_target(
        preflight_sheet_id=preflight_sheet_id,
        active_entry_id=selection["active_entry_id"],
        is_new_entry_mode=is_new_entry_mode,
    )

    preflight_payload = None
    if preflight_target["preflight_available"]:
        preflight_payload = build_vendor_work_entry_preflight_payload(
            conn,
            sheet_id=int(preflight_target["sheet_id"]),
            business_date=business_date,
            entry_id=preflight_target["entry_id"],
        )

    pending_items = get_pending_items_by_vendor(preflight_sheet_id) if preflight_sheet_id is not None else {}
    vendor_pending_items = list(pending_items.get(str(business_identity["vendor_name"]), []))
    preflight_context = preflight_payload.get("preflight") if isinstance(preflight_payload, dict) else {}
    has_work_content = bool(str((active_today_entry or {}).get("work_content", "")).strip())
    today_entries_summary = {
        "entry_count": len(today_entries),
        "active_index": int(selection["active_index"]),
        "entries": today_entries,
    }
    draft_mode = "create" if is_new_entry_mode else "selected_entry"
    draft_entry_source = resolve_vendor_work_entry_draft_entry_source(
        is_new_entry_mode=is_new_entry_mode,
        active_today_entry=active_today_entry,
        today_entries=today_entries,
    )
    readiness_summary = {
        "vendor_name": str(profile_payload["vendor_name"]),
        "vendor_username": str(profile_payload["vendor_username"]),
        "business_date": business_date,
        "preflight_available": preflight_payload is not None,
        "active_entry_id": (active_today_entry or {}).get("entry_id"),
        "sheet_id": preflight_context.get("sheet_id"),
        "write_mode": preflight_context.get("write_mode"),
        "entry_id": preflight_context.get("entry_id"),
        "has_today_entry": active_today_entry is not None,
        "planned_at": str((active_today_entry or {}).get("planned_at", "")),
        "planned_headcount": int((active_today_entry or {}).get("planned_headcount", 0) or 0),
        "actual_headcount": int((active_today_entry or {}).get("actual_headcount", 0) or 0),
        "has_work_content": has_work_content,
        "work_headcount": int((active_today_entry or {}).get("work_headcount", 0) or 0),
        "pending_items": vendor_pending_items,
        "pending_item_count": len(vendor_pending_items),
    }
    draft_submit_preparation = build_vendor_work_entry_draft_submit_preparation(
        preflight_context=preflight_context,
        business_date=business_date,
        vendor_name=str(profile_payload["vendor_name"]),
        draft_entry_source=draft_entry_source,
    )
    submit_result = {
        "status": str(submit_status_raw or ""),
        "mode": str(submit_mode_raw or ""),
    }
    return {
        "profile_payload": profile_payload,
        "scope_payload": scope_payload,
        "preview_payload": preview_payload,
        "preflight_payload": preflight_payload,
        "preflight_sheet_id": preflight_sheet_id,
        "today_entries_summary": today_entries_summary,
        "readiness_summary": readiness_summary,
        "draft_submit_preparation": draft_submit_preparation,
        "submit_result": submit_result,
        "draft_mode": draft_mode,
        "is_new_entry_mode": is_new_entry_mode,
        "preflight_business_date": business_date,
    }


def init_schema(conn: sqlite3.Connection) -> None:
    conn.executescript(
        """
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT NOT NULL UNIQUE,
            display_name TEXT,
            password_hash TEXT NOT NULL,
            role TEXT NOT NULL DEFAULT 'member',
            created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
        );

        CREATE TABLE IF NOT EXISTS tasks (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            sheet_id INTEGER,
            col_index INTEGER NOT NULL UNIQUE,
            vendor TEXT,
            location TEXT,
            name TEXT NOT NULL
        );

        CREATE TABLE IF NOT EXISTS sheets (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            sort_order INTEGER NOT NULL DEFAULT 1,
            site_id INTEGER,
            created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
        );

        CREATE TABLE IF NOT EXISTS floors (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            sheet_id INTEGER,
            sort_order INTEGER NOT NULL UNIQUE,
            name TEXT NOT NULL,
            block_name TEXT,
            unit_count INTEGER NOT NULL
        );

        CREATE TABLE IF NOT EXISTS units (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            floor_id INTEGER NOT NULL,
            sort_order INTEGER NOT NULL,
            name TEXT NOT NULL,
            FOREIGN KEY (floor_id) REFERENCES floors(id)
        );

        CREATE TABLE IF NOT EXISTS progress (
            unit_id INTEGER NOT NULL,
            task_id INTEGER NOT NULL,
            value TEXT NOT NULL DEFAULT 'X',
            updated_by INTEGER,
            updated_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
            PRIMARY KEY (unit_id, task_id),
            FOREIGN KEY (unit_id) REFERENCES units(id),
            FOREIGN KEY (task_id) REFERENCES tasks(id),
            FOREIGN KEY (updated_by) REFERENCES users(id)
        );

        CREATE TABLE IF NOT EXISTS unit_extra (
            unit_id INTEGER PRIMARY KEY,
            initial_check TEXT NOT NULL DEFAULT '',
            recheck_1 TEXT NOT NULL DEFAULT '',
            recheck_2 TEXT NOT NULL DEFAULT '',
            handover TEXT NOT NULL DEFAULT 'X',
            updated_by INTEGER,
            updated_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (unit_id) REFERENCES units(id),
            FOREIGN KEY (updated_by) REFERENCES users(id)
        );

        CREATE TABLE IF NOT EXISTS extra_fields (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            sheet_id INTEGER NOT NULL,
            field_key TEXT NOT NULL,
            name TEXT NOT NULL,
            field_type TEXT NOT NULL DEFAULT 'date',
            sort_order INTEGER NOT NULL DEFAULT 0,
            is_builtin INTEGER NOT NULL DEFAULT 0,
            active INTEGER NOT NULL DEFAULT 1,
            UNIQUE(sheet_id, field_key),
            FOREIGN KEY (sheet_id) REFERENCES sheets(id)
        );

        CREATE TABLE IF NOT EXISTS unit_extra_values (
            unit_id INTEGER NOT NULL,
            field_key TEXT NOT NULL,
            value TEXT NOT NULL DEFAULT '',
            updated_by INTEGER,
            updated_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
            PRIMARY KEY (unit_id, field_key),
            FOREIGN KEY (unit_id) REFERENCES units(id),
            FOREIGN KEY (updated_by) REFERENCES users(id)
        );

        CREATE TABLE IF NOT EXISTS meta (
            key TEXT PRIMARY KEY,
            value TEXT NOT NULL
        );

        CREATE TABLE IF NOT EXISTS sites (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            site_name TEXT NOT NULL,
            site_code TEXT NOT NULL DEFAULT '',
            is_active INTEGER NOT NULL DEFAULT 1,
            created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
            updated_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
            UNIQUE(site_name)
        );

        CREATE TABLE IF NOT EXISTS user_site_permissions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            site_id INTEGER NOT NULL,
            role TEXT NOT NULL,
            created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
            updated_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
            UNIQUE(user_id, site_id)
        );

        CREATE TABLE IF NOT EXISTS vendor_accounts (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT NOT NULL,
            password_hash TEXT NOT NULL,
            vendor_name TEXT NOT NULL,
            is_active INTEGER NOT NULL DEFAULT 1,
            created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
            updated_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
            UNIQUE(username)
        );

        CREATE TABLE IF NOT EXISTS vendor_contacts (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            sheet_id INTEGER NOT NULL,
            vendor_name TEXT NOT NULL,
            contact_name TEXT NOT NULL DEFAULT '',
            contact_title TEXT NOT NULL DEFAULT '',
            contact_phone TEXT NOT NULL DEFAULT '',
            is_primary INTEGER NOT NULL DEFAULT 0,
            contact_order INTEGER NOT NULL DEFAULT 0,
            created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
            updated_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (sheet_id) REFERENCES sheets(id)
        );

        CREATE TABLE IF NOT EXISTS vendor_work_entries (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            sheet_id INTEGER NOT NULL,
            vendor_name TEXT NOT NULL,
            business_date TEXT NOT NULL,
            planned_at TEXT NOT NULL DEFAULT '',
            planned_headcount INTEGER NOT NULL DEFAULT 0,
            actual_headcount INTEGER NOT NULL DEFAULT 0,
            work_content TEXT NOT NULL DEFAULT '',
            work_headcount INTEGER NOT NULL DEFAULT 0,
            entry_order INTEGER NOT NULL DEFAULT 0,
            pre_entry_requirement TEXT,
            requirement_status TEXT DEFAULT 'pending',
            requirement_confirmed_by TEXT,
            requirement_confirmed_at TEXT,
            created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
            updated_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (sheet_id) REFERENCES sheets(id)
        );

        CREATE TABLE IF NOT EXISTS formal_approvals (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            entry_id INTEGER NOT NULL,
            sheet_id INTEGER NOT NULL,
            action TEXT NOT NULL DEFAULT 'crew_formal_approve_entry',
            approval_status TEXT NOT NULL DEFAULT 'approved',
            approved_by TEXT,
            approved_at TEXT,
            created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
            updated_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (entry_id) REFERENCES vendor_work_entries(id),
            FOREIGN KEY (sheet_id) REFERENCES sheets(id)
        );

        CREATE TABLE IF NOT EXISTS scheduling_entries (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            entry_id INTEGER NOT NULL,
            sheet_id INTEGER NOT NULL,
            action TEXT NOT NULL DEFAULT 'schedule_entry',
            schedule_status TEXT NOT NULL DEFAULT 'scheduled',
            scheduled_date TEXT NOT NULL DEFAULT '',
            scheduled_time TEXT NOT NULL DEFAULT '',
            scheduled_by TEXT,
            scheduled_at TEXT,
            created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
            updated_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (entry_id) REFERENCES vendor_work_entries(id),
            FOREIGN KEY (sheet_id) REFERENCES sheets(id)
        );
        """
    )
    conn.execute("CREATE INDEX IF NOT EXISTS idx_user_site_permissions_user_id ON user_site_permissions (user_id)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_user_site_permissions_site_id ON user_site_permissions (site_id)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_vendor_accounts_vendor_name ON vendor_accounts (vendor_name)")
    conn.executescript(
        """
        CREATE INDEX IF NOT EXISTS idx_vendor_work_entries_sheet_business_date
        ON vendor_work_entries (sheet_id, business_date);

        CREATE INDEX IF NOT EXISTS idx_vendor_work_entries_sheet_vendor_date
        ON vendor_work_entries (sheet_id, vendor_name, business_date);

        CREATE INDEX IF NOT EXISTS idx_vendor_work_entries_business_date
        ON vendor_work_entries (business_date);
        """
    )
    ensure_vendor_work_entries_schema(conn)
    ensure_formal_approvals_schema(conn)
    ensure_scheduling_entries_schema(conn)
    ensure_vendor_contacts_schema(conn)


def seed_admin(conn: sqlite3.Connection) -> None:
    exists = conn.execute("SELECT id FROM users WHERE username = ?", ("admin",)).fetchone()
    if exists:
        return
    conn.execute(
        "INSERT INTO users (username, display_name, password_hash, role) VALUES (?, ?, ?, ?)",
        ("admin", "管理員", generate_password_hash("admin"), "admin"),
    )


def _table_columns(conn: sqlite3.Connection, table_name: str) -> set[str]:
    return {row["name"] for row in conn.execute(f"PRAGMA table_info({table_name})")}


def ensure_vendor_work_entries_schema(conn: sqlite3.Connection) -> None:
    existing_tables = {
        row["name"] for row in conn.execute("SELECT name FROM sqlite_master WHERE type = 'table'")
    }
    if "vendor_work_entries" not in existing_tables:
        conn.executescript(
            """
            CREATE TABLE vendor_work_entries (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                sheet_id INTEGER NOT NULL,
                vendor_name TEXT NOT NULL,
                business_date TEXT NOT NULL,
                planned_at TEXT NOT NULL DEFAULT '',
                planned_headcount INTEGER NOT NULL DEFAULT 0,
                actual_headcount INTEGER NOT NULL DEFAULT 0,
                work_content TEXT NOT NULL DEFAULT '',
                work_headcount INTEGER NOT NULL DEFAULT 0,
                entry_order INTEGER NOT NULL DEFAULT 0,
                pre_entry_requirement TEXT,
                requirement_status TEXT DEFAULT 'pending',
                requirement_confirmed_by TEXT,
                requirement_confirmed_at TEXT,
                created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                updated_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (sheet_id) REFERENCES sheets(id)
            );

            CREATE INDEX IF NOT EXISTS idx_vendor_work_entries_sheet_business_date
            ON vendor_work_entries (sheet_id, business_date);

            CREATE INDEX IF NOT EXISTS idx_vendor_work_entries_sheet_vendor_date
            ON vendor_work_entries (sheet_id, vendor_name, business_date);

            CREATE INDEX IF NOT EXISTS idx_vendor_work_entries_business_date
            ON vendor_work_entries (business_date);
            """
        )
        return

    vendor_work_entry_columns = _table_columns(conn, "vendor_work_entries")
    if "pre_entry_requirement" not in vendor_work_entry_columns:
        conn.execute("ALTER TABLE vendor_work_entries ADD COLUMN pre_entry_requirement TEXT")
    if "requirement_status" not in vendor_work_entry_columns:
        conn.execute("ALTER TABLE vendor_work_entries ADD COLUMN requirement_status TEXT DEFAULT 'pending'")
    if "requirement_confirmed_by" not in vendor_work_entry_columns:
        conn.execute("ALTER TABLE vendor_work_entries ADD COLUMN requirement_confirmed_by TEXT")
    if "requirement_confirmed_at" not in vendor_work_entry_columns:
        conn.execute("ALTER TABLE vendor_work_entries ADD COLUMN requirement_confirmed_at TEXT")
    conn.executescript(
        """
        CREATE INDEX IF NOT EXISTS idx_vendor_work_entries_sheet_business_date
        ON vendor_work_entries (sheet_id, business_date);

        CREATE INDEX IF NOT EXISTS idx_vendor_work_entries_sheet_vendor_date
        ON vendor_work_entries (sheet_id, vendor_name, business_date);

        CREATE INDEX IF NOT EXISTS idx_vendor_work_entries_business_date
        ON vendor_work_entries (business_date);
        """
    )


def ensure_formal_approvals_schema(conn: sqlite3.Connection) -> None:
    existing_tables = {
        row["name"] for row in conn.execute("SELECT name FROM sqlite_master WHERE type = 'table'")
    }
    if "formal_approvals" not in existing_tables:
        conn.executescript(
            """
            CREATE TABLE formal_approvals (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                entry_id INTEGER NOT NULL,
                sheet_id INTEGER NOT NULL,
                action TEXT NOT NULL DEFAULT 'crew_formal_approve_entry',
                approval_status TEXT NOT NULL DEFAULT 'approved',
                approved_by TEXT,
                approved_at TEXT,
                created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                updated_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (entry_id) REFERENCES vendor_work_entries(id),
                FOREIGN KEY (sheet_id) REFERENCES sheets(id)
            );

            CREATE UNIQUE INDEX IF NOT EXISTS idx_formal_approvals_entry_action_unique
            ON formal_approvals (entry_id, action);

            CREATE INDEX IF NOT EXISTS idx_formal_approvals_sheet_id
            ON formal_approvals (sheet_id);
            """
        )
        return

    formal_approvals_columns = _table_columns(conn, "formal_approvals")
    if "entry_id" not in formal_approvals_columns:
        conn.execute("ALTER TABLE formal_approvals ADD COLUMN entry_id INTEGER")
    if "sheet_id" not in formal_approvals_columns:
        conn.execute("ALTER TABLE formal_approvals ADD COLUMN sheet_id INTEGER")
    if "action" not in formal_approvals_columns:
        conn.execute(
            "ALTER TABLE formal_approvals ADD COLUMN action TEXT NOT NULL DEFAULT 'crew_formal_approve_entry'"
        )
    if "approval_status" not in formal_approvals_columns:
        conn.execute("ALTER TABLE formal_approvals ADD COLUMN approval_status TEXT NOT NULL DEFAULT 'approved'")
    if "approved_by" not in formal_approvals_columns:
        conn.execute("ALTER TABLE formal_approvals ADD COLUMN approved_by TEXT")
    if "approved_at" not in formal_approvals_columns:
        conn.execute("ALTER TABLE formal_approvals ADD COLUMN approved_at TEXT")
    if "created_at" not in formal_approvals_columns:
        conn.execute("ALTER TABLE formal_approvals ADD COLUMN created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP")
    if "updated_at" not in formal_approvals_columns:
        conn.execute("ALTER TABLE formal_approvals ADD COLUMN updated_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP")
    conn.executescript(
        """
        CREATE UNIQUE INDEX IF NOT EXISTS idx_formal_approvals_entry_action_unique
        ON formal_approvals (entry_id, action);

        CREATE INDEX IF NOT EXISTS idx_formal_approvals_sheet_id
        ON formal_approvals (sheet_id);
        """
    )


def ensure_scheduling_entries_schema(conn: sqlite3.Connection) -> None:
    existing_tables = {
        row["name"] for row in conn.execute("SELECT name FROM sqlite_master WHERE type = 'table'")
    }
    if "scheduling_entries" not in existing_tables:
        conn.executescript(
            """
            CREATE TABLE scheduling_entries (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                entry_id INTEGER NOT NULL,
                sheet_id INTEGER NOT NULL,
                action TEXT NOT NULL DEFAULT 'schedule_entry',
                schedule_status TEXT NOT NULL DEFAULT 'scheduled',
                scheduled_date TEXT NOT NULL DEFAULT '',
                scheduled_time TEXT NOT NULL DEFAULT '',
                scheduled_by TEXT,
                scheduled_at TEXT,
                created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                updated_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (entry_id) REFERENCES vendor_work_entries(id),
                FOREIGN KEY (sheet_id) REFERENCES sheets(id)
            );

            CREATE UNIQUE INDEX IF NOT EXISTS idx_scheduling_entries_entry_action_unique
            ON scheduling_entries (entry_id, action);

            CREATE INDEX IF NOT EXISTS idx_scheduling_entries_sheet_id
            ON scheduling_entries (sheet_id);

            CREATE INDEX IF NOT EXISTS idx_scheduling_entries_scheduled_date
            ON scheduling_entries (scheduled_date);
            """
        )
        return

    scheduling_entries_columns = _table_columns(conn, "scheduling_entries")
    if "entry_id" not in scheduling_entries_columns:
        conn.execute("ALTER TABLE scheduling_entries ADD COLUMN entry_id INTEGER")
    if "sheet_id" not in scheduling_entries_columns:
        conn.execute("ALTER TABLE scheduling_entries ADD COLUMN sheet_id INTEGER")
    if "action" not in scheduling_entries_columns:
        conn.execute("ALTER TABLE scheduling_entries ADD COLUMN action TEXT NOT NULL DEFAULT 'schedule_entry'")
    if "schedule_status" not in scheduling_entries_columns:
        conn.execute("ALTER TABLE scheduling_entries ADD COLUMN schedule_status TEXT NOT NULL DEFAULT 'scheduled'")
    if "scheduled_date" not in scheduling_entries_columns:
        conn.execute("ALTER TABLE scheduling_entries ADD COLUMN scheduled_date TEXT NOT NULL DEFAULT ''")
    if "scheduled_time" not in scheduling_entries_columns:
        conn.execute("ALTER TABLE scheduling_entries ADD COLUMN scheduled_time TEXT NOT NULL DEFAULT ''")
    if "scheduled_by" not in scheduling_entries_columns:
        conn.execute("ALTER TABLE scheduling_entries ADD COLUMN scheduled_by TEXT")
    if "scheduled_at" not in scheduling_entries_columns:
        conn.execute("ALTER TABLE scheduling_entries ADD COLUMN scheduled_at TEXT")
    if "created_at" not in scheduling_entries_columns:
        conn.execute("ALTER TABLE scheduling_entries ADD COLUMN created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP")
    if "updated_at" not in scheduling_entries_columns:
        conn.execute("ALTER TABLE scheduling_entries ADD COLUMN updated_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP")
    conn.executescript(
        """
        CREATE UNIQUE INDEX IF NOT EXISTS idx_scheduling_entries_entry_action_unique
        ON scheduling_entries (entry_id, action);

        CREATE INDEX IF NOT EXISTS idx_scheduling_entries_sheet_id
        ON scheduling_entries (sheet_id);

        CREATE INDEX IF NOT EXISTS idx_scheduling_entries_scheduled_date
        ON scheduling_entries (scheduled_date);
        """
    )


def seed_default_site(conn: sqlite3.Connection) -> int:
    row = conn.execute("SELECT id FROM sites WHERE site_name = ?", (DEFAULT_SITE_NAME,)).fetchone()
    if row:
        return row["id"]
    cur = conn.execute(
        """
        INSERT INTO sites (site_name, site_code, is_active)
        VALUES (?, ?, 1)
        """,
        (DEFAULT_SITE_NAME, DEFAULT_SITE_CODE),
    )
    return cur.lastrowid


def ensure_sheets_site_backfill(conn: sqlite3.Connection, default_site_id: int) -> None:
    sheet_columns = _table_columns(conn, "sheets")
    if "site_id" not in sheet_columns:
        conn.execute("ALTER TABLE sheets ADD COLUMN site_id INTEGER")
    conn.execute("UPDATE sheets SET site_id = ? WHERE site_id IS NULL", (default_site_id,))
    conn.execute("CREATE INDEX IF NOT EXISTS idx_sheets_site_id ON sheets (site_id)")


def ensure_site_foundation_schema(conn: sqlite3.Connection) -> int:
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS sites (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            site_name TEXT NOT NULL,
            site_code TEXT NOT NULL DEFAULT '',
            is_active INTEGER NOT NULL DEFAULT 1,
            created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
            updated_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
            UNIQUE(site_name)
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS user_site_permissions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            site_id INTEGER NOT NULL,
            role TEXT NOT NULL,
            created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
            updated_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
            UNIQUE(user_id, site_id)
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS vendor_accounts (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT NOT NULL,
            password_hash TEXT NOT NULL,
            vendor_name TEXT NOT NULL,
            is_active INTEGER NOT NULL DEFAULT 1,
            created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
            updated_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
            UNIQUE(username)
        )
        """
    )
    conn.execute("CREATE INDEX IF NOT EXISTS idx_user_site_permissions_user_id ON user_site_permissions (user_id)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_user_site_permissions_site_id ON user_site_permissions (site_id)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_vendor_accounts_vendor_name ON vendor_accounts (vendor_name)")
    default_site_id = seed_default_site(conn)
    ensure_sheets_site_backfill(conn, default_site_id)
    return default_site_id


def _site_row_payload(row: sqlite3.Row | None) -> dict[str, object] | None:
    if row is None:
        return None
    return {
        "id": int(row["id"]),
        "site_name": str(row["site_name"]),
        "site_code": str(row["site_code"]),
        "is_active": int(row["is_active"]),
    }


def _fetch_default_site_row(conn: sqlite3.Connection) -> sqlite3.Row | None:
    return conn.execute(
        """
        SELECT id, site_name, site_code, is_active
        FROM sites
        WHERE site_name = ?
        """,
        (DEFAULT_SITE_NAME,),
    ).fetchone()


def _fetch_site_row_by_id(conn: sqlite3.Connection, site_id: int) -> sqlite3.Row | None:
    return conn.execute(
        """
        SELECT id, site_name, site_code, is_active
        FROM sites
        WHERE id = ?
        """,
        (site_id,),
    ).fetchone()


def _fetch_active_sites(conn: sqlite3.Connection) -> list[dict[str, object]]:
    rows = conn.execute(
        """
        SELECT id, site_name, site_code, is_active
        FROM sites
        WHERE is_active = 1
        ORDER BY id
        """
    ).fetchall()
    return [_site_row_payload(row) for row in rows if row is not None]


def _site_permission_payload(row: sqlite3.Row | None) -> dict[str, object] | None:
    if row is None:
        return None
    role = str(row["role"] or "")
    return {
        "id": int(row["id"]),
        "user_id": int(row["user_id"]),
        "site_id": int(row["site_id"]),
        "site_name": str(row["site_name"] or ""),
        "site_code": str(row["site_code"] or ""),
        "role": role,
        "role_label": SITE_PERMISSION_ROLE_LABELS.get(role, role),
        "is_active": int(row["is_active"]) == 1,
    }


def _fetch_user_site_permission_row(conn: sqlite3.Connection, permission_id: int) -> sqlite3.Row | None:
    return conn.execute(
        """
        SELECT usp.id, usp.user_id, usp.site_id, usp.role, s.site_name, s.site_code, s.is_active
        FROM user_site_permissions usp
        JOIN sites s ON s.id = usp.site_id
        WHERE usp.id = ?
        """,
        (permission_id,),
    ).fetchone()


def get_user_site_permissions_map(
    conn: sqlite3.Connection,
    *,
    user_ids: list[int] | None = None,
) -> dict[int, list[dict[str, object]]]:
    query = """
        SELECT usp.id, usp.user_id, usp.site_id, usp.role, s.site_name, s.site_code, s.is_active
        FROM user_site_permissions usp
        JOIN sites s ON s.id = usp.site_id
    """
    params: list[object] = []
    if user_ids:
        placeholders = ",".join("?" for _ in user_ids)
        query += f" WHERE usp.user_id IN ({placeholders})"
        params.extend(user_ids)
    query += " ORDER BY usp.user_id, s.site_name COLLATE NOCASE, usp.id"
    permission_map: dict[int, list[dict[str, object]]] = {}
    for row in conn.execute(query, params).fetchall():
        payload = _site_permission_payload(row)
        if payload is None:
            continue
        permission_map.setdefault(int(payload["user_id"]), []).append(payload)
    return permission_map


def _write_current_site_session(*, site_id: int, site_name: str) -> None:
    session["current_site_id"] = int(site_id)
    session["current_site_name"] = str(site_name)
    session["site_selection_required"] = False


def get_default_site_id(conn: sqlite3.Connection | None = None) -> int | None:
    if conn is not None:
        row = _fetch_default_site_row(conn)
        return int(row["id"]) if row else None
    with db() as temp_conn:
        row = _fetch_default_site_row(temp_conn)
    return int(row["id"]) if row else None


def get_current_site_id() -> int | None:
    if not has_request_context():
        return None
    raw = session.get("current_site_id")
    if raw in (None, ""):
        return None
    try:
        return int(raw)
    except (TypeError, ValueError):
        return None


def set_current_site_id(site_id: int) -> None:
    with db() as conn:
        row = _fetch_site_row_by_id(conn, int(site_id))
    if row is None:
        raise LookupError("site_not_found")
    _write_current_site_session(site_id=int(row["id"]), site_name=str(row["site_name"]))


def clear_current_site() -> None:
    if not has_request_context():
        return
    session.pop("current_site_id", None)
    session.pop("current_site_name", None)
    session.pop("site_selection_required", None)


def clear_current_site_selection_context() -> None:
    if not has_request_context():
        return
    session.pop("current_site_id", None)
    session.pop("current_site_name", None)


def _site_selection_redirect_target(resolution: dict[str, object]) -> str:
    if resolution["status"] == "site_selection_required":
        return url_for("site_selector")
    return url_for("sheet")


def _current_internal_user():
    user_id = session.get("user_id")
    if not user_id:
        return None
    return get_user_by_id(int(user_id))


def is_global_admin(user) -> bool:
    if not user:
        return False
    try:
        role = user["role"]
    except (KeyError, TypeError, IndexError):
        role = getattr(user, "role", None)
    return str(role or "").strip() == "admin"


def get_user_accessible_sites(user_id: int) -> list[dict[str, object]]:
    user = get_user_by_id(user_id)
    if user is None:
        return []
    with db() as conn:
        if is_global_admin(user):
            return _fetch_active_sites(conn)
        rows = conn.execute(
            """
            SELECT DISTINCT s.id, s.site_name, s.site_code, s.is_active
            FROM user_site_permissions usp
            JOIN sites s ON s.id = usp.site_id
            WHERE usp.user_id = ? AND s.is_active = 1
            ORDER BY s.id
            """,
            (user_id,),
        ).fetchall()
    return [_site_row_payload(row) for row in rows if row is not None]


def user_can_access_site(user_id: int, site_id: int) -> bool:
    user = get_user_by_id(user_id)
    if user is None:
        return False
    with db() as conn:
        if is_global_admin(user):
            row = conn.execute("SELECT 1 FROM sites WHERE id = ? AND is_active = 1", (site_id,)).fetchone()
            return row is not None
        row = conn.execute(
            """
            SELECT 1
            FROM user_site_permissions usp
            JOIN sites s ON s.id = usp.site_id
            WHERE usp.user_id = ? AND usp.site_id = ? AND s.is_active = 1
            """,
            (user_id, site_id),
        ).fetchone()
    return row is not None


def get_user_role_for_site(user_id: int, site_id: int) -> str | None:
    user = get_user_by_id(user_id)
    if user is None:
        return None
    with db() as conn:
        if is_global_admin(user):
            site_row = conn.execute("SELECT 1 FROM sites WHERE id = ? AND is_active = 1", (site_id,)).fetchone()
            return "admin" if site_row else None
        row = conn.execute(
            """
            SELECT usp.role
            FROM user_site_permissions usp
            JOIN sites s ON s.id = usp.site_id
            WHERE usp.user_id = ? AND usp.site_id = ? AND s.is_active = 1
            """,
            (user_id, site_id),
        ).fetchone()
    return str(row["role"]) if row else None


def resolve_current_site_for_user(user) -> dict[str, object]:
    if not user:
        return {
            "status": "access_denied_no_site_permission",
            "site_id": None,
            "site_name": "",
            "site_selection_required": False,
            "accessible_sites": [],
            "reason": "missing_user",
        }

    current_site_id = get_current_site_id()
    with db() as conn:
        default_site_row = _fetch_default_site_row(conn)
        accessible_sites = get_user_accessible_sites(int(user["id"]))
        accessible_site_ids = {int(site["id"]) for site in accessible_sites}

        if is_global_admin(user):
            if default_site_row is None:
                return {
                    "status": "default_site_missing",
                    "site_id": None,
                    "site_name": "",
                    "site_selection_required": False,
                    "accessible_sites": accessible_sites,
                    "reason": "default_site_missing",
                }
            current_site_row = _fetch_site_row_by_id(conn, current_site_id) if current_site_id else None
            if current_site_row is not None and int(current_site_row["is_active"]) == 1:
                chosen_row = current_site_row
                reason = "session_site_valid"
            else:
                chosen_row = default_site_row
                reason = "admin_default_fallback"
            return {
                "status": "resolved",
                "site_id": int(chosen_row["id"]),
                "site_name": str(chosen_row["site_name"]),
                "site_selection_required": False,
                "accessible_sites": accessible_sites,
                "reason": reason,
            }

        accessible_count = len(accessible_sites)
        if accessible_count == 0:
            return {
                "status": "access_denied_no_site_permission",
                "site_id": None,
                "site_name": "",
                "site_selection_required": False,
                "accessible_sites": accessible_sites,
                "reason": "no_accessible_sites",
            }

        if accessible_count == 1:
            only_site = accessible_sites[0]
            if current_site_id == int(only_site["id"]):
                reason = "session_site_valid"
            else:
                reason = "single_site_auto_select"
            return {
                "status": "resolved",
                "site_id": int(only_site["id"]),
                "site_name": str(only_site["site_name"]),
                "site_selection_required": False,
                "accessible_sites": accessible_sites,
                "reason": reason,
            }

        if current_site_id and current_site_id in accessible_site_ids:
            current_site_row = _fetch_site_row_by_id(conn, current_site_id)
            if current_site_row is not None and int(current_site_row["is_active"]) == 1:
                return {
                    "status": "resolved",
                    "site_id": int(current_site_row["id"]),
                    "site_name": str(current_site_row["site_name"]),
                    "site_selection_required": False,
                    "accessible_sites": accessible_sites,
                    "reason": "session_site_valid",
                }

        return {
            "status": "site_selection_required",
            "site_id": None,
            "site_name": "",
            "site_selection_required": True,
            "accessible_sites": accessible_sites,
            "reason": "multiple_accessible_sites",
        }


def normalize_current_site_for_user(user) -> dict[str, object]:
    resolution = resolve_current_site_for_user(user)
    status = resolution["status"]
    if status == "resolved":
        _write_current_site_session(
            site_id=int(resolution["site_id"]),
            site_name=str(resolution["site_name"]),
        )
    elif status == "site_selection_required":
        clear_current_site()
        if has_request_context():
            session["site_selection_required"] = True
    else:
        clear_current_site()
    return resolution


def _render_site_selector(*, user, status_code: int = 200, error_message: str = ""):
    resolution = resolve_current_site_for_user(user)
    accessible_sites = list(resolution.get("accessible_sites", []))
    current_site = None
    current_site_id = get_current_site_id()
    if current_site_id is not None:
        current_site = next(
            (site for site in accessible_sites if int(site["id"]) == int(current_site_id)),
            None,
        )
    settings = query_settings()
    app.logger.info(
        "selector_page_viewed user_id=%s current_site_id=%s accessible_site_count=%s",
        user["id"],
        current_site_id,
        len(accessible_sites),
    )
    return (
        render_template(
            "site_selector.html",
            settings=settings,
            accessible_sites=accessible_sites,
            current_site=current_site,
            selector_error=error_message,
        ),
        status_code,
    )


def _vendor_contacts_has_legacy_unique(conn: sqlite3.Connection) -> bool:
    for row in conn.execute("PRAGMA index_list(vendor_contacts)"):
        if row["unique"]:
            cols = tuple(index_row["name"] for index_row in conn.execute(f"PRAGMA index_info({row['name']})"))
            if cols == ("sheet_id", "vendor_name"):
                return True
    return False


def ensure_vendor_contacts_schema(conn: sqlite3.Connection) -> None:
    existing_tables = {
        row["name"] for row in conn.execute("SELECT name FROM sqlite_master WHERE type = 'table'")
    }
    if "vendor_contacts" not in existing_tables:
        conn.executescript(
            """
            CREATE TABLE vendor_contacts (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                sheet_id INTEGER NOT NULL,
                vendor_name TEXT NOT NULL,
                contact_name TEXT NOT NULL DEFAULT '',
                contact_title TEXT NOT NULL DEFAULT '',
                contact_phone TEXT NOT NULL DEFAULT '',
                is_primary INTEGER NOT NULL DEFAULT 0,
                contact_order INTEGER NOT NULL DEFAULT 0,
                created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                updated_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (sheet_id) REFERENCES sheets(id)
            );

            CREATE INDEX IF NOT EXISTS idx_vendor_contacts_sheet_id
            ON vendor_contacts (sheet_id);

            CREATE INDEX IF NOT EXISTS idx_vendor_contacts_sheet_vendor
            ON vendor_contacts (sheet_id, vendor_name);

            CREATE INDEX IF NOT EXISTS idx_vendor_contacts_sheet_vendor_order
            ON vendor_contacts (sheet_id, vendor_name, contact_order);
            """
        )
        return

    vendor_contacts_columns = _table_columns(conn, "vendor_contacts")
    has_multi_contact_columns = {"contact_title", "is_primary", "contact_order"}.issubset(vendor_contacts_columns)
    has_legacy_unique = _vendor_contacts_has_legacy_unique(conn)
    if has_multi_contact_columns and not has_legacy_unique:
        if "vendor_contacts_old" in existing_tables:
            conn.execute("DROP TABLE vendor_contacts_old")
        conn.executescript(
            """
            CREATE INDEX IF NOT EXISTS idx_vendor_contacts_sheet_id
            ON vendor_contacts (sheet_id);

            CREATE INDEX IF NOT EXISTS idx_vendor_contacts_sheet_vendor
            ON vendor_contacts (sheet_id, vendor_name);

            CREATE INDEX IF NOT EXISTS idx_vendor_contacts_sheet_vendor_order
            ON vendor_contacts (sheet_id, vendor_name, contact_order);
            """
        )
        return

    old_row_count = conn.execute("SELECT COUNT(*) AS count FROM vendor_contacts").fetchone()["count"]
    conn.executescript(
        """
        DROP INDEX IF EXISTS idx_vendor_contacts_sheet_id;
        DROP INDEX IF EXISTS idx_vendor_contacts_vendor_name;
        DROP INDEX IF EXISTS idx_vendor_contacts_sheet_vendor;
        DROP INDEX IF EXISTS idx_vendor_contacts_sheet_vendor_order;
        """
    )
    conn.execute("ALTER TABLE vendor_contacts RENAME TO vendor_contacts_old")
    conn.executescript(
        """
        CREATE TABLE vendor_contacts (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            sheet_id INTEGER NOT NULL,
            vendor_name TEXT NOT NULL,
            contact_name TEXT NOT NULL DEFAULT '',
            contact_title TEXT NOT NULL DEFAULT '',
            contact_phone TEXT NOT NULL DEFAULT '',
            is_primary INTEGER NOT NULL DEFAULT 0,
            contact_order INTEGER NOT NULL DEFAULT 0,
            created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
            updated_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (sheet_id) REFERENCES sheets(id)
        );
        """
    )
    old_columns = _table_columns(conn, "vendor_contacts_old")
    select_contact_title = "contact_title" if "contact_title" in old_columns else "''"
    select_is_primary = "is_primary" if "is_primary" in old_columns else "1"
    select_contact_order = "contact_order" if "contact_order" in old_columns else "0"
    conn.execute(
        f"""
        INSERT INTO vendor_contacts (
            sheet_id,
            vendor_name,
            contact_name,
            contact_title,
            contact_phone,
            is_primary,
            contact_order,
            created_at,
            updated_at
        )
        SELECT
            sheet_id,
            vendor_name,
            contact_name,
            {select_contact_title},
            contact_phone,
            {select_is_primary},
            {select_contact_order},
            created_at,
            updated_at
        FROM vendor_contacts_old
        ORDER BY id
        """
    )
    new_row_count = conn.execute("SELECT COUNT(*) AS count FROM vendor_contacts").fetchone()["count"]
    if new_row_count != old_row_count:
        raise RuntimeError("vendor_contacts_row_count_mismatch_after_migration")
    conn.executescript(
        """
        CREATE INDEX IF NOT EXISTS idx_vendor_contacts_sheet_id
        ON vendor_contacts (sheet_id);

        CREATE INDEX IF NOT EXISTS idx_vendor_contacts_sheet_vendor
        ON vendor_contacts (sheet_id, vendor_name);

        CREATE INDEX IF NOT EXISTS idx_vendor_contacts_sheet_vendor_order
        ON vendor_contacts (sheet_id, vendor_name, contact_order);
        """
    )
    conn.execute("DROP TABLE vendor_contacts_old")


def get_setting(conn: sqlite3.Connection, key: str) -> str:
    row = conn.execute("SELECT value FROM meta WHERE key = ?", (key,)).fetchone()
    return row["value"] if row else DEFAULT_SETTINGS[key]


def get_settings(conn: sqlite3.Connection) -> dict[str, str]:
    settings = DEFAULT_SETTINGS.copy()
    rows = conn.execute("SELECT key, value FROM meta").fetchall()
    for row in rows:
        if row["key"] in settings:
            settings[row["key"]] = row["value"]
    return settings


def set_setting(conn: sqlite3.Connection, key: str, value: str) -> None:
    conn.execute(
        """
        INSERT INTO meta (key, value) VALUES (?, ?)
        ON CONFLICT(key) DO UPDATE SET value = excluded.value
        """,
        (key, value),
    )


def save_admin_global_settings(conn: sqlite3.Connection, *, form) -> None:
    for key in DEFAULT_SETTINGS:
        if key in form:
            set_setting(conn, key, form.get(key, DEFAULT_SETTINGS[key]).strip())


def save_admin_site_content(conn: sqlite3.Connection, *, sheet_id: int, form) -> None:
    authorize_admin_site_scoped_write(conn, sheet_id=sheet_id)
    conn.execute(
        "UPDATE sheets SET name = ? WHERE id = ?",
        (form.get("sheet_name", "").strip() or "???", sheet_id),
    )
    for task in conn.execute("SELECT id FROM tasks WHERE sheet_id = ?", (sheet_id,)):
        task_id = task["id"]
        conn.execute(
            "UPDATE tasks SET vendor = ?, location = ?, name = ? WHERE id = ?",
            (
                form.get(f"task_vendor_{task_id}", "").strip(),
                form.get(f"task_location_{task_id}", "").strip(),
                form.get(f"task_name_{task_id}", "").strip(),
                task_id,
            ),
        )
    for field in conn.execute("SELECT id FROM extra_fields WHERE sheet_id = ? AND active = 1", (sheet_id,)):
        field_id = field["id"]
        field_type = form.get(f"extra_type_{field_id}", "date")
        if field_type not in EXTRA_FIELD_TYPES:
            field_type = "date"
        conn.execute(
            "UPDATE extra_fields SET name = ?, field_type = ? WHERE id = ? AND sheet_id = ?",
            (form.get(f"extra_name_{field_id}", "").strip() or "??", field_type, field_id, sheet_id),
        )
    for floor in conn.execute("SELECT id FROM floors WHERE sheet_id = ?", (sheet_id,)):
        floor_id = floor["id"]
        floor_name = form.get(f"floor_name_{floor_id}", "").strip()
        floor_block_name = form.get(f"floor_block_{floor_id}", "").strip()
        update_floor_fields_sqlite(conn, floor_id, name=floor_name, block_name=floor_block_name)
        maybe_dual_write_floor_update(floor_id, name=floor_name, block_name=floor_block_name)
    for unit in conn.execute("SELECT u.id FROM units u JOIN floors f ON f.id = u.floor_id WHERE f.sheet_id = ?", (sheet_id,)):
        unit_id = unit["id"]
        conn.execute("UPDATE units SET name = ? WHERE id = ?", (form.get(f"unit_name_{unit_id}", "").strip(), unit_id))


def seed_settings(conn: sqlite3.Connection) -> None:
    for key, value in DEFAULT_SETTINGS.items():
        row = conn.execute("SELECT key FROM meta WHERE key = ?", (key,)).fetchone()
        if not row:
            set_setting(conn, key, value)


def seed_from_excel(conn: sqlite3.Connection) -> None:
    seeded = conn.execute("SELECT value FROM meta WHERE key = 'excel_seeded'").fetchone()
    if seeded:
        return
    if not SOURCE_XLSX.exists():
        raise FileNotFoundError(f"找不到 {SOURCE_XLSX}")

    wb = load_workbook(SOURCE_XLSX, data_only=False)
    ws = wb.active
    sheet_row = conn.execute("SELECT id FROM sheets ORDER BY sort_order, id LIMIT 1").fetchone()
    if not sheet_row:
        default_site_id = ensure_site_foundation_schema(conn)
        cur = conn.execute(
            "INSERT INTO sheets (name, sort_order, site_id) VALUES (?, ?, ?)",
            (get_setting(conn, "tab_title"), 1, default_site_id),
        )
        sheet_id = cur.lastrowid
    else:
        sheet_id = sheet_row["id"]
    task_ids: dict[int, int] = {}
    for col in range(4, MAX_WORK_COL + 1):
        task_name = ws.cell(4, col).value
        if not task_name:
            continue
        cur = conn.execute(
            "INSERT INTO tasks (sheet_id, col_index, vendor, location, name) VALUES (?, ?, ?, ?, ?)",
            (sheet_id, col, ws.cell(2, col).value or "", ws.cell(3, col).value or "", str(task_name)),
        )
        task_ids[col] = cur.lastrowid

    for sort_order, floor_name in enumerate(
        ["20F", "19F", "18F", "17F", "16F", "15F", "14F", "13F", "12F", "11F", "10F", "9F", "8F", "7F", "6F", "5F", "4F", "3F", "2F", "1MF", "1F"],
        start=1,
    ):
        units = desired_units_for_floor(floor_name)
        cur = conn.execute(
            "INSERT INTO floors (sheet_id, sort_order, name, block_name, unit_count) VALUES (?, ?, ?, ?, ?)",
            (sheet_id, sort_order, floor_name, block_for_floor(floor_name), len(units)),
        )
        floor_id = cur.lastrowid
        for unit_order, unit_name in enumerate(units, start=1):
            unit_cur = conn.execute(
                "INSERT INTO units (floor_id, sort_order, name) VALUES (?, ?, ?)",
                (floor_id, unit_order, unit_name),
            )
            unit_id = unit_cur.lastrowid
            for task_id in task_ids.values():
                conn.execute(
                    "INSERT INTO progress (unit_id, task_id, value) VALUES (?, ?, ?)",
                    (unit_id, task_id, WORKING_VALUE),
                )
            conn.execute("INSERT INTO unit_extra (unit_id) VALUES (?)", (unit_id,))

    conn.execute("INSERT INTO meta (key, value) VALUES ('excel_seeded', CURRENT_TIMESTAMP)")


def block_for_floor(floor_name: str) -> str:
    if floor_name in ("1MF", "1F"):
        return "S"
    return "A/B"


def desired_units_for_floor(floor_name: str) -> list[str]:
    if floor_name in ("1MF", "1F"):
        return ["S1", "S2", "S3", "S4", "S5", "S7", "S8"]
    if floor_name == "20F":
        return ["A1", "A2", "A6", "A7", "B1", "B2", "B5", "B7"]
    return ["A1", "A2", "A3", "A5", "A6", "A7", "B1", "B2", "B3", "B5", "B6", "B7"]


def migrate_schema(conn: sqlite3.Connection) -> None:
    existing_tables = {
        row["name"] for row in conn.execute("SELECT name FROM sqlite_master WHERE type = 'table'")
    }
    if "unit_extra" not in existing_tables:
        conn.executescript(
            """
            CREATE TABLE unit_extra (
                unit_id INTEGER PRIMARY KEY,
                initial_check TEXT NOT NULL DEFAULT '',
                recheck_1 TEXT NOT NULL DEFAULT '',
                recheck_2 TEXT NOT NULL DEFAULT '',
                handover TEXT NOT NULL DEFAULT 'X',
                updated_by INTEGER,
                updated_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (unit_id) REFERENCES units(id),
                FOREIGN KEY (updated_by) REFERENCES users(id)
            );
            """
        )
    task_cols = {row["name"] for row in conn.execute("PRAGMA table_info(tasks)")}
    floor_cols = {row["name"] for row in conn.execute("PRAGMA table_info(floors)")}
    user_cols = {row["name"] for row in conn.execute("PRAGMA table_info(users)")}
    if "display_name" not in user_cols:
        conn.execute("ALTER TABLE users ADD COLUMN display_name TEXT")
        conn.execute("UPDATE users SET display_name = username WHERE display_name IS NULL OR display_name = ''")
    if "sheets" not in existing_tables:
        conn.execute(
            """
            CREATE TABLE sheets (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL,
                sort_order INTEGER NOT NULL DEFAULT 1,
                site_id INTEGER,
                created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
            )
            """
        )
    sheet = conn.execute("SELECT id FROM sheets ORDER BY sort_order, id LIMIT 1").fetchone()
    if not sheet:
        default_site_id = ensure_site_foundation_schema(conn)
        cur = conn.execute(
            "INSERT INTO sheets (name, sort_order, site_id) VALUES (?, ?, ?)",
            (get_setting(conn, "tab_title"), 1, default_site_id),
        )
        default_sheet_id = cur.lastrowid
    else:
        default_sheet_id = sheet["id"]
    if "sheet_id" not in task_cols:
        conn.execute("ALTER TABLE tasks ADD COLUMN sheet_id INTEGER")
        conn.execute("UPDATE tasks SET sheet_id = ?", (default_sheet_id,))
    if "sheet_id" not in floor_cols:
        conn.execute("ALTER TABLE floors ADD COLUMN sheet_id INTEGER")
        conn.execute("UPDATE floors SET sheet_id = ?", (default_sheet_id,))
    conn.executescript(
        """
        CREATE TABLE IF NOT EXISTS vendor_work_entries (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            sheet_id INTEGER NOT NULL,
            vendor_name TEXT NOT NULL,
            business_date TEXT NOT NULL,
            planned_at TEXT NOT NULL DEFAULT '',
            planned_headcount INTEGER NOT NULL DEFAULT 0,
            actual_headcount INTEGER NOT NULL DEFAULT 0,
            work_content TEXT NOT NULL DEFAULT '',
            work_headcount INTEGER NOT NULL DEFAULT 0,
            entry_order INTEGER NOT NULL DEFAULT 0,
            pre_entry_requirement TEXT,
            requirement_status TEXT DEFAULT 'pending',
            requirement_confirmed_by TEXT,
            requirement_confirmed_at TEXT,
            created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
            updated_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (sheet_id) REFERENCES sheets(id)
        );

        CREATE TABLE IF NOT EXISTS formal_approvals (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            entry_id INTEGER NOT NULL,
            sheet_id INTEGER NOT NULL,
            action TEXT NOT NULL DEFAULT 'crew_formal_approve_entry',
            approval_status TEXT NOT NULL DEFAULT 'approved',
            approved_by TEXT,
            approved_at TEXT,
            created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
            updated_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (entry_id) REFERENCES vendor_work_entries(id),
            FOREIGN KEY (sheet_id) REFERENCES sheets(id)
        );

        CREATE TABLE IF NOT EXISTS scheduling_entries (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            entry_id INTEGER NOT NULL,
            sheet_id INTEGER NOT NULL,
            action TEXT NOT NULL DEFAULT 'schedule_entry',
            schedule_status TEXT NOT NULL DEFAULT 'scheduled',
            scheduled_date TEXT NOT NULL DEFAULT '',
            scheduled_time TEXT NOT NULL DEFAULT '',
            scheduled_by TEXT,
            scheduled_at TEXT,
            created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
            updated_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (entry_id) REFERENCES vendor_work_entries(id),
            FOREIGN KEY (sheet_id) REFERENCES sheets(id)
        );

        CREATE INDEX IF NOT EXISTS idx_vendor_work_entries_sheet_business_date
        ON vendor_work_entries (sheet_id, business_date);

        CREATE INDEX IF NOT EXISTS idx_vendor_work_entries_sheet_vendor_date
        ON vendor_work_entries (sheet_id, vendor_name, business_date);

        CREATE INDEX IF NOT EXISTS idx_vendor_work_entries_business_date
        ON vendor_work_entries (business_date);

        CREATE UNIQUE INDEX IF NOT EXISTS idx_formal_approvals_entry_action_unique
        ON formal_approvals (entry_id, action);

        CREATE INDEX IF NOT EXISTS idx_formal_approvals_sheet_id
        ON formal_approvals (sheet_id);

        CREATE UNIQUE INDEX IF NOT EXISTS idx_scheduling_entries_entry_action_unique
        ON scheduling_entries (entry_id, action);

        CREATE INDEX IF NOT EXISTS idx_scheduling_entries_sheet_id
        ON scheduling_entries (sheet_id);

        CREATE INDEX IF NOT EXISTS idx_scheduling_entries_scheduled_date
        ON scheduling_entries (scheduled_date);
        """
    )
    ensure_vendor_work_entries_schema(conn)
    ensure_formal_approvals_schema(conn)
    ensure_scheduling_entries_schema(conn)
    ensure_site_foundation_schema(conn)
    ensure_vendor_contacts_schema(conn)


def normalize_progress_values(conn: sqlite3.Connection) -> None:
    rows = conn.execute("SELECT unit_id, task_id, value FROM progress").fetchall()
    for row in rows:
        value = row["value"]
        if value in ("1", "O"):
            new_value = DONE_VALUE
        else:
            new_value = WORKING_VALUE
        if new_value != value:
            conn.execute(
                "UPDATE progress SET value = ? WHERE unit_id = ? AND task_id = ?",
                (new_value, row["unit_id"], row["task_id"]),
            )


def migrate_unit_layout(conn: sqlite3.Connection) -> None:
    version = conn.execute("SELECT value FROM meta WHERE key = 'unit_layout_version'").fetchone()
    if version and version["value"] == "2026-06-26-ab":
        return

    tasks = conn.execute("SELECT id FROM tasks ORDER BY col_index").fetchall()
    floors = conn.execute("SELECT * FROM floors ORDER BY sort_order").fetchall()

    for floor in floors:
        desired = desired_units_for_floor(floor["name"])
        old_units = conn.execute(
            "SELECT * FROM units WHERE floor_id = ? ORDER BY sort_order", (floor["id"],)
        ).fetchall()
        old_by_name = {unit["name"]: unit for unit in old_units}
        old_progress: dict[tuple[str, int], str] = {}
        old_extra: dict[str, sqlite3.Row] = {}
        for unit in old_units:
            for progress in conn.execute("SELECT task_id, value FROM progress WHERE unit_id = ?", (unit["id"],)):
                old_progress[(unit["name"], progress["task_id"])] = progress["value"]
            extra = conn.execute("SELECT * FROM unit_extra WHERE unit_id = ?", (unit["id"],)).fetchone()
            if extra:
                old_extra[unit["name"]] = extra

        conn.execute("DELETE FROM progress WHERE unit_id IN (SELECT id FROM units WHERE floor_id = ?)", (floor["id"],))
        conn.execute("DELETE FROM unit_extra WHERE unit_id IN (SELECT id FROM units WHERE floor_id = ?)", (floor["id"],))
        conn.execute("DELETE FROM units WHERE floor_id = ?", (floor["id"],))

        for order, unit_name in enumerate(desired, start=1):
            cur = conn.execute(
                "INSERT INTO units (floor_id, sort_order, name) VALUES (?, ?, ?)",
                (floor["id"], order, unit_name),
            )
            unit_id = cur.lastrowid
            for task in tasks:
                old_value = old_progress.get((unit_name, task["id"]), WORKING_VALUE)
                value = DONE_VALUE if old_value in ("1", "O") else WORKING_VALUE
                conn.execute(
                    "INSERT INTO progress (unit_id, task_id, value) VALUES (?, ?, ?)",
                    (unit_id, task["id"], value),
                )

            extra = old_extra.get(unit_name)
            if extra:
                conn.execute(
                    """
                    INSERT INTO unit_extra
                    (unit_id, initial_check, recheck_1, recheck_2, handover, updated_by, updated_at)
                    VALUES (?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        unit_id,
                        extra["initial_check"],
                        extra["recheck_1"],
                        extra["recheck_2"],
                        DONE_VALUE if extra["handover"] in ("1", "O") else WORKING_VALUE,
                        extra["updated_by"],
                        extra["updated_at"],
                    ),
                )
            else:
                conn.execute("INSERT INTO unit_extra (unit_id, handover) VALUES (?, ?)", (unit_id, WORKING_VALUE))

        conn.execute(
            "UPDATE floors SET block_name = ?, unit_count = ? WHERE id = ?",
            (block_for_floor(floor["name"]), len(desired), floor["id"]),
        )

    set_setting(conn, "unit_layout_version", "2026-06-26-ab")


def ensure_unit_extra_rows(conn: sqlite3.Connection) -> None:
    for unit in conn.execute("SELECT id FROM units"):
        conn.execute(
            "INSERT OR IGNORE INTO unit_extra (unit_id, handover) VALUES (?, ?)",
            (unit["id"], WORKING_VALUE),
        )


def ensure_extra_fields(conn: sqlite3.Connection) -> None:
    for sheet in conn.execute("SELECT id FROM sheets"):
        for field_key, field in BUILTIN_EXTRA_FIELDS.items():
            conn.execute(
                """
                INSERT OR IGNORE INTO extra_fields
                (sheet_id, field_key, name, field_type, sort_order, is_builtin, active)
                VALUES (?, ?, ?, ?, ?, 1, 1)
                """,
                (sheet["id"], field_key, field["name"], field["type"], field["sort_order"]),
            )


def bootstrap() -> None:
    with db() as conn:
        init_schema(conn)
        seed_admin(conn)
        seed_settings(conn)
        seed_from_excel(conn)
        migrate_schema(conn)
        ensure_site_foundation_schema(conn)
        normalize_progress_values(conn)
        ensure_unit_extra_rows(conn)
        ensure_extra_fields(conn)
        migrate_unit_layout(conn)


def available_sheets(conn: sqlite3.Connection) -> list[sqlite3.Row]:
    _user, current_site_id, is_admin = _resolve_read_scope(conn)
    if is_admin:
        return conn.execute("SELECT * FROM sheets ORDER BY sort_order, id").fetchall()
    return conn.execute(
        "SELECT * FROM sheets WHERE site_id = ? ORDER BY sort_order, id",
        (current_site_id,),
    ).fetchall()


def resolve_sheet_id(conn: sqlite3.Connection, sheet_id: int | None = None) -> int:
    _user, current_site_id, is_admin = _resolve_read_scope(conn)
    if is_admin:
        if sheet_id:
            row = conn.execute("SELECT id FROM sheets WHERE id = ?", (sheet_id,)).fetchone()
            if row:
                return row["id"]
        row = conn.execute("SELECT id FROM sheets ORDER BY sort_order, id LIMIT 1").fetchone()
        if row:
            return row["id"]
        default_site_id = ensure_site_foundation_schema(conn)
        cur = conn.execute(
            "INSERT INTO sheets (name, sort_order, site_id) VALUES (?, ?, ?)",
            (get_setting(conn, "tab_title"), 1, default_site_id),
        )
        return cur.lastrowid

    if sheet_id is not None:
        row = conn.execute("SELECT id, site_id FROM sheets WHERE id = ?", (sheet_id,)).fetchone()
        if row is None:
            raise LookupError("sheet_not_found")
        if int(row["site_id"] or 0) != int(current_site_id):
            raise LookupError("sheet_not_in_current_site")
        return int(row["id"])

    row = conn.execute(
        "SELECT id FROM sheets WHERE site_id = ? ORDER BY sort_order, id LIMIT 1",
        (current_site_id,),
    ).fetchone()
    if row:
        return int(row["id"])
    raise LookupError("no_sheets_in_current_site")


def load_grid(sheet_id: int | None = None) -> dict:
    with db() as conn:
        current_sheet_id = resolve_sheet_id(conn, sheet_id)
        settings = get_settings(conn)
        sheets = available_sheets(conn)
        current_sheet = conn.execute("SELECT * FROM sheets WHERE id = ?", (current_sheet_id,)).fetchone()
        tasks = conn.execute("SELECT * FROM tasks WHERE sheet_id = ? ORDER BY col_index", (current_sheet_id,)).fetchall()
        floors = conn.execute("SELECT * FROM floors WHERE sheet_id = ? ORDER BY sort_order", (current_sheet_id,)).fetchall()
        units_by_floor = {
            floor["id"]: conn.execute(
                "SELECT * FROM units WHERE floor_id = ? ORDER BY sort_order", (floor["id"],)
            ).fetchall()
            for floor in floors
        }
        progress_rows = conn.execute(
            """
            SELECT p.unit_id, p.task_id, p.value
            FROM progress p
            JOIN units u ON u.id = p.unit_id
            JOIN floors f ON f.id = u.floor_id
            WHERE f.sheet_id = ?
            """,
            (current_sheet_id,),
        ).fetchall()
        extra_rows = conn.execute(
            """
            SELECT ue.*
            FROM unit_extra ue
            JOIN units u ON u.id = ue.unit_id
            JOIN floors f ON f.id = u.floor_id
            WHERE f.sheet_id = ?
            """,
            (current_sheet_id,),
        ).fetchall()
        extra_fields = conn.execute(
            "SELECT * FROM extra_fields WHERE sheet_id = ? AND active = 1 ORDER BY sort_order, id",
            (current_sheet_id,),
        ).fetchall()
        extra_value_rows = conn.execute(
            """
            SELECT v.unit_id, v.field_key, v.value
            FROM unit_extra_values v
            JOIN units u ON u.id = v.unit_id
            JOIN floors f ON f.id = u.floor_id
            WHERE f.sheet_id = ?
            """,
            (current_sheet_id,),
        ).fetchall()

    progress = {(row["unit_id"], row["task_id"]): row["value"] for row in progress_rows}
    extras = {row["unit_id"]: dict(row) for row in extra_rows}
    for row in extra_value_rows:
        extras.setdefault(row["unit_id"], {})[row["field_key"]] = row["value"]
    floor_rows = []
    summary = {task["id"]: {"done": 0, "total": 0} for task in tasks}
    extra_summary = {field["field_key"]: {"done": 0, "total": 0} for field in extra_fields}

    for floor in floors:
        units = units_by_floor[floor["id"]]
        parent_status = {}
        for task in tasks:
            values = [progress.get((unit["id"], task["id"]), WORKING_VALUE) for unit in units]
            done_count = sum(1 for value in values if value == DONE_VALUE)
            parent_status[task["id"]] = DONE_VALUE if units and done_count == len(units) else WORKING_VALUE
            summary[task["id"]]["done"] += done_count
            summary[task["id"]]["total"] += len(units)
        for unit in units:
            extra = extras.get(unit["id"], {})
            for field in extra_fields:
                field_key = field["field_key"]
                extra_summary[field_key]["done"] += 1 if extra_done(field, extra) else 0
                extra_summary[field_key]["total"] += 1
        floor_rows.append({"floor": floor, "units": units, "parent_status": parent_status})

    return {
        "settings": settings,
        "sheets": sheets,
        "current_sheet": current_sheet,
        "tasks": tasks,
        "extra_fields": extra_fields,
        "floor_rows": floor_rows,
        "progress": progress,
        "extras": extras,
        "summary": summary,
        "extra_summary": extra_summary,
    }


def extra_done(field: dict | sqlite3.Row, extra: dict) -> bool:
    field_key = field["field_key"]
    if field_key == "initial_check":
        return bool(extra.get("recheck_1")) or extra.get("handover") == DONE_VALUE
    if field_key == "recheck_1":
        return bool(extra.get("recheck_2")) or extra.get("handover") == DONE_VALUE
    if field_key == "recheck_2":
        return bool(extra.get("recheck_2")) or extra.get("handover") == DONE_VALUE
    if field_key == "handover":
        return extra.get("handover") == DONE_VALUE
    if field["field_type"] == "status":
        return extra.get(field_key) == DONE_VALUE
    return bool(extra.get(field_key))


def render_grid_payload(sheet_id: int | None = None) -> dict:
    grid = load_grid(sheet_id)
    floors = []
    for item in grid["floor_rows"]:
        floor = dict(item["floor"])
        units = [dict(unit) for unit in item["units"]]
        floors.append({"floor": floor, "units": units, "parent_status": item["parent_status"]})
    return {
        "settings": grid["settings"],
        "sheets": [dict(sheet) for sheet in grid["sheets"]],
        "current_sheet": dict(grid["current_sheet"]) if grid["current_sheet"] else None,
        "tasks": [dict(task) for task in grid["tasks"]],
        "extra_fields": [dict(field) for field in grid["extra_fields"]],
        "floors": floors,
        "progress": {f"{unit_id}:{task_id}": value for (unit_id, task_id), value in grid["progress"].items()},
        "extras": {str(unit_id): extra for unit_id, extra in grid["extras"].items()},
        "summary": grid["summary"],
        "extra_summary": grid["extra_summary"],
    }


def build_unit_handover_report_context(conn: sqlite3.Connection, *, unit_id: int) -> dict[str, object]:
    unit_context = conn.execute(
        """
        SELECT u.name AS unit_name,
               f.name AS floor_name,
               f.block_name AS block_name,
               s.id AS sheet_id,
               s.name AS sheet_name,
               st.site_name AS site_name,
               st.site_code AS site_code
        FROM units u
        JOIN floors f ON f.id = u.floor_id
        JOIN sheets s ON s.id = f.sheet_id
        JOIN sites st ON st.id = s.site_id
        WHERE u.id = ?
        """,
        (unit_id,),
    ).fetchone()
    if unit_context is None:
        raise LookupError("unit_not_found")

    authorize_dashboard_read(conn, sheet_id=int(unit_context["sheet_id"]))

    task_rows = conn.execute(
        """
        SELECT t.id,
               t.col_index,
               t.vendor,
               t.location,
               t.name,
               COALESCE(p.value, ?) AS progress_value
        FROM tasks t
        LEFT JOIN progress p
               ON p.task_id = t.id
              AND p.unit_id = ?
        WHERE t.sheet_id = ?
        ORDER BY t.col_index, t.id
        """,
        (WORKING_VALUE, unit_id, int(unit_context["sheet_id"])),
    ).fetchall()
    extra_field_rows = conn.execute(
        """
        SELECT id, field_key, name, field_type, sort_order, is_builtin
        FROM extra_fields
        WHERE sheet_id = ? AND active = 1
        ORDER BY sort_order, id
        """,
        (int(unit_context["sheet_id"]),),
    ).fetchall()
    unit_extra_row = conn.execute("SELECT * FROM unit_extra WHERE unit_id = ?", (unit_id,)).fetchone()
    extra_values = dict(unit_extra_row) if unit_extra_row is not None else {}
    for row in conn.execute(
        "SELECT field_key, value FROM unit_extra_values WHERE unit_id = ?",
        (unit_id,),
    ).fetchall():
        extra_values[str(row["field_key"])] = row["value"]

    return {
        "settings": get_settings(conn),
        "site": {
            "name": str(unit_context["site_name"] or ""),
            "code": str(unit_context["site_code"] or ""),
        },
        "sheet": {
            "id": int(unit_context["sheet_id"]),
            "name": str(unit_context["sheet_name"] or ""),
        },
        "floor": {
            "name": str(unit_context["floor_name"] or ""),
            "block_name": str(unit_context["block_name"] or ""),
        },
        "unit": {
            "name": str(unit_context["unit_name"] or ""),
        },
        "tasks": [dict(row) for row in task_rows],
        "extra_fields": [dict(row) for row in extra_field_rows],
        "extra_values": extra_values,
    }


def build_floor_handover_report_context(conn: sqlite3.Connection, *, floor_id: int) -> dict[str, object]:
    floor_context = conn.execute(
        """
        SELECT f.name AS floor_name,
               f.block_name AS block_name,
               s.id AS sheet_id,
               s.name AS sheet_name,
               st.site_name AS site_name,
               st.site_code AS site_code
        FROM floors f
        JOIN sheets s ON s.id = f.sheet_id
        JOIN sites st ON st.id = s.site_id
        WHERE f.id = ?
        """,
        (floor_id,),
    ).fetchone()
    if floor_context is None:
        raise LookupError("floor_not_found")

    authorize_dashboard_read(conn, sheet_id=int(floor_context["sheet_id"]))

    unit_rows = conn.execute(
        "SELECT id, sort_order, name FROM units WHERE floor_id = ? ORDER BY sort_order, id",
        (floor_id,),
    ).fetchall()
    task_rows = conn.execute(
        """
        SELECT id, col_index, vendor, location, name
        FROM tasks
        WHERE sheet_id = ?
        ORDER BY col_index, id
        """,
        (int(floor_context["sheet_id"]),),
    ).fetchall()
    extra_field_rows = conn.execute(
        """
        SELECT id, field_key, name, field_type, sort_order, is_builtin
        FROM extra_fields
        WHERE sheet_id = ? AND active = 1
        ORDER BY sort_order, id
        """,
        (int(floor_context["sheet_id"]),),
    ).fetchall()
    progress_rows = conn.execute(
        """
        SELECT p.unit_id, p.task_id, p.value
        FROM progress p
        JOIN units u ON u.id = p.unit_id
        JOIN tasks t ON t.id = p.task_id
        WHERE u.floor_id = ? AND t.sheet_id = ?
        """,
        (floor_id, int(floor_context["sheet_id"])),
    ).fetchall()
    unit_extra_rows = conn.execute(
        """
        SELECT ue.*
        FROM unit_extra ue
        JOIN units u ON u.id = ue.unit_id
        WHERE u.floor_id = ?
        """,
        (floor_id,),
    ).fetchall()
    extra_value_rows = conn.execute(
        """
        SELECT v.unit_id, v.field_key, v.value
        FROM unit_extra_values v
        JOIN units u ON u.id = v.unit_id
        WHERE u.floor_id = ?
        """,
        (floor_id,),
    ).fetchall()

    progress = {(int(row["unit_id"]), int(row["task_id"])): row["value"] for row in progress_rows}
    extras = {int(row["unit_id"]): dict(row) for row in unit_extra_rows}
    for row in extra_value_rows:
        extras.setdefault(int(row["unit_id"]), {})[str(row["field_key"])] = row["value"]

    return {
        "settings": get_settings(conn),
        "site": {
            "name": str(floor_context["site_name"] or ""),
            "code": str(floor_context["site_code"] or ""),
        },
        "sheet": {
            "id": int(floor_context["sheet_id"]),
            "name": str(floor_context["sheet_name"] or ""),
        },
        "floor": {
            "name": str(floor_context["floor_name"] or ""),
            "block_name": str(floor_context["block_name"] or ""),
        },
        "units": [dict(row) for row in unit_rows],
        "tasks": [dict(row) for row in task_rows],
        "extra_fields": [dict(row) for row in extra_field_rows],
        "progress": progress,
        "extras": extras,
    }


@app.route("/")
def index():
    if not session.get("user_id"):
        return redirect(url_for("login"))
    return redirect(url_for("sheet"))


def _render_vendor_login_page(*, error_message: str = "") -> str:
    return render_template(
        "vendor_login.html",
        settings=query_settings(),
        error_message=error_message,
        is_authenticated=is_vendor_session(),
        vendor_name=session.get("vendor_name"),
        vendor_username=session.get("vendor_username"),
    )


@app.route("/login", methods=["GET", "POST"])
def login():
    settings = query_settings()
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        display_name = request.form.get("display_name", "").strip() or username
        password = request.form.get("password", "")
        user = get_user_by_username(username)
        if user and check_password_hash(user["password_hash"], password):
            session.clear()
            session["user_id"] = user["id"]
            session["username"] = user["username"]
            session["display_name"] = user["display_name"] or user["username"]
            session["role"] = user["role"]
            resolution = normalize_current_site_for_user(user)
            status = str(resolution["status"])
            if status == "access_denied_no_site_permission":
                app.logger.info("zero_site_login_blocked user_id=%s", user["id"])
                session.clear()
                flash("您目前沒有可進入的工地", "error")
                return redirect(url_for("login"))
            if status == "site_selection_required":
                app.logger.info("login_requires_site_selection user_id=%s", user["id"])
            elif status == "resolved":
                app.logger.info(
                    "login_auto_selected_site user_id=%s site_id=%s reason=%s",
                    user["id"],
                    resolution["site_id"],
                    resolution.get("reason", ""),
                )
            return redirect(_site_selection_redirect_target(resolution))
        flash("帳號或密碼錯誤。", "error")
    return render_template("login.html", settings=settings)


@app.route("/vendor/login", methods=["GET", "POST"])
def vendor_login():
    if request.method == "POST":
        session.clear()
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "")
        vendor_account = verify_vendor_account(username, password)
        if vendor_account is not None:
            set_vendor_session(vendor_account)
            return redirect(url_for("vendor_login"))
        return _render_vendor_login_page(error_message="Invalid vendor username or password.")
    return _render_vendor_login_page()


def query_settings() -> dict[str, str]:
    with db() as conn:
        return get_settings(conn)


@app.route("/logout", methods=["POST"])
@login_required
def logout():
    session.clear()
    return redirect(url_for("login"))


@app.route("/vendor/logout", methods=["GET"])
def vendor_logout():
    session.clear()
    return redirect(url_for("vendor_login"))


@app.route("/vendor/home", methods=["GET"])
@vendor_login_required
def vendor_home():
    return vendor_work_entry_page()


@app.route("/vendor/work-entry", methods=["GET"])
@vendor_login_required
def vendor_work_entry_page():
    vendor_account = require_current_vendor_account()
    business_identity = authorize_vendor_business_read()
    scope = current_vendor_scope()
    if scope is None:
        return redirect(url_for("vendor_login"))

    business_date = resolve_crew_business_date()
    with db() as conn:
        page_context = build_vendor_work_entry_page_context(
            conn,
            vendor_account,
            business_identity=business_identity,
            scope=scope,
            business_date=business_date,
            today_entry_index_raw=request.args.get("today_entry_index", ""),
            selected_entry_id_raw=request.args.get("selected_entry_id", ""),
            new_entry_raw=request.args.get("new_entry", ""),
            submit_status_raw=request.args.get("submit_status", ""),
            submit_mode_raw=request.args.get("submit_mode", ""),
        )

    return render_template(
        "vendor_work_entry.html",
        settings=query_settings(),
        **page_context,
    )


@app.route("/vendor/profile", methods=["GET"])
@vendor_login_required
def vendor_profile():
    vendor_account = require_current_vendor_account()
    return jsonify(
        {
            "ok": True,
            "vendor_account_id": int(vendor_account["id"]),
            "vendor_username": str(vendor_account["username"]),
            "vendor_name": str(vendor_account["vendor_name"]),
        }
    )


@app.route("/vendor/scope", methods=["GET"])
@vendor_login_required
def vendor_scope():
    scope = current_vendor_scope()
    if scope is None:
        return redirect(url_for("vendor_login"))
    return jsonify({"ok": True, "scope": scope})


@app.route("/vendor/business-read-preview", methods=["GET"])
@vendor_login_required
def vendor_business_read_preview():
    business_identity = authorize_vendor_business_read()

    with db() as conn:
        rows = fetch_vendor_business_read_preview(
            conn,
            vendor_name=str(business_identity["vendor_name"]),
        )

    return jsonify(
        serialize_vendor_business_read_preview(
            business_identity=business_identity,
            rows=rows,
        )
    )


@app.route("/site-selector", methods=["GET", "POST"])
@login_required
def site_selector():
    user = _current_internal_user()
    if user is None:
        session.clear()
        return redirect(url_for("login"))

    resolution = resolve_current_site_for_user(user)
    if resolution["status"] == "access_denied_no_site_permission":
        app.logger.info("zero_site_login_blocked user_id=%s route=site_selector", user["id"])
        session.clear()
        flash("您目前沒有可進入的工地", "error")
        return redirect(url_for("login"))

    if request.method == "POST":
        selector_required = session.get("site_selection_required") is True
        raw_site_id = request.form.get("site_id", "").strip()
        try:
            site_id = int(raw_site_id)
        except (TypeError, ValueError):
            if selector_required:
                clear_current_site_selection_context()
            app.logger.info(
                "invalid_site_selection user_id=%s site_id=%s reason=invalid_integer",
                user["id"],
                raw_site_id,
            )
            return _render_site_selector(user=user, status_code=400, error_message="請選擇有效的工地")

        if not user_can_access_site(int(user["id"]), site_id):
            if selector_required:
                clear_current_site_selection_context()
            app.logger.info(
                "invalid_site_selection user_id=%s site_id=%s reason=not_accessible",
                user["id"],
                site_id,
            )
            return _render_site_selector(user=user, status_code=403, error_message="你沒有這個工地的使用權限")

        set_current_site_id(site_id)
        session["site_selection_required"] = False
        app.logger.info("site_switched user_id=%s site_id=%s", user["id"], site_id)
        return redirect(url_for("sheet"))

    return _render_site_selector(user=user)


@app.route("/sheet")
@app.route("/sheet/<int:sheet_id>")
@login_required
def sheet(sheet_id: int | None = None):
    if session.get("site_selection_required") is True:
        app.logger.info(
            "current_site_cleared user_id=%s reason=sheet_recovery_redirect",
            session.get("user_id"),
        )
        return redirect(url_for("site_selector"))
    try:
        with db() as conn:
            resolved = resolve_sheet_id(conn, sheet_id)
        session["sheet_id"] = resolved
        grid = load_grid(resolved)
    except LookupError as exc:
        return _handle_sheet_read_lookup_error(exc)
    return render_template(
        "sheet.html",
        grid=grid,
        settings=grid["settings"],
        done_value=DONE_VALUE,
        working_value=WORKING_VALUE,
    )


@app.route("/reports/unit/<int:unit_id>", methods=["GET"])
def unit_handover_report(unit_id: int):
    if current_vendor_account() is not None:
        return _handle_unit_handover_report_lookup_error(LookupError("vendor_auth_forbidden"))
    user = _current_internal_user()
    if user is None:
        return _handle_unit_handover_report_lookup_error(LookupError("auth_required"))

    try:
        with db() as conn:
            report = build_unit_handover_report_context(conn, unit_id=unit_id)
    except LookupError as exc:
        return _handle_unit_handover_report_lookup_error(exc)

    report["generated_at"] = datetime.now().astimezone().strftime("%Y-%m-%d %H:%M %Z")
    report["generated_by"] = str(user["display_name"] or user["username"] or "")
    return render_template(
        "unit_handover_report.html",
        report=report,
        settings=report["settings"],
        done_value=DONE_VALUE,
    )


@app.route("/reports/floor/<int:floor_id>", methods=["GET"])
def floor_handover_report(floor_id: int):
    if current_vendor_account() is not None:
        return _handle_floor_handover_report_lookup_error(LookupError("vendor_auth_forbidden"))
    user = _current_internal_user()
    if user is None:
        return _handle_floor_handover_report_lookup_error(LookupError("auth_required"))

    try:
        with db() as conn:
            report = build_floor_handover_report_context(conn, floor_id=floor_id)
    except LookupError as exc:
        return _handle_floor_handover_report_lookup_error(exc)

    report["generated_at"] = datetime.now().astimezone().strftime("%Y-%m-%d %H:%M %Z")
    report["generated_by"] = str(user["display_name"] or user["username"] or "")
    return render_template(
        "floor_handover_report.html",
        report=report,
        settings=report["settings"],
        done_value=DONE_VALUE,
        working_value=WORKING_VALUE,
    )


@app.route("/api/grid")
@login_required
def api_grid():
    sheet_id = request.args.get("sheet_id", type=int) or session.get("sheet_id")
    try:
        return jsonify(render_grid_payload(sheet_id))
    except LookupError as exc:
        return _handle_grid_read_lookup_error(exc)


@app.route("/api/progress", methods=["POST"])
@login_required
def api_progress():
    data = request.get_json(force=True)
    unit_id = int(data.get("unit_id"))
    task_id = int(data.get("task_id"))
    value = data.get("value", WORKING_VALUE)
    if value not in (DONE_VALUE, WORKING_VALUE):
        return jsonify({"ok": False, "message": "狀態只能是 O 或 X。"}), 400
    with db() as conn:
        try:
            progress_context = authorize_progress_write(conn, unit_id=unit_id, task_id=task_id)
        except LookupError as exc:
            return _handle_progress_write_lookup_error(exc)
        conn.execute(
            """
            INSERT INTO progress (unit_id, task_id, value, updated_by, updated_at)
            VALUES (?, ?, ?, ?, CURRENT_TIMESTAMP)
            ON CONFLICT(unit_id, task_id) DO UPDATE SET
                value = excluded.value,
                updated_by = excluded.updated_by,
                updated_at = CURRENT_TIMESTAMP
            """,
            (unit_id, task_id, value, session["user_id"]),
        )
    return jsonify({"ok": True, "grid": render_grid_payload(progress_context["sheet_id"])})


@app.route("/api/unit-extra", methods=["POST"])
@login_required
def api_unit_extra():
    data = request.get_json(force=True)
    unit_id = int(data.get("unit_id"))
    field = data.get("field", "")
    value = data.get("value", "")
    with db() as conn:
        try:
            unit_extra_context = authorize_unit_extra_write(conn, unit_id=unit_id, field_key=field)
        except LookupError as exc:
            return _handle_unit_extra_write_lookup_error(exc)
        if unit_extra_context["field_type"] == "status" and value not in (DONE_VALUE, WORKING_VALUE):
            return jsonify({"ok": False, "message": "O/X 欄位只能是 O 或 X。"}), 400
        conn.execute(
            "INSERT OR IGNORE INTO unit_extra (unit_id, handover) VALUES (?, ?)",
            (unit_id, WORKING_VALUE),
        )
        if field in EXTRA_FIELDS:
            conn.execute(
                f"""
                UPDATE unit_extra
                SET {field} = ?, updated_by = ?, updated_at = CURRENT_TIMESTAMP
                WHERE unit_id = ?
                """,
                (value, session["user_id"], unit_id),
            )
        else:
            conn.execute(
                """
                INSERT INTO unit_extra_values (unit_id, field_key, value, updated_by, updated_at)
                VALUES (?, ?, ?, ?, CURRENT_TIMESTAMP)
                ON CONFLICT(unit_id, field_key) DO UPDATE SET
                    value = excluded.value,
                    updated_by = excluded.updated_by,
                    updated_at = CURRENT_TIMESTAMP
                """,
                (unit_id, field, value, session["user_id"]),
            )
    return jsonify({"ok": True, "grid": render_grid_payload(int(unit_extra_context["sheet_id"]))})


@app.route("/api/reset-sheet", methods=["POST"])
@admin_required
def api_reset_sheet():
    data = request.get_json(force=True)
    password = data.get("password", "")
    user = get_user_by_id(session["user_id"])
    if not user or not check_password_hash(user["password_hash"], password):
        return jsonify({"ok": False, "message": "管理員密碼錯誤。"}), 403

    raw_sheet_id = data.get("sheet_id", session.get("sheet_id"))
    try:
        if raw_sheet_id in (None, ""):
            raise LookupError("invalid_request")
        sheet_id = int(raw_sheet_id)
    except LookupError as exc:
        return _handle_admin_reset_sheet_lookup_error(exc)
    except (TypeError, ValueError):
        return progress_api_error("sheet_id 無效。", status=400)

    with db() as conn:
        try:
            authorize_admin_site_scoped_write(conn, sheet_id=sheet_id)
        except LookupError as exc:
            return _handle_admin_reset_sheet_lookup_error(exc)
        conn.execute(
            """
            UPDATE progress
            SET value = ?, updated_by = ?, updated_at = CURRENT_TIMESTAMP
            WHERE task_id IN (SELECT id FROM tasks WHERE sheet_id = ?)
            """,
            (WORKING_VALUE, session["user_id"], sheet_id),
        )
        conn.execute(
            """
            UPDATE unit_extra
            SET initial_check = '',
                recheck_1 = '',
                recheck_2 = '',
                handover = ?,
                updated_by = ?,
                updated_at = CURRENT_TIMESTAMP
            WHERE unit_id IN (
                SELECT u.id
                FROM units u
                JOIN floors f ON f.id = u.floor_id
                WHERE f.sheet_id = ?
            )
            """,
            (WORKING_VALUE, session["user_id"], sheet_id),
        )
        conn.execute(
            """
            DELETE FROM unit_extra_values
            WHERE unit_id IN (
                SELECT u.id
                FROM units u
                JOIN floors f ON f.id = u.floor_id
                WHERE f.sheet_id = ?
            )
            AND field_key IN (SELECT field_key FROM extra_fields WHERE sheet_id = ?)
            """,
            (sheet_id, sheet_id),
        )
    return jsonify({"ok": True, "grid": render_grid_payload(sheet_id)})


@app.route("/api/crew-forms")
@login_required
def api_crew_forms():
    sheet_id = request.args.get("sheet_id", type=int)
    if sheet_id is None:
        return crew_api_error("invalid_sheet_id", "sheet_id is required and must be a valid integer.")

    raw_business_date = request.args.get("business_date", "").strip()
    try:
        business_date = parse_crew_business_date(raw_business_date) if raw_business_date else resolve_crew_business_date()
    except ValueError as exc:
        return crew_api_error("invalid_business_date", str(exc))

    with db() as conn:
        try:
            authorize_sheet_read(conn, sheet_id)
        except LookupError as exc:
            return _handle_grid_read_lookup_error(exc)

        active_vendors = get_active_crew_vendors(sheet_id)
        active_vendor_set = set(active_vendors)
        contacts_by_vendor = fetch_vendor_contacts_grouped(conn, sheet_id)
        entries_by_vendor = fetch_vendor_work_entries(conn, sheet_id=sheet_id, business_date=business_date)

        active_vendor_payload = []
        for vendor_name in active_vendors:
            contacts = contacts_by_vendor.get(vendor_name, [])
            contact = first_or_empty_vendor_contact(contacts, sheet_id=sheet_id, vendor_name=vendor_name)
            active_vendor_payload.append(
                {
                    "vendor_name": vendor_name,
                    "contact": contact,
                    "contacts": contacts,
                    "work_entries": entries_by_vendor.get(vendor_name, []),
                }
            )

        inactive_contacts = [
            {
                "vendor_name": vendor_name,
                "contact": first_or_empty_vendor_contact(contacts, sheet_id=sheet_id, vendor_name=vendor_name),
                "contacts": contacts,
            }
            for vendor_name, contacts in contacts_by_vendor.items()
            if vendor_name not in active_vendor_set
        ]

    return jsonify(
        {
            "ok": True,
            "sheet_id": sheet_id,
            "business_date": business_date,
            "active_vendors": active_vendor_payload,
            "inactive_contacts": inactive_contacts,
        }
    )


@app.route("/api/dashboard")
def api_dashboard():
    sheet_id = request.args.get("sheet_id", type=int)
    if sheet_id is None:
        return crew_api_error("invalid_sheet_id", "sheet_id is required and must be a valid integer.")

    raw_business_date = request.args.get("business_date", "").strip()
    try:
        business_date = parse_crew_business_date(raw_business_date) if raw_business_date else resolve_crew_business_date()
    except ValueError as exc:
        return crew_api_error("invalid_business_date", str(exc))

    with db() as conn:
        try:
            dashboard_context = authorize_dashboard_read(conn, sheet_id=sheet_id)
        except LookupError as exc:
            return _handle_dashboard_lookup_error(exc)
        payload = build_dashboard_payload(
            conn,
            sheet_id=int(dashboard_context["sheet_id"]),
            business_date=business_date,
        )

    return jsonify(payload)


@app.route("/api/scheduling")
def api_scheduling():
    sheet_id = request.args.get("sheet_id", type=int)
    if sheet_id is None:
        return crew_api_error("invalid_sheet_id", "sheet_id is required and must be a valid integer.")

    raw_business_date = request.args.get("business_date", "").strip()
    try:
        business_date = parse_crew_business_date(raw_business_date) if raw_business_date else resolve_crew_business_date()
    except ValueError as exc:
        return crew_api_error("invalid_business_date", str(exc))

    with db() as conn:
        try:
            scheduling_context = authorize_dashboard_read(conn, sheet_id=sheet_id)
        except LookupError as exc:
            return _handle_scheduling_lookup_error(exc)
        payload = build_scheduling_payload(
            conn,
            sheet_id=int(scheduling_context["sheet_id"]),
            business_date=business_date,
        )

    return jsonify(payload)


@app.route("/api/work-hub-runtime")
def api_work_hub_runtime():
    sheet_id = request.args.get("sheet_id", type=int)
    if sheet_id is None:
        return crew_api_error("invalid_sheet_id", "sheet_id is required and must be a valid integer.")

    raw_business_date = request.args.get("business_date", "").strip()
    try:
        business_date = parse_crew_business_date(raw_business_date) if raw_business_date else resolve_crew_business_date()
    except ValueError as exc:
        return crew_api_error("invalid_business_date", str(exc))

    with db() as conn:
        try:
            work_hub_runtime_context = authorize_dashboard_read(conn, sheet_id=sheet_id)
        except LookupError as exc:
            return _handle_work_hub_runtime_lookup_error(exc)
        payload = build_work_hub_runtime_payload(
            conn,
            sheet_id=int(work_hub_runtime_context["sheet_id"]),
            business_date=business_date,
        )

    return jsonify(payload)


@app.route("/api/management-read-model")
def api_management_read_model():
    sheet_id = request.args.get("sheet_id", type=int)
    if sheet_id is None:
        return crew_api_error("invalid_sheet_id", "sheet_id is required and must be a valid integer.")

    raw_business_date = request.args.get("business_date", "").strip()
    try:
        business_date = parse_crew_business_date(raw_business_date) if raw_business_date else resolve_crew_business_date()
    except ValueError as exc:
        return crew_api_error("invalid_business_date", str(exc))

    with db() as conn:
        try:
            management_read_model_context = authorize_dashboard_read(conn, sheet_id=sheet_id)
        except LookupError as exc:
            return _handle_management_read_model_lookup_error(exc)
        payload = build_management_read_model_payload(
            conn,
            sheet_id=int(management_read_model_context["sheet_id"]),
            business_date=business_date,
        )

    return jsonify(payload)


@app.route("/api/vendor-contact", methods=["POST"])
@login_required
def api_vendor_contact():
    data = request.get_json(silent=True) or {}
    try:
        contact_id = parse_optional_positive_int(data.get("id"), field_name="id")
        sheet_id = parse_non_negative_int(data.get("sheet_id"), field_name="sheet_id")
        if sheet_id <= 0:
            raise ValueError("sheet_id must be a positive integer.")
        vendor_name = normalize_vendor_name(str(data.get("vendor_name", "")))
        contact_order = parse_optional_non_negative_int(data.get("contact_order"), field_name="contact_order")
        is_primary = parse_contact_primary_flag(data.get("is_primary", 0))
    except ValueError as exc:
        return crew_api_error("invalid_request", str(exc))

    contact_name = str(data.get("contact_name", "")).strip()
    contact_title = str(data.get("contact_title", "")).strip()
    contact_phone = str(data.get("contact_phone", "")).strip()
    if len(contact_name) > 100:
        return crew_api_error("invalid_contact_name", "contact_name must be 100 characters or fewer.")
    if len(contact_title) > 100:
        return crew_api_error("invalid_contact_title", "contact_title must be 100 characters or fewer.")
    if len(contact_phone) > 50:
        return crew_api_error("invalid_contact_phone", "contact_phone must be 50 characters or fewer.")

    with db() as conn:
        try:
            vendor_contact_context = authorize_vendor_contact_write(
                conn,
                sheet_id=sheet_id,
                vendor_name=vendor_name,
                contact_id=contact_id,
            )
        except LookupError as exc:
            return _handle_vendor_contact_lookup_error(exc)
        if contact_id is not None:
            if contact_order is None:
                contact_order = int(vendor_contact_context["existing_contact_order"] or 0)
            conn.execute(
                """
                UPDATE vendor_contacts
                SET vendor_name = ?, contact_name = ?, contact_title = ?, contact_phone = ?,
                    is_primary = ?, contact_order = ?, updated_at = CURRENT_TIMESTAMP
                WHERE id = ?
                """,
                (vendor_name, contact_name, contact_title, contact_phone, is_primary, contact_order, contact_id),
            )
            if is_primary == 1:
                set_primary_contact(conn, sheet_id=sheet_id, vendor_name=vendor_name, contact_id=contact_id)
            saved_contact_id = contact_id
        else:
            if contact_order is None:
                contact_order = next_contact_order(conn, sheet_id=sheet_id, vendor_name=vendor_name)
            cur = conn.execute(
                """
                INSERT INTO vendor_contacts (
                    sheet_id, vendor_name, contact_name, contact_title, contact_phone, is_primary, contact_order,
                    created_at, updated_at
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, CURRENT_TIMESTAMP, CURRENT_TIMESTAMP)
                """,
                (sheet_id, vendor_name, contact_name, contact_title, contact_phone, is_primary, contact_order),
            )
            saved_contact_id = cur.lastrowid
            if is_primary == 1:
                set_primary_contact(conn, sheet_id=sheet_id, vendor_name=vendor_name, contact_id=saved_contact_id)
        contacts = fetch_vendor_contacts_by_vendor(conn, sheet_id=sheet_id, vendor_name=vendor_name)
        contact = first_or_empty_vendor_contact(
            contacts,
            sheet_id=sheet_id,
            vendor_name=vendor_name,
        )

    return jsonify({"ok": True, "contact": contact, "contacts": contacts})


@app.route("/api/vendor-work-entry", methods=["POST"])
@login_required
def api_vendor_work_entry():
    data = request.get_json(silent=True) or {}
    try:
        payload = normalize_vendor_work_entry_submit_payload(data)
    except ValueError as exc:
        return crew_api_error("invalid_request", str(exc))

    entry_id = int(payload["entry_id"]) if payload["entry_id"] is not None else None
    sheet_id = int(payload["sheet_id"])
    vendor_name = str(payload["vendor_name"])
    business_date = str(payload["business_date"])
    planned_at = str(payload["planned_at"])
    planned_headcount = int(payload["planned_headcount"])
    actual_headcount = int(payload["actual_headcount"])
    work_content = str(payload["work_content"])
    pre_entry_requirement = str(payload["pre_entry_requirement"])
    work_headcount = int(payload["work_headcount"])
    entry_order = int(payload["entry_order"])

    if len(work_content) > 500:
        return crew_api_error("invalid_work_content", "work_content must be 500 characters or fewer.")
    if len(pre_entry_requirement) > 500:
        return crew_api_error(
            "invalid_pre_entry_requirement",
            "pre_entry_requirement must be 500 characters or fewer.",
        )

    with db() as conn:
        try:
            vendor_work_entry_context = authorize_vendor_work_entry_write(
                conn,
                sheet_id=sheet_id,
                vendor_name=vendor_name,
                entry_id=entry_id,
            )
        except LookupError as exc:
            return _handle_vendor_work_entry_lookup_error(exc)

        if entry_id is None:
            cur = conn.execute(
                """
                INSERT INTO vendor_work_entries (
                    sheet_id, vendor_name, business_date, planned_at, planned_headcount,
                    actual_headcount, work_content, pre_entry_requirement, work_headcount, entry_order,
                    created_at, updated_at
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, CURRENT_TIMESTAMP, CURRENT_TIMESTAMP)
                """,
                (
                    int(vendor_work_entry_context["sheet_id"]),
                    str(vendor_work_entry_context["vendor_name"]),
                    business_date,
                    planned_at,
                    planned_headcount,
                    actual_headcount,
                    work_content,
                    pre_entry_requirement,
                    work_headcount,
                    entry_order,
                ),
            )
            target_id = int(cur.lastrowid)
        else:
            conn.execute(
                """
                UPDATE vendor_work_entries
                SET vendor_name = ?,
                    business_date = ?,
                    planned_at = ?,
                    planned_headcount = ?,
                    actual_headcount = ?,
                    work_content = ?,
                    pre_entry_requirement = ?,
                    work_headcount = ?,
                    entry_order = ?,
                    updated_at = CURRENT_TIMESTAMP
                WHERE id = ? AND sheet_id = ?
                """,
                (
                    str(vendor_work_entry_context["vendor_name"]),
                    business_date,
                    planned_at,
                    planned_headcount,
                    actual_headcount,
                    work_content,
                    pre_entry_requirement,
                    work_headcount,
                    entry_order,
                    entry_id,
                    int(vendor_work_entry_context["sheet_id"]),
                ),
            )
            target_id = entry_id

        row = conn.execute(
            """
            SELECT id, sheet_id, vendor_name, business_date, planned_at, planned_headcount,
                   actual_headcount, work_content, pre_entry_requirement, work_headcount, entry_order, created_at, updated_at
            FROM vendor_work_entries
            WHERE id = ?
            """,
            (target_id,),
        ).fetchone()

    return jsonify({"ok": True, "entry": dict(row) if row else None})


@app.route("/api/crew-work-entry-requirement-confirm", methods=["POST"])
def api_crew_work_entry_requirement_confirm():
    data = request.get_json(silent=True) or {}
    try:
        entry_id = parse_non_negative_int(data.get("entry_id"), field_name="entry_id")
        sheet_id = parse_non_negative_int(data.get("sheet_id"), field_name="sheet_id")
        if entry_id <= 0:
            raise ValueError("entry_id must be a positive integer.")
        if sheet_id <= 0:
            raise ValueError("sheet_id must be a positive integer.")
    except ValueError as exc:
        return crew_api_error("invalid_request", str(exc))

    with db() as conn:
        try:
            confirmation_context = authorize_vendor_work_entry_requirement_confirmation(
                conn,
                sheet_id=sheet_id,
                entry_id=entry_id,
            )
        except LookupError as exc:
            return _handle_vendor_work_entry_requirement_confirmation_lookup_error(exc)

        row = conn.execute(
            """
            SELECT username
            FROM users
            WHERE id = ?
            """,
            (int(_current_internal_user()["id"]),),
        ).fetchone()
        confirmed_by = str(row["username"] if row is not None else "")

        if confirmation_context["requirement_status"] != "confirmed":
            conn.execute(
                """
                UPDATE vendor_work_entries
                SET requirement_status = 'confirmed',
                    requirement_confirmed_by = ?,
                    requirement_confirmed_at = CURRENT_TIMESTAMP,
                    updated_at = CURRENT_TIMESTAMP
                WHERE id = ? AND sheet_id = ?
                """,
                (
                    confirmed_by,
                    int(confirmation_context["entry_id"]),
                    int(confirmation_context["sheet_id"]),
                ),
            )

        confirmed_entry = conn.execute(
            """
            SELECT id, requirement_status, requirement_confirmed_by, requirement_confirmed_at
            FROM vendor_work_entries
            WHERE id = ?
            """,
            (int(confirmation_context["entry_id"]),),
        ).fetchone()

    return jsonify({"ok": True, "entry": dict(confirmed_entry) if confirmed_entry else None})


@app.route("/api/crew-work-entry/formal-approve", methods=["POST"])
def api_crew_work_entry_formal_approve():
    data = request.get_json(silent=True) or {}
    try:
        entry_id = parse_non_negative_int(data.get("entry_id"), field_name="entry_id")
        sheet_id = parse_non_negative_int(data.get("sheet_id"), field_name="sheet_id")
        action = str(data.get("action") or "").strip()
        if entry_id <= 0:
            raise ValueError("entry_id must be a positive integer.")
        if sheet_id <= 0:
            raise ValueError("sheet_id must be a positive integer.")
        if action != "crew_formal_approve_entry":
            raise ValueError("action must be crew_formal_approve_entry.")
    except ValueError as exc:
        return crew_api_error("invalid_request", str(exc))

    with db() as conn:
        try:
            approval_context = authorize_vendor_work_entry_formal_approve(
                conn,
                sheet_id=sheet_id,
                entry_id=entry_id,
            )
        except LookupError as exc:
            return _handle_vendor_work_entry_formal_approve_lookup_error(exc)

        if str(approval_context["scheduling_gate_state"]) != "allowed":
            return crew_api_error("entry_not_ready", "Entry is not ready for this action.", status=409)

        existing_approval = conn.execute(
            """
            SELECT id
            FROM formal_approvals
            WHERE entry_id = ? AND action = ?
            """,
            (int(approval_context["entry_id"]), action),
        ).fetchone()
        if existing_approval is not None:
            return crew_api_error(
                "duplicate_approval",
                "Formal approval already exists for this entry.",
                status=409,
            )

        user = _current_internal_user()
        approved_by = str(user["username"] if user is not None else "")
        try:
            conn.execute(
                """
                INSERT INTO formal_approvals (
                    entry_id,
                    sheet_id,
                    action,
                    approval_status,
                    approved_by,
                    approved_at,
                    updated_at
                ) VALUES (?, ?, ?, 'approved', ?, CURRENT_TIMESTAMP, CURRENT_TIMESTAMP)
                """,
                (
                    int(approval_context["entry_id"]),
                    int(approval_context["sheet_id"]),
                    action,
                    approved_by,
                ),
            )
        except sqlite3.IntegrityError:
            return crew_api_error(
                "duplicate_approval",
                "Formal approval already exists for this entry.",
                status=409,
            )

    return jsonify(
        {
            "ok": True,
            "action": "crew_formal_approve_entry",
            "entry": {
                "id": int(approval_context["entry_id"]),
                "sheet_id": int(approval_context["sheet_id"]),
            },
        }
    )


@app.route("/api/schedule-entry", methods=["POST"])
def api_schedule_entry():
    data = request.get_json(silent=True) or {}
    try:
        entry_id = parse_non_negative_int(data.get("entry_id"), field_name="entry_id")
        sheet_id = parse_non_negative_int(data.get("sheet_id"), field_name="sheet_id")
        action = str(data.get("action") or "").strip()
        scheduled_date = parse_schedule_date(data.get("scheduled_date"))
        scheduled_time = parse_schedule_time(data.get("scheduled_time"))
        if entry_id <= 0:
            raise ValueError("entry_id must be a positive integer.")
        if sheet_id <= 0:
            raise ValueError("sheet_id must be a positive integer.")
        if action != "schedule_entry":
            raise ValueError("action must be schedule_entry.")
    except ValueError as exc:
        return crew_api_error("invalid_request", str(exc))

    with db() as conn:
        try:
            schedule_context = authorize_schedule_entry_write(
                conn,
                sheet_id=sheet_id,
                entry_id=entry_id,
            )
        except LookupError as exc:
            return _handle_schedule_entry_lookup_error(exc)

        if not (
            str(schedule_context.get("formal_approval_state")) == "approved"
            and str(schedule_context.get("scheduling_gate_state")) == "allowed"
        ):
            return crew_api_error("entry_not_schedulable", "Entry is not schedulable for this action.", status=409)

        existing_schedule = conn.execute(
            """
            SELECT id
            FROM scheduling_entries
            WHERE entry_id = ? AND action = ?
            """,
            (int(schedule_context["entry_id"]), action),
        ).fetchone()
        if existing_schedule is not None:
            return crew_api_error(
                "duplicate_schedule",
                "Schedule already exists for this entry.",
                status=409,
            )

        user = _current_internal_user()
        scheduled_by = str(user["username"] if user is not None else "")
        try:
            row = conn.execute(
                """
                INSERT INTO scheduling_entries (
                    entry_id,
                    sheet_id,
                    action,
                    schedule_status,
                    scheduled_date,
                    scheduled_time,
                    scheduled_by,
                    scheduled_at,
                    updated_at
                ) VALUES (?, ?, ?, 'scheduled', ?, ?, ?, CURRENT_TIMESTAMP, CURRENT_TIMESTAMP)
                RETURNING id, entry_id, sheet_id, scheduled_date, scheduled_time
                """,
                (
                    int(schedule_context["entry_id"]),
                    int(schedule_context["sheet_id"]),
                    action,
                    scheduled_date,
                    scheduled_time,
                    scheduled_by,
                ),
            ).fetchone()
        except sqlite3.IntegrityError:
            return crew_api_error(
                "duplicate_schedule",
                "Schedule already exists for this entry.",
                status=409,
            )

    return jsonify(
        {
            "ok": True,
            "action": "schedule_entry",
            "schedule": {
                "id": int(row["id"]),
                "entry_id": int(row["entry_id"]),
                "sheet_id": int(row["sheet_id"]),
                "scheduled_date": str(row["scheduled_date"]),
                "scheduled_time": str(row["scheduled_time"]),
            },
        }
    )


@app.route("/api/vendor/work-entry/preflight", methods=["POST"])
def api_vendor_work_entry_preflight():
    data = request.get_json(silent=True) or {}
    try:
        entry_id = parse_optional_positive_int(data.get("id"), field_name="id")
        sheet_id = parse_non_negative_int(data.get("sheet_id"), field_name="sheet_id")
        if sheet_id <= 0:
            raise ValueError("sheet_id must be a positive integer.")
        raw_vendor_name = str(data.get("vendor_name", "")).strip()
        payload_vendor_name = normalize_vendor_name(raw_vendor_name) if raw_vendor_name else None
        raw_business_date = str(data.get("business_date", "")).strip()
        business_date = (
            parse_crew_business_date(raw_business_date)
            if raw_business_date
            else resolve_crew_business_date()
        )
    except ValueError as exc:
        return crew_api_error("invalid_request", str(exc))

    with db() as conn:
        try:
            preflight_payload = build_vendor_work_entry_preflight_payload(
                conn,
                sheet_id=sheet_id,
                business_date=business_date,
                entry_id=entry_id,
                payload_vendor_name=payload_vendor_name,
                payload_business_date_provided=bool(raw_business_date),
            )
        except LookupError as exc:
            return _handle_vendor_write_preflight_error(exc)

    return jsonify(preflight_payload)


@app.route("/api/crew-followups")
@login_required
def api_crew_followups():
    sheet_id = request.args.get("sheet_id", type=int)
    if sheet_id is None:
        return crew_api_error("invalid_sheet_id", "sheet_id is required and must be a valid integer.")

    raw_business_date = request.args.get("business_date", "").strip()
    try:
        business_date = parse_crew_business_date(raw_business_date) if raw_business_date else resolve_crew_business_date()
    except ValueError as exc:
        return crew_api_error("invalid_business_date", str(exc))

    with db() as conn:
        try:
            authorize_sheet_read(conn, sheet_id)
        except LookupError as exc:
            return _handle_grid_read_lookup_error(exc)

        active_vendors = get_active_crew_vendors(sheet_id)
        pending_by_vendor = get_pending_items_by_vendor(sheet_id)
        contacts_by_vendor = fetch_vendor_contact_map(conn, sheet_id)
        entries_by_vendor = fetch_vendor_work_entries(conn, sheet_id=sheet_id, business_date=business_date)

        items = []
        for vendor_name in active_vendors:
            entries = entries_by_vendor.get(vendor_name, [])
            has_valid_planned_at = any(bool(str(entry.get("planned_at", "")).strip()) for entry in entries)
            if has_valid_planned_at:
                continue
            contact = contacts_by_vendor.get(vendor_name, {})
            items.append(
                {
                    "vendor_name": vendor_name,
                    "pending_items": pending_by_vendor.get(vendor_name, []),
                    "contact_name": str(contact.get("contact_name", "")),
                    "contact_phone": str(contact.get("contact_phone", "")),
                    "planned_at": "",
                }
            )

    return jsonify({"ok": True, "sheet_id": sheet_id, "business_date": business_date, "items": items})


@app.route("/api/crew-daily-summary")
@login_required
def api_crew_daily_summary():
    sheet_id = request.args.get("sheet_id", type=int)
    if sheet_id is None:
        return crew_api_error("invalid_sheet_id", "sheet_id is required and must be a valid integer.")

    raw_business_date = request.args.get("business_date", "").strip()
    try:
        business_date = parse_crew_business_date(raw_business_date) if raw_business_date else resolve_crew_business_date()
    except ValueError as exc:
        return crew_api_error("invalid_business_date", str(exc))

    with db() as conn:
        try:
            authorize_sheet_read(conn, sheet_id)
        except LookupError as exc:
            return _handle_grid_read_lookup_error(exc)

        rows = conn.execute(
            """
            SELECT id, vendor_name, actual_headcount, work_content, work_headcount,
                   planned_at, planned_headcount, entry_order
            FROM vendor_work_entries
            WHERE sheet_id = ? AND business_date = ? AND actual_headcount > 0
            ORDER BY vendor_name, entry_order, id
            """,
            (sheet_id, business_date),
        ).fetchall()
        items = [
            {
                "id": row["id"],
                "vendor_name": row["vendor_name"],
                "actual_headcount": row["actual_headcount"],
                "work_content": row["work_content"],
                "work_headcount": row["work_headcount"],
                "planned_at": row["planned_at"],
                "planned_headcount": row["planned_headcount"],
                "entry_order": row["entry_order"],
            }
            for row in rows
        ]
        totals = {
            "vendors": 0,
            "actual_headcount_sum": 0,
            "work_headcount_sum": 0,
        }
        if items:
            totals = {
                "vendors": len({item["vendor_name"] for item in items}),
                "actual_headcount_sum": sum(int(item["actual_headcount"]) for item in items),
                "work_headcount_sum": sum(int(item["work_headcount"]) for item in items),
            }

    return jsonify(
        {
            "ok": True,
            "sheet_id": sheet_id,
            "business_date": business_date,
            "items": items,
            "totals": totals,
        }
    )


@app.route("/api/crew-missing")
@login_required
def api_crew_missing():
    sheet_id = request.args.get("sheet_id", type=int)
    if sheet_id is None:
        return crew_api_error("invalid_sheet_id", "sheet_id is required and must be a valid integer.")

    raw_business_date = request.args.get("business_date", "").strip()
    try:
        business_date = parse_crew_business_date(raw_business_date) if raw_business_date else resolve_crew_business_date()
    except ValueError as exc:
        return crew_api_error("invalid_business_date", str(exc))

    with db() as conn:
        try:
            authorize_sheet_read(conn, sheet_id)
        except LookupError as exc:
            return _handle_grid_read_lookup_error(exc)

        active_vendors = set(get_active_crew_vendors(sheet_id))
        pending_by_vendor = get_pending_items_by_vendor(sheet_id)
        contacts_by_vendor = fetch_vendor_contact_map(conn, sheet_id)
        entries_by_vendor = fetch_vendor_work_entries(conn, sheet_id=sheet_id, business_date=business_date)

        now = datetime.now()
        items = []
        for vendor_name, entries in entries_by_vendor.items():
            if vendor_name not in active_vendors:
                continue
            contact = contacts_by_vendor.get(vendor_name, {})
            for entry in entries:
                planned_at = str(entry.get("planned_at", "")).strip()
                if not planned_at:
                    continue
                planned_at_dt = parse_planned_at_datetime(planned_at)
                if planned_at_dt is None or planned_at_dt > now:
                    continue
                actual_headcount = int(entry.get("actual_headcount", 0) or 0)
                if actual_headcount > 0:
                    continue
                items.append(
                    {
                        "vendor_name": vendor_name,
                        "contact_name": str(contact.get("contact_name", "")),
                        "contact_phone": str(contact.get("contact_phone", "")),
                        "planned_at": planned_at,
                        "planned_headcount": int(entry.get("planned_headcount", 0) or 0),
                        "actual_headcount": actual_headcount,
                        "pending_items": pending_by_vendor.get(vendor_name, []),
                    }
                )

    return jsonify({"ok": True, "sheet_id": sheet_id, "business_date": business_date, "items": items})


@app.route("/admin/users", methods=["GET", "POST"])
@admin_required
def users():
    if request.method == "POST":
        action = request.form.get("action", "create_user")
        username = request.form.get("username", "").strip()
        display_name = request.form.get("display_name", "").strip() or username
        password = request.form.get("password", "")
        role = request.form.get("role", "member")

        if action.startswith("add_site_permission:"):
            try:
                user_id = int(action.split(":", 1)[1])
            except (TypeError, ValueError):
                flash("\u64cd\u4f5c\u7121\u6548\u3002", "error")
                return redirect(url_for("users"))
            site_role = request.form.get("site_role", "")
            try:
                site_id = int(request.form.get("site_id", ""))
            except (TypeError, ValueError):
                flash("\u5de5\u5730\u9078\u64c7\u7121\u6548\u3002", "error")
                return redirect(url_for("users"))
            try:
                with db() as conn:
                    target_user = get_user_by_id(user_id)
                    if target_user is None:
                        raise LookupError("user_not_found")
                    if is_global_admin(target_user):
                        raise ValueError("admin_permissions_forbidden")
                    site_row = conn.execute(
                        "SELECT id, site_name, is_active FROM sites WHERE id = ?",
                        (site_id,),
                    ).fetchone()
                    if site_row is None or int(site_row["is_active"]) != 1:
                        raise LookupError("site_not_found")
                    create_user_site_permission_sqlite(
                        conn,
                        user_id=user_id,
                        site_id=site_id,
                        role=site_role,
                    )
                flash("\u5de5\u5730\u6b0a\u9650\u5df2\u65b0\u589e\u3002", "success")
            except ValueError as exc:
                if str(exc) == "admin_permissions_forbidden":
                    flash("\u7ba1\u7406\u54e1\u4e0d\u9700\u8981\u5de5\u5730\u6388\u6b0a\u3002", "error")
                else:
                    flash("\u5de5\u5730\u89d2\u8272\u8a2d\u5b9a\u932f\u8aa4\u3002", "error")
            except LookupError as exc:
                if str(exc) == "user_not_found":
                    flash("\u627e\u4e0d\u5230\u6210\u54e1\u3002", "error")
                else:
                    flash("\u627e\u4e0d\u5230\u5de5\u5730\u6216\u5de5\u5730\u5df2\u505c\u7528\u3002", "error")
            except sqlite3.IntegrityError:
                flash("\u8a72\u6210\u54e1\u5df2\u64c1\u6709\u6b64\u5de5\u5730\u6388\u6b0a\u3002", "error")
        elif action.startswith("update_site_permission:"):
            try:
                permission_id = int(action.split(":", 1)[1])
            except (TypeError, ValueError):
                flash("\u64cd\u4f5c\u7121\u6548\u3002", "error")
                return redirect(url_for("users"))
            site_role = request.form.get("site_role", "")
            try:
                with db() as conn:
                    permission_row = _fetch_user_site_permission_row(conn, permission_id)
                    if permission_row is None:
                        raise LookupError("permission_not_found")
                    target_user = get_user_by_id(int(permission_row["user_id"]))
                    if target_user is None:
                        raise LookupError("user_not_found")
                    if is_global_admin(target_user):
                        raise ValueError("admin_permissions_forbidden")
                    update_user_site_permission_role_sqlite(conn, permission_id, role=site_role)
                flash("\u5de5\u5730\u89d2\u8272\u5df2\u66f4\u65b0\u3002", "success")
            except ValueError:
                flash("\u5de5\u5730\u89d2\u8272\u8a2d\u5b9a\u932f\u8aa4\u3002", "error")
            except LookupError as exc:
                if str(exc) == "permission_not_found":
                    flash("\u627e\u4e0d\u5230\u5de5\u5730\u6388\u6b0a\u3002", "error")
                else:
                    flash("\u627e\u4e0d\u5230\u6210\u54e1\u3002", "error")
        elif action.startswith("delete_site_permission:"):
            try:
                permission_id = int(action.split(":", 1)[1])
            except (TypeError, ValueError):
                flash("\u64cd\u4f5c\u7121\u6548\u3002", "error")
                return redirect(url_for("users"))
            try:
                with db() as conn:
                    permission_row = _fetch_user_site_permission_row(conn, permission_id)
                    if permission_row is None:
                        raise LookupError("permission_not_found")
                    target_user = get_user_by_id(int(permission_row["user_id"]))
                    if target_user is None:
                        raise LookupError("user_not_found")
                    if is_global_admin(target_user):
                        raise ValueError("admin_permissions_forbidden")
                    delete_user_site_permission_sqlite(conn, permission_id)
                flash("\u5de5\u5730\u6388\u6b0a\u5df2\u522a\u9664\u3002", "success")
            except ValueError:
                flash("\u7ba1\u7406\u54e1\u4e0d\u9700\u8981\u5de5\u5730\u6388\u6b0a\u3002", "error")
            except LookupError as exc:
                if str(exc) == "permission_not_found":
                    flash("\u627e\u4e0d\u5230\u5de5\u5730\u6388\u6b0a\u3002", "error")
                else:
                    flash("\u627e\u4e0d\u5230\u6210\u54e1\u3002", "error")
        elif action == "create_user":
            if role not in ("member", "admin"):
                flash("\u89d2\u8272\u8a2d\u5b9a\u932f\u8aa4\u3002", "error")
            elif not username or not password:
                flash("\u8acb\u8f38\u5165\u5e33\u865f\u8207\u5bc6\u78bc\u3002", "error")
            else:
                try:
                    password_hash = generate_password_hash(password)
                    with db() as conn:
                        user_row = create_user_sqlite(
                            conn,
                            username=username,
                            display_name=display_name,
                            password_hash=password_hash,
                            role=role,
                        )
                    maybe_dual_write_user_create(user_row)
                    flash("\u6210\u54e1\u5df2\u65b0\u589e\u3002", "success")
                except sqlite3.IntegrityError:
                    flash("\u5e33\u865f\u5df2\u5b58\u5728\u3002", "error")
        elif action.startswith("update_user:"):
            if role not in ("member", "admin"):
                flash("\u89d2\u8272\u8a2d\u5b9a\u932f\u8aa4\u3002", "error")
                return redirect(url_for("users"))
            user_id = int(action.split(":", 1)[1])
            if user_id == session.get("user_id"):
                flash("\u672c\u968e\u6bb5\u4e0d\u5141\u8a31\u4fee\u6539\u81ea\u5df1\u7684\u89d2\u8272", "error")
            else:
                try:
                    with db() as conn:
                        update_user_role_sqlite(conn, user_id, role=role)
                    maybe_dual_write_user_role_update(user_id, role=role)
                    if user_id == session.get("user_id"):
                        session["username"] = username
                        session["display_name"] = display_name
                        session["role"] = role
                    flash("\u6210\u54e1\u8cc7\u6599\u5df2\u66f4\u65b0\u3002", "success")
                except sqlite3.IntegrityError:
                    flash("\u5e33\u865f\u5df2\u5b58\u5728\u3002", "error")
        elif action.startswith("delete_user:"):
            user_id = int(action.split(":", 1)[1])
            try:
                with db() as conn:
                    deleted_user_row = delete_user_sqlite(conn, user_id)
                maybe_dual_write_user_delete(deleted_user_row)
                flash("\u6210\u54e1\u5df2\u522a\u9664\u3002", "success")
            except ValueError:
                flash("\u7ba1\u7406\u54e1\u6216\u4fdd\u8b77\u5e33\u865f\u4e0d\u53ef\u522a\u9664\u3002", "error")
            except LookupError:
                flash("\u627e\u4e0d\u5230\u8981\u522a\u9664\u7684\u6210\u54e1\u3002", "error")
        else:
            flash("\u64cd\u4f5c\u7121\u6548\u3002", "error")

        return redirect(url_for("users"))

    all_users = list_users()
    with db() as conn:
        settings = get_settings(conn)
        active_sites = _fetch_active_sites(conn)
        permission_map = get_user_site_permissions_map(conn, user_ids=[int(user["id"]) for user in all_users])

    users_payload = []
    for user in all_users:
        is_admin = is_global_admin(user)
        users_payload.append(
            {
                "id": int(user["id"]),
                "username": str(user["username"]),
                "display_name": str(user["display_name"] or user["username"]),
                "role": str(user["role"]),
                "created_at": str(user["created_at"]),
                "is_global_admin": is_admin,
                "site_permissions": [] if is_admin else permission_map.get(int(user["id"]), []),
            }
        )

    return render_template(
        "users.html",
        users=users_payload,
        settings=settings,
        active_sites=active_sites,
        site_permission_roles=get_site_permission_role_options(),
    )


@app.route("/admin/table", methods=["GET", "POST"])
@admin_required
def table_admin():
    with db() as conn:
        sheet_id = resolve_sheet_id(conn, request.values.get("sheet_id", type=int) or session.get("sheet_id"))
        session["sheet_id"] = sheet_id
        if request.method == "POST":
            actions = request.form.getlist("action")
            action = actions[-1] if actions else "save"
            if action == "create_sheet":
                name = request.form.get("new_sheet_name", "").strip() or "????"
                next_order = conn.execute("SELECT COALESCE(MAX(sort_order), 0) + 1 FROM sheets").fetchone()[0]
                try:
                    create_sheet_context = authorize_admin_create_sheet_site(conn)
                except LookupError as exc:
                    return _handle_admin_site_write_lookup_error(exc)
                cur = conn.execute(
                    "INSERT INTO sheets (name, sort_order, site_id) VALUES (?, ?, ?)",
                    (name, next_order, create_sheet_context["site_id"]),
                )
                for field_key, field in BUILTIN_EXTRA_FIELDS.items():
                    conn.execute(
                        """
                        INSERT INTO extra_fields
                        (sheet_id, field_key, name, field_type, sort_order, is_builtin, active)
                        VALUES (?, ?, ?, ?, ?, 1, 1)
                        """,
                        (cur.lastrowid, field_key, field["name"], field["type"], field["sort_order"]),
                    )
                flash("???????", "success")
                return redirect(url_for("table_admin", sheet_id=cur.lastrowid))
            if action == "delete_sheet":
                count = conn.execute("SELECT COUNT(*) FROM sheets").fetchone()[0]
                if count <= 1:
                    flash("????????????", "error")
                    return redirect(url_for("table_admin", sheet_id=sheet_id))
                try:
                    delete_sheet_context = authorize_admin_site_scoped_write(conn, sheet_id=sheet_id)
                except LookupError as exc:
                    return _handle_admin_site_write_lookup_error(exc, sheet_id=sheet_id)
                unit_ids = [row["id"] for row in conn.execute("SELECT u.id FROM units u JOIN floors f ON f.id = u.floor_id WHERE f.sheet_id = ?", (sheet_id,))]
                if unit_ids:
                    placeholders = ",".join("?" for _ in unit_ids)
                    conn.execute(f"DELETE FROM progress WHERE unit_id IN ({placeholders})", unit_ids)
                    conn.execute(f"DELETE FROM unit_extra WHERE unit_id IN ({placeholders})", unit_ids)
                    conn.execute(f"DELETE FROM unit_extra_values WHERE unit_id IN ({placeholders})", unit_ids)
                    conn.execute(f"DELETE FROM units WHERE id IN ({placeholders})", unit_ids)
                conn.execute("DELETE FROM tasks WHERE sheet_id = ?", (sheet_id,))
                conn.execute("DELETE FROM floors WHERE sheet_id = ?", (sheet_id,))
                conn.execute("DELETE FROM unit_extra_values WHERE field_key IN (SELECT field_key FROM extra_fields WHERE sheet_id = ?)", (sheet_id,))
                conn.execute("DELETE FROM extra_fields WHERE sheet_id = ?", (sheet_id,))
                conn.execute("DELETE FROM sheets WHERE id = ?", (sheet_id,))
                next_sheet = resolve_sheet_id(conn)
                flash("???????", "success")
                return redirect(url_for("table_admin", sheet_id=next_sheet))
            if action == "add_task":
                try:
                    authorize_admin_site_scoped_write(conn, sheet_id=sheet_id)
                except LookupError as exc:
                    return _handle_admin_site_write_lookup_error(exc, sheet_id=sheet_id)
                next_col = conn.execute("SELECT COALESCE(MAX(col_index), 3) + 1 FROM tasks").fetchone()[0]
                cur = conn.execute(
                    "INSERT INTO tasks (sheet_id, col_index, vendor, location, name) VALUES (?, ?, ?, ?, ?)",
                    (sheet_id, next_col, request.form.get("new_task_vendor", ""), request.form.get("new_task_location", ""), request.form.get("new_task_name", "???") or "???"),
                )
                for unit in conn.execute("SELECT u.id FROM units u JOIN floors f ON f.id = u.floor_id WHERE f.sheet_id = ?", (sheet_id,)):
                    conn.execute("INSERT INTO progress (unit_id, task_id, value) VALUES (?, ?, ?)", (unit["id"], cur.lastrowid, WORKING_VALUE))
                flash("??????", "success")
                return redirect(url_for("table_admin", sheet_id=sheet_id))
            if action.startswith("delete_task:"):
                task_id = int(action.split(":", 1)[1])
                try:
                    task_row = resolve_task_sheet_for_admin_write(conn, task_id=task_id)
                    if int(task_row["sheet_id"]) != int(sheet_id):
                        raise LookupError("task_sheet_mismatch")
                    authorize_admin_site_scoped_write(conn, sheet_id=sheet_id)
                except LookupError as exc:
                    return _handle_admin_site_write_lookup_error(exc, sheet_id=sheet_id)
                conn.execute("DELETE FROM progress WHERE task_id = ?", (task_id,))
                conn.execute("DELETE FROM tasks WHERE id = ? AND sheet_id = ?", (task_id, sheet_id))
                flash("??????", "success")
                return redirect(url_for("table_admin", sheet_id=sheet_id))
            if action == "add_extra_field":
                field_name = request.form.get("new_extra_name", "").strip() or "新增欄位"
                field_type = request.form.get("new_extra_type", "date")
                if field_type not in EXTRA_FIELD_TYPES:
                    field_type = "date"
                try:
                    authorize_admin_site_scoped_write(conn, sheet_id=sheet_id)
                except LookupError as exc:
                    return _handle_admin_site_write_lookup_error(exc, sheet_id=sheet_id)
                next_order = conn.execute(
                    "SELECT COALESCE(MAX(sort_order), 0) + 1 FROM extra_fields WHERE sheet_id = ?",
                    (sheet_id,),
                ).fetchone()[0]
                conn.execute(
                    """
                    INSERT INTO extra_fields
                    (sheet_id, field_key, name, field_type, sort_order, is_builtin, active)
                    VALUES (?, ?, ?, ?, ?, 0, 1)
                    """,
                    (sheet_id, f"custom_{uuid.uuid4().hex[:12]}", field_name, field_type, next_order),
                )
                flash("欄位已新增。", "success")
                return redirect(url_for("table_admin", sheet_id=sheet_id))
            if action.startswith("delete_extra_field:"):
                field_id = int(action.split(":", 1)[1])
                try:
                    field_row = resolve_extra_field_sheet_for_admin_write(conn, field_id=field_id)
                    if int(field_row["sheet_id"]) != int(sheet_id):
                        raise LookupError("extra_field_sheet_mismatch")
                    authorize_admin_site_scoped_write(conn, sheet_id=int(field_row["sheet_id"]))
                except LookupError as exc:
                    return _handle_admin_site_write_lookup_error(exc, sheet_id=sheet_id)
                conn.execute("UPDATE extra_fields SET active = 0 WHERE id = ? AND sheet_id = ?", (field_id, sheet_id))
                flash("欄位已刪除。", "success")
                return redirect(url_for("table_admin", sheet_id=sheet_id))
            if action == "add_floor":
                try:
                    authorize_admin_site_scoped_write(conn, sheet_id=sheet_id)
                except LookupError as exc:
                    return _handle_admin_site_write_lookup_error(exc, sheet_id=sheet_id)
                next_order = conn.execute("SELECT COALESCE(MAX(sort_order), 0) + 1 FROM floors").fetchone()[0]
                conn.execute(
                    "INSERT INTO floors (sheet_id, sort_order, name, block_name, unit_count) VALUES (?, ?, ?, ?, 0)",
                    (sheet_id, next_order, request.form.get("new_floor_name", "???") or "???", request.form.get("new_floor_block", "")),
                )
                flash("??????", "success")
                return redirect(url_for("table_admin", sheet_id=sheet_id))
            if action.startswith("delete_floor:"):
                floor_id = int(action.split(":", 1)[1])
                try:
                    floor_row = resolve_floor_sheet_for_admin_write(conn, floor_id=floor_id)
                    if int(floor_row["sheet_id"]) != int(sheet_id):
                        raise LookupError("floor_sheet_mismatch")
                    authorize_admin_site_scoped_write(conn, sheet_id=int(floor_row["sheet_id"]))
                except LookupError as exc:
                    return _handle_admin_site_write_lookup_error(exc, sheet_id=sheet_id)
                unit_ids = [row["id"] for row in conn.execute("SELECT id FROM units WHERE floor_id = ?", (floor_id,))]
                if unit_ids:
                    placeholders = ",".join("?" for _ in unit_ids)
                    conn.execute(f"DELETE FROM progress WHERE unit_id IN ({placeholders})", unit_ids)
                    conn.execute(f"DELETE FROM unit_extra WHERE unit_id IN ({placeholders})", unit_ids)
                    conn.execute(f"DELETE FROM unit_extra_values WHERE unit_id IN ({placeholders})", unit_ids)
                    conn.execute(f"DELETE FROM units WHERE id IN ({placeholders})", unit_ids)
                conn.execute("DELETE FROM floors WHERE id = ? AND sheet_id = ?", (floor_id, sheet_id))
                flash("??????", "success")
                return redirect(url_for("table_admin", sheet_id=sheet_id))
            if action.startswith("add_unit:"):
                floor_id = int(action.split(":", 1)[1])
                try:
                    floor_row = resolve_floor_sheet_for_admin_write(conn, floor_id=floor_id)
                    if int(floor_row["sheet_id"]) != int(sheet_id):
                        raise LookupError("floor_sheet_mismatch")
                    authorize_admin_site_scoped_write(conn, sheet_id=int(floor_row["sheet_id"]))
                except LookupError as exc:
                    return _handle_admin_site_write_lookup_error(exc, sheet_id=sheet_id)
                next_order = conn.execute("SELECT COALESCE(MAX(sort_order), 0) + 1 FROM units WHERE floor_id = ?", (floor_id,)).fetchone()[0]
                cur = conn.execute(
                    "INSERT INTO units (floor_id, sort_order, name) VALUES (?, ?, ?)",
                    (floor_id, next_order, request.form.get(f"new_unit_name_{floor_id}", "???") or "???"),
                )
                for task in conn.execute("SELECT id FROM tasks WHERE sheet_id = ?", (sheet_id,)):
                    conn.execute("INSERT INTO progress (unit_id, task_id, value) VALUES (?, ?, ?)", (cur.lastrowid, task["id"], WORKING_VALUE))
                conn.execute("INSERT INTO unit_extra (unit_id, handover) VALUES (?, ?)", (cur.lastrowid, WORKING_VALUE))
                conn.execute("UPDATE floors SET unit_count = (SELECT COUNT(*) FROM units WHERE floor_id = ?) WHERE id = ?", (floor_id, floor_id))
                flash("??????", "success")
                return redirect(url_for("table_admin", sheet_id=sheet_id))
            if action.startswith("delete_unit:"):
                unit_id = int(action.split(":", 1)[1])
                try:
                    unit_row = resolve_unit_sheet_for_admin_write(conn, unit_id=unit_id)
                    if int(unit_row["sheet_id"]) != int(sheet_id):
                        raise LookupError("unit_sheet_mismatch")
                    authorize_admin_site_scoped_write(conn, sheet_id=int(unit_row["sheet_id"]))
                except LookupError as exc:
                    return _handle_admin_site_write_lookup_error(exc, sheet_id=sheet_id)
                conn.execute("DELETE FROM progress WHERE unit_id = ?", (unit_id,))
                conn.execute("DELETE FROM unit_extra WHERE unit_id = ?", (unit_id,))
                conn.execute("DELETE FROM unit_extra_values WHERE unit_id = ?", (unit_id,))
                conn.execute("DELETE FROM units WHERE id = ?", (unit_id,))
                conn.execute(
                    "UPDATE floors SET unit_count = (SELECT COUNT(*) FROM units WHERE floor_id = ?) WHERE id = ?",
                    (unit_row["floor_id"], unit_row["floor_id"]),
                )
                flash("??????", "success")
                return redirect(url_for("table_admin", sheet_id=sheet_id))

            try:
                authorize_admin_site_scoped_write(conn, sheet_id=sheet_id)
            except LookupError as exc:
                return _handle_admin_site_write_lookup_error(exc, sheet_id=sheet_id)
            save_admin_global_settings(conn, form=request.form)
            save_admin_site_content(conn, sheet_id=sheet_id, form=request.form)
            flash("????????", "success")
            return redirect(url_for("table_admin", sheet_id=sheet_id))

        settings = get_settings(conn)
        sheets = available_sheets(conn)
        current_sheet = conn.execute("SELECT * FROM sheets WHERE id = ?", (sheet_id,)).fetchone()
        tasks = conn.execute("SELECT * FROM tasks WHERE sheet_id = ? ORDER BY col_index", (sheet_id,)).fetchall()
        extra_fields = conn.execute("SELECT * FROM extra_fields WHERE sheet_id = ? AND active = 1 ORDER BY sort_order, id", (sheet_id,)).fetchall()
        floors = conn.execute("SELECT * FROM floors WHERE sheet_id = ? ORDER BY sort_order", (sheet_id,)).fetchall()
        units = {floor["id"]: conn.execute("SELECT * FROM units WHERE floor_id = ? ORDER BY sort_order", (floor["id"],)).fetchall() for floor in floors}
        unit_write_state = build_admin_site_write_render_state(
            conn,
            sheet_id=sheet_id,
            missing_current_site_message="請先選擇目前工地，才能管理戶別。",
            sheet_not_in_current_site_message="目前工地與此表單不一致，不能修改戶別。",
        )
        floor_write_state = build_admin_site_write_render_state(
            conn,
            sheet_id=sheet_id,
            missing_current_site_message="請先選擇目前工地，才能管理樓層。",
            sheet_not_in_current_site_message="目前工地與此表單不一致，不能修改樓層。",
        )
        task_write_state = build_admin_site_write_render_state(
            conn,
            sheet_id=sheet_id,
            missing_current_site_message="請先選擇目前工地，才能管理工項。",
            sheet_not_in_current_site_message="目前工地與此表單不一致，不能修改工項。",
        )
        extra_field_write_state = build_admin_site_write_render_state(
            conn,
            sheet_id=sheet_id,
            missing_current_site_message="請先選擇目前工地，才能管理驗收／交屋欄位。",
            sheet_not_in_current_site_message="目前工地與此表單不一致，不能修改驗收／交屋欄位。",
        )
    return render_template(
        "table_admin.html",
        settings=settings,
        sheets=sheets,
        current_sheet=current_sheet,
        tasks=tasks,
        unit_write_enabled=unit_write_state["enabled"],
        unit_write_block_reason=unit_write_state["block_reason"],
        unit_write_block_message=unit_write_state["block_message"],
        floor_write_enabled=floor_write_state["enabled"],
        floor_write_block_reason=floor_write_state["block_reason"],
        floor_write_block_message=floor_write_state["block_message"],
        task_write_enabled=task_write_state["enabled"],
        task_write_block_reason=task_write_state["block_reason"],
        task_write_block_message=task_write_state["block_message"],
        extra_fields=extra_fields,
        floors=floors,
        units=units,
        extra_field_write_enabled=extra_field_write_state["enabled"],
        extra_field_write_block_reason=extra_field_write_state["block_reason"],
        extra_field_write_block_message=extra_field_write_state["block_message"],
    )


bootstrap()

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=False)
